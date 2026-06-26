import logging
from datetime import datetime

import openpyxl  # Install with: pip install openpyxl (required for Excel export)

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, get_user_model, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.mail import EmailMessage
from django.db import connection
from django.db.models import Count, Q, Sum, Value
from django.db.models.functions import Coalesce, Concat
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from django.views.decorators.clickjacking import xframe_options_sameorigin

from .forms import DesktopLoginForm
from .models import (
    OeesAccessCards, OeesAccessCardsPins, OeesAccessCardsStates, OeesAccessCardsVisitors,
    OeesAccessCardsVisitorsNotes, OeesAccessKeys, OeesCompanies, OeesDelegations, OeesDevices,
    OeesDocs, OeesFiberLines, OeesFiberLinesIncidences, OeesIncorporations,
    OeesLicenses, OeesMobileLines, OeesMobilePhones, OeesOrders, OeesPrinters,
    OeesProvinces, OeesStaff, OeesUnderRepair,
)
from .reports import build_staff_inventory_pdf

logger = logging.getLogger(__name__)

PERMITS_LIST = [
    ("activo", "Active"), ("reader", "Reader"), ("users", "Users"),
    ("staff", "Staff"), ("devices", "Devices"), ("licenses", "Licenses"),
    ("phones", "Phones"), ("mobile_lines", "Mobile Lines"), ("fiber_lines", "Fiber Lines"),
    ("allocations", "Allocations"), ("incorporations", "Incorporations"),
    ("incorporator", "Incorporator"), ("orders", "Orders"), ("disponibility", "Disponibility"),
    ("delegation", "Delegation"), ("access_cards", "Access Cards"),
    ("visitors_cards", "Visitors Cards"), ("access_keys", "Access Keys"),
    ("under_repair", "Under Repair"), ("facturas", "Invoices distrib."),
    ("printers", "Printers"), ("not_returned", "Not Returned"), ("omada", "Omada"),
    ("net_overview", "Net Overview")
]

def api_get_device(request):
    serial = request.GET.get('serial_number', '').strip()
    
    if not serial:
        return JsonResponse({'success': False, 'error': 'Serial number not provided.'}, status=400)
    
    try:
        # Look up the device by serial number in isolation.
        device = OeesDevices.objects.get(serial_number=serial)

        # 1. Safely extract the company.
        company_val = ''
        if hasattr(device, 'company') and device.company:
            if hasattr(device.company, 'id_company'):
                company_val = device.company.id_company
            else:
                company_val = str(device.company)
        
        # 2. Safely extract the assigned employee.
        staff_val = 'Unassigned'
        try:
            if hasattr(device, 'persone') and device.persone and getattr(device, 'persone_id', None):
                staff_val = device.persone.name if hasattr(device.persone, 'name') else str(device.persone)
        except Exception:
            staff_val = 'Unassigned'
            
        # 3. Handle the mobile-line YES/NO flag.
        # It's a plain boolean: read it directly and safely, defaulting to False.
        is_mobile = False
        try:
            is_mobile = bool(getattr(device, 'mobile_line', False))
        except Exception:
            # In case the DB tries to resolve a broken OeesMobileLines relation in the background.
            is_mobile = False
            
        mobile_val = 'true' if is_mobile else 'false'
        
        # 4. Safe date formatting.
        date_val = ''
        if device.insert_date:
            try:
                date_val = device.insert_date.strftime('%Y-%m-%d')
            except AttributeError:
                date_val = str(device.insert_date)[:10]

        # 5. Safe numeric value formatting.
        value_val = '0'
        if device.value is not None:
            value_val = str(device.value).replace(',', '.')

        # Build the final JSON response.
        data = {
            'success': True,
            'serial': getattr(device, 'serial_number', '') or '',
            'type': getattr(device, 'type', '') or '',
            'brand': getattr(device, 'brand', '') or '',
            'model': getattr(device, 'model', '') or '',
            'company': company_val,
            'staff': staff_val,
            'screen': getattr(device, 'screen_size', '') or '',
            'hd': getattr(device, 'hd', '') or '',
            'memory': getattr(device, 'memory', '') or '',
            'imei': getattr(device, 'imei', '') or '',
            'mobile': mobile_val,
            'pin': getattr(device, 'pin_puk', '') or '',
            'origin': getattr(device, 'origin', '') or '',
            'date': date_val,
            'bill': getattr(device, 'bill_number', '') or '',
            'value': value_val,
            'obs': getattr(device, 'obs', '') or '',
            'notes': getattr(device, 'notes', '') or '',
            'under_repair': OeesUnderRepair.objects.filter(
                serial_number=serial, date_in__isnull=True).exists(),
        }
        return JsonResponse(data)
        
    except OeesDevices.DoesNotExist:
        return JsonResponse({'success': False, 'error': f'Device with SN {serial} not found.'})
    except Exception:
        # Log the real cause server-side; never leak internals to the client.
        logger.exception("Unexpected error fetching device with SN %s", serial)
        return JsonResponse({'success': False, 'error': 'An internal error occurred. Please try again later.'}, status=500)

def login_view(request):
    # If the user is already logged in, send them straight to the inventory.
    if request.user.is_authenticated:
        return redirect('mdi_home')

    if request.method == 'POST':
        form = DesktopLoginForm(request, data=request.POST)
        
        if form.is_valid():
            # Extract the cleaned username and password.
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')

            # Django checks the credentials against the MySQL database.
            user = authenticate(username=username, password=password)

            if user is not None:
                login(request, user)  # Create the session (stored in django_session).
                return redirect('mdi_home')  # Redirect to the main screen.
    else:
        form = DesktopLoginForm()

    return render(request, 'oe_inventory_py_web/login.html', {'form': form})


@login_required
def api_online_users(request):
    """Names of the users currently online, for the footer 'Online' popup."""
    from .context_processors import online_user_ids
    User = get_user_model()
    ids = online_user_ids()
    ids.add(str(request.user.pk))  # the requester is online by definition
    users = User.objects.filter(pk__in=ids)
    names = []
    for u in users:
        full = u.get_full_name().strip()
        names.append(full or u.username)
    names.sort(key=str.lower)
    return JsonResponse({'count': len(names), 'users': names})


# Language code -> manual file name (the files ship inside the app package so
# they are always deployed, unlike a top-level docs/ folder).
MANUAL_FILES = {'es': 'MANUAL_USUARIO.md', 'en': 'USER_MANUAL.md'}


@login_required
def manual_view(request):
    """Render the user manual (Markdown) as an in-app HTML page.

    Single source of truth: the same .md files used in the repo are rendered
    here. Use ?lang=en|es to switch language and #<anchor> to jump to a screen.
    """
    import os
    import markdown as md

    lang = request.GET.get('lang', 'es')
    if lang not in MANUAL_FILES:
        lang = 'es'

    manuals_dir = os.path.join(os.path.dirname(__file__), 'manuals')
    path = os.path.join(manuals_dir, MANUAL_FILES[lang])

    try:
        with open(path, encoding='utf-8') as fh:
            text = fh.read()
        html = md.markdown(text, extensions=['tables', 'sane_lists', 'toc', 'attr_list'])
        # Point the markdown image references at the served static files.
        static_base = '/' + settings.STATIC_URL.strip('/') + '/manual_images/'
        html = html.replace('src="images/', f'src="{static_base}')
    except FileNotFoundError:
        logger.exception("Manual file not found: %s", path)
        html = "<p>The manual is not available right now.</p>"

    return render(request, 'oe_inventory_py_web/manual.html', {
        'manual_html': html,
        'lang': lang,
    })


@login_required
def frm_devices_view(request):
    companies = OeesCompanies.objects.all()
    device_data = {}
    under_repair = False

    # 1. Handle POST actions.
    if request.method == 'POST':
        action = request.POST.get('action', '')
        serial_number = request.POST.get('serial_number', '').strip()

        if action == 'clear':
            return redirect('frm_devices')

        elif action == 'find':
            if serial_number:
                try:
                    device = OeesDevices.objects.get(serial_number=serial_number)
                    device_data = device
                    under_repair = OeesUnderRepair.objects.filter(
                        serial_number=serial_number,
                        date_in__isnull=True
                    ).exists()
                except OeesDevices.DoesNotExist:
                    messages.error(request, f"Device with SN '{serial_number}' not found.")
            else:
                messages.warning(request, "Please, insert a serial number to find it.")

        elif action == 'save':
            if not serial_number:
                messages.error(request, "Please, insert a serial number.")
            else:
                id_company = request.POST.get('id_company')
                company = None
                if id_company:
                    try:
                        company = OeesCompanies.objects.get(id_company=id_company)
                    except OeesCompanies.DoesNotExist:
                        company = None

                if company is None:
                    # The company column is NOT NULL in the database, so a
                    # company must always be chosen when saving a device.
                    messages.error(request, "Please select a company.")
                else:
                    has_mobile = request.POST.get('mobile_line') == '1'

                    device, created = OeesDevices.objects.get_or_create(
                        serial_number=serial_number,
                        defaults={
                            'insert_date': datetime.now().date(),
                            'value': 0.0,
                            'company': company,
                            # mobile_line is a plain 0/1 flag (has SIM?), NOT NULL.
                            'mobile_line': 0,
                        }
                    )

                    device.company = company
                    device.type = request.POST.get('type', '')
                    device.brand = request.POST.get('brand', '')
                    device.model = request.POST.get('model', '')
                    device.screen_size = request.POST.get('screen_size', '')
                    device.hd = request.POST.get('hd', '')
                    device.memory = request.POST.get('memory', '')
                    device.imei = request.POST.get('imei', '')
                    device.pin_puk = request.POST.get('pin_puk', '')
                    device.origin = request.POST.get('origin', '')
                    device.bill_number = request.POST.get('bill_number', '')
                    device.obs = request.POST.get('obs', '')

                    # Log this save to the device history (newest entry on top),
                    # the same way the technical-support action does.
                    now_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                    posted_notes = request.POST.get('notes', '')
                    action_txt = 'Created' if created else 'Updated'
                    device.notes = (f"{now_str} - {action_txt} by {request.user.username}\n"
                                    f"{posted_notes}").strip()

                    val_str = request.POST.get('value', '0')
                    try:
                        device.value = float(val_str)
                    except ValueError:
                        device.value = 0.0

                    device.mobile_line = 1 if has_mobile else 0   # 0/1 flag

                    date_str = request.POST.get('insert_date')
                    if date_str:
                        try:
                            device.insert_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                        except ValueError:
                            pass

                    device.save()
                    device_data = device
                    under_repair = OeesUnderRepair.objects.filter(
                        serial_number=serial_number,
                        date_in__isnull=True
                    ).exists()
                    messages.success(request, f"Device '{serial_number}' saved successfully.")

        elif action == 'support':
            if not serial_number:
                messages.error(request, "Please, insert a serial number to find it.")
            else:
                try:
                    device = OeesDevices.objects.get(serial_number=serial_number)
                    active_repair = OeesUnderRepair.objects.filter(
                        serial_number=serial_number,
                        date_in__isnull=True
                    ).first()

                    now_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

                    if active_repair:
                        # Receiving: record the repair cost entered by the user.
                        active_repair.date_in = datetime.now().date()
                        cost = None
                        repair_value = request.POST.get('repair_value', '').strip()
                        if repair_value:
                            try:
                                cost = float(repair_value.replace(',', '.'))
                                active_repair.value = cost
                            except ValueError:
                                pass
                        active_repair.save()
                        cost_txt = f" (cost {cost} €)" if cost is not None else ""
                        device.notes = (f"{now_str} - Received from maintenance{cost_txt} "
                                        f"by {request.user.username}\n{device.notes or ''}").strip()
                        # Update by primary key (unique, even if 0): a queryset update
                        # skips the AutoField check that save() applies to id_device=0.
                        OeesDevices.objects.filter(pk=device.pk).update(notes=device.notes)
                        messages.success(request, f"Device '{serial_number}' received from technical support.")
                        under_repair = False
                    else:
                        # Sending: store the destination entered by the user. type='D'
                        # marks it as a device so it shows up in frmUnderRepair.
                        destiny = request.POST.get('repair_destiny', '').strip() or "Technical Service"
                        OeesUnderRepair.objects.create(
                            serial_number=serial_number,
                            type='D',
                            date_out=datetime.now().date(),
                            destiny=destiny,
                            notes="Sent from the web inventory panel",
                            value=device.value or 0.0
                        )
                        device.notes = (f"{now_str} - Sent to maintenance ({destiny}) "
                                        f"by {request.user.username}\n{device.notes or ''}").strip()
                        # Update by primary key (unique, even if 0): a queryset update
                        # skips the AutoField check that save() applies to id_device=0.
                        OeesDevices.objects.filter(pk=device.pk).update(notes=device.notes)
                        messages.success(request, f"Device '{serial_number}' sent to technical support.")
                        under_repair = True
                    device_data = device
                except OeesDevices.DoesNotExist:
                    messages.error(request, f"Can't manage support: Device with SN '{serial_number}' not found.")

    # 2. Load the device list (QuerySet) applying GET filters.
    devices_qs = OeesDevices.objects.all().select_related('persone')
    staff_id = request.GET.get('id_staff', '')
    if staff_id and staff_id.isdigit():
        devices_qs = devices_qs.filter(persone__id_staff=staff_id)

    # 3. Export to Excel if requested via GET parameter.
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="devices_inventory.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Devices"
        
        headers = [
            'Serial Number', 'Type', 'Brand', 'Model', 'Screen Size', 'HD', 'Memory',
            'IMEI', 'Mobile Line', 'PIN/PUK', 'Origin', 'Insert Date', 'Bill Number',
            'Assigned To', 'Value (€)'
        ]
        ws.append(headers)
        
        for item in devices_qs:
            assigned = item.persone.name if item.persone else "Sin Asignar"
            mobile = "Yes" if item.mobile_line else "No"
            insert_date = item.insert_date.strftime('%Y-%m-%d') if item.insert_date else ""
            ws.append([
                item.serial_number, item.type, item.brand, item.model, item.screen_size,
                item.hd, item.memory, item.imei, mobile, item.pin_puk, item.origin,
                insert_date, item.bill_number, assigned, item.value
            ])
        
        wb.save(response)
        return response

    # 4. The grid itself is now loaded asynchronously by DataTables in
    #    server-side mode (see api_devices_datatable), so we don't materialise
    #    every device here. We only need the lightweight summary totals (one
    #    aggregate query) and the distinct types for the autocomplete datalist.
    agg = devices_qs.aggregate(total=Count('id_device'), value=Sum('value'))
    total_devices = agg['total'] or 0
    total_value = agg['value'] or 0.0
    unique_types = sorted(
        t for t in devices_qs.values_list('type', flat=True).distinct() if t
    )

    # Resolve the assigned staff name safely. Legacy rows may store an empty
    # string in `persone` instead of NULL, which would make the template's FK
    # lookup (device_data.persone) raise when converting '' to a number.
    device_staff = 'Unassigned'
    if device_data:
        try:
            if getattr(device_data, 'persone_id', None):
                device_staff = device_data.persone.name
        except Exception:
            device_staff = 'Unassigned'

    context = {
        'companies': companies,
        'total_devices': total_devices,
        'total_value': total_value,
        'device_data': device_data,
        'device_staff': device_staff,
        'unique_types': unique_types,
        'under_repair': under_repair,
        'staff_filter': staff_id,
    }

    return render(request, 'oe_inventory_py_web/frmDevices.html', context)


@login_required
def api_devices_datatable(request):
    """Server-side data source for the Devices grid (DataTables protocol).

    Returns one page of rows at a time (with search/order/paginate done in the
    database), so the screen no longer ships the entire inventory on load.
    """
    qs = OeesDevices.objects.select_related('persone')

    # Honour the optional "show only this staff's devices" filter.
    staff_id = request.GET.get('id_staff', '')
    if staff_id and staff_id.isdigit():
        qs = qs.filter(persone__id_staff=staff_id)

    records_total = qs.count()

    # Global search box.
    search = request.GET.get('search[value]', '').strip()
    if search:
        qs = qs.filter(
            Q(serial_number__icontains=search) |
            Q(type__icontains=search) |
            Q(brand__icontains=search) |
            Q(model__icontains=search) |
            Q(imei__icontains=search) |
            Q(origin__icontains=search) |
            Q(bill_number__icontains=search) |
            Q(persone__name__icontains=search)
        )

    records_filtered = qs.count()

    # Ordering: column index -> model field, matching the table's columns.
    order_columns = [
        'serial_number', 'type', 'brand', 'model', 'screen_size', 'hd',
        'memory', 'imei', 'mobile_line', 'pin_puk', 'origin', 'insert_date',
        'bill_number', 'persone__name', 'value',
    ]
    order_col = request.GET.get('order[0][column]', '')
    order_dir = request.GET.get('order[0][dir]', 'asc')
    if order_col.isdigit() and 0 <= int(order_col) < len(order_columns):
        field = order_columns[int(order_col)]
        qs = qs.order_by(('-' if order_dir == 'desc' else '') + field)
    else:
        qs = qs.order_by('serial_number')

    # Pagination.
    try:
        start = int(request.GET.get('start', 0))
    except ValueError:
        start = 0
    try:
        length = int(request.GET.get('length', 50))
    except ValueError:
        length = 50
    page = qs[start:] if length == -1 else qs[start:start + length]

    def _safe_date(value):
        # Legacy rows can carry odd dates (e.g. '0000-00-00'); never let one
        # bad value turn an ordering request into a 500.
        try:
            return value.strftime('%d-%m-%Y') if value else ''
        except Exception:
            return ''

    def _row(d):
        return {
            'serial': d.serial_number,
            'type': d.type or '',
            'brand': d.brand or '',
            'model': d.model or '',
            'screen': d.screen_size or '-',
            'hd': d.hd or '-',
            'memory': d.memory or '-',
            'imei': d.imei or '-',
            'mobile': bool(d.mobile_line),
            'pin': d.pin_puk or '-',
            'origin': d.origin or '-',
            'date': _safe_date(d.insert_date),
            'bill': d.bill_number or '-',
            'staff': d.persone.name if d.persone else '-',
            'value': d.value or 0,
        }

    # Build the rows defensively: a single corrupt record must not break the
    # whole grid (which DataTables would surface as a generic "Ajax error").
    data = []
    try:
        for d in page:
            try:
                data.append(_row(d))
            except Exception:
                logger.exception("Skipping a device row that failed to serialize")
    except Exception:
        logger.exception("Failed to iterate the devices page")

    return JsonResponse({
        'draw': int(request.GET.get('draw', 1)),
        'recordsTotal': records_total,
        'recordsFiltered': records_filtered,
        'data': data,
    })

def api_finder(request):
    """
    Generic finder that returns JSON.
    Expected GET params: 'term' (what the user types), 'field' (column to search).
    """
    term = request.GET.get('term', '').strip()
    field = request.GET.get('field', 'device_name')
    option = request.GET.get('option', None)

    # User finder: search CustomUser by name (first_name) or login (username).
    if option == 'users':
        User = get_user_model()
        usr_qs = (User.objects.filter(Q(first_name__icontains=term) | Q(username__icontains=term))
                  .order_by('first_name').values('username', 'first_name')[:50])
        results = [{'code': u['username'], 'description': u['first_name'] or ''} for u in usr_qs]
        return JsonResponse({'status': 'success', 'data': results})

    # Staff finder: search staff by name within the user's scope.
    if option == 'staff':
        staff_qs = scope_staff_queryset(
            request.user, OeesStaff.objects.all()
        ).filter(name__icontains=term).order_by('name')[:50]
        results = [{'code': s.id_staff, 'description': s.name} for s in staff_qs]
        return JsonResponse({'status': 'success', 'data': results})

    # License finder: search licenses by serial number.
    if option == 'licenses':
        lic_qs = (OeesLicenses.objects.filter(serial_number__icontains=term)
                  .order_by('serial_number').values('serial_number', 'type')[:50])
        results = [{'code': lic['serial_number'], 'description': lic['type']} for lic in lic_qs]
        return JsonResponse({'status': 'success', 'data': results})

    # Phone finder: search mobile phones by serial number.
    if option == 'phones':
        ph_qs = (OeesMobilePhones.objects.filter(serial_number__icontains=term)
                 .exclude(serial_number='Personal Mobile')
                 .order_by('serial_number').values('serial_number', 'brand')[:50])
        results = [{'code': p['serial_number'], 'description': p['brand']} for p in ph_qs]
        return JsonResponse({'status': 'success', 'data': results})

    # Fiber line finder: search by description.
    if option == 'fiber':
        fb_qs = (OeesFiberLines.objects.filter(description__icontains=term)
                 .order_by('description').values('id_fiber_line', 'description')[:50])
        results = [{'code': f['id_fiber_line'], 'description': f['description']} for f in fb_qs]
        return JsonResponse({'status': 'success', 'data': results})

    # Printer finder: search by description (returns the serial number as the key).
    if option == 'printers':
        pr_qs = (OeesPrinters.objects.filter(description__icontains=term)
                 .order_by('description').values('serial_number', 'description')[:50])
        results = [{'code': p['serial_number'], 'description': p['description']} for p in pr_qs]
        return JsonResponse({'status': 'success', 'data': results})

    # Mobile line finder: search by number.
    if option == 'lines':
        ln_qs = (OeesMobileLines.objects.filter(number__icontains=term)
                 .order_by('number').values('number', 'origin')[:50])
        results = [{'code': ln['number'], 'description': ln['origin'] or ''} for ln in ln_qs]
        return JsonResponse({'status': 'success', 'data': results})

    # Delegation finder: search by name.
    if option == 'delegations':
        dl_qs = (OeesDelegations.objects.filter(delegation__icontains=term)
                 .order_by('delegation').values('id_delegation', 'delegation')[:50])
        results = [{'code': d['id_delegation'], 'description': d['delegation']} for d in dl_qs]
        return JsonResponse({'status': 'success', 'data': results})

    # Access card finder: search by card code (ac_max).
    if option == 'access_cards':
        ac_qs = (OeesAccessCards.objects.filter(ac_max__icontains=term)
                 .order_by('ac_max').values('ac_max')[:50])
        results = [{'code': c['ac_max'], 'description': c['ac_max']} for c in ac_qs]
        return JsonResponse({'status': 'success', 'data': results})

    # Visitor card finder: search by card code (ac_max).
    if option == 'visitor_cards':
        vc_qs = (OeesAccessCardsVisitors.objects.filter(ac_max__icontains=term)
                 .order_by('ac_max').values('ac_max', 'name')[:50])
        results = [{'code': v['ac_max'], 'description': v['name'] or ''} for v in vc_qs]
        return JsonResponse({'status': 'success', 'data': results})

    # Access key finder: search by type.
    if option == 'access_keys':
        ak_qs = (OeesAccessKeys.objects.filter(type__icontains=term)
                 .order_by('id_key').values('id_key', 'type')[:50])
        results = [{'code': k['id_key'], 'description': k['type'] or ''} for k in ak_qs]
        return JsonResponse({'status': 'success', 'data': results})

    # Build the filter dynamically (equivalent to a SQL 'like').
    lookup = f"{field}__icontains"
    filters = Q(**{lookup: term})

    # Replicate the special logic for option 5 (sPhones example).
    if option == "5":
        filters &= ~Q(serial_number__icontains='Personal Mobile')

    # Run the query (assumes it returns code/serial and description),
    # ordered dynamically by the search field. Limit to 50 for performance.
    queryset = OeesDevices.objects.filter(filters).order_by(field)[:50]

    # Format the result mimicking the "Code" and "Description" columns.
    results = []
    for item in queryset:
        results.append({
            'code': item.serial_number,  # The value to inject.
            'description': getattr(item, field)  # The description the user sees.
        })
        
    return JsonResponse({'status': 'success', 'data': results})

def dict_fetchall(cursor):
    """Return all rows from a cursor as a list of dicts."""
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]

@login_required
def frm_users(request):
    user_session = request.user.username

    if request.method == "POST":
        login = request.POST.get('login', '').strip()
        nombre = request.POST.get('nombre', '').strip()
        is_new_raw = request.POST.get('is_new', '0')
        is_new = is_new_raw == '1'

        # 1. Collect multi-select grids (outside any loop).
        list_companies = request.POST.getlist('companies_selected')  
        list_delegations = request.POST.getlist('delegations_selected')
        list_departments = request.POST.getlist('departments_selected')
        
        sCompanies = ",".join(list_companies) if list_companies else None     
        sDelegations = ",".join(list_delegations) if list_delegations else None
        sDepartments = ",".join(list_departments) if list_departments else None

        hoy = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if login and nombre:
            try:
                User = get_user_model()

                # An initial password may only be set by users holding the Users
                # permit; '********' is just the on-screen mask, not a real value.
                raw_pw = request.POST.get('password', '').strip()
                can_set_pw = getattr(request.user, 'users', 0) == 1
                pw_to_set = raw_pw if (can_set_pw and raw_pw and raw_pw != '********') else None

                # Get or create the user instance.
                if is_new:
                    if User.objects.filter(username=login).exists():
                        messages.error(request, f"Login {login} already exists.")
                        return redirect('frm_users')

                    objeto_usuario = User.objects.create_user(username=login, password=pw_to_set)
                    objeto_usuario.notes = f"{hoy} - Created by {user_session}"
                else:
                    if User.objects.filter(username=login).exists():
                        objeto_usuario = User.objects.get(username=login)

                        # If 'notes' came in empty from the POST, keep whatever was
                        # already stored so we don't wipe the audit history.
                        old_notes_form = request.POST.get('notes', '').strip()
                        old_notes_db = objeto_usuario.notes if objeto_usuario.notes else ""
                        final_old_notes = old_notes_form if old_notes_form else old_notes_db

                        # Prepend the new history block.
                        objeto_usuario.notes = f"{hoy} - Modified by {user_session}\n{final_old_notes}".strip()

                        # Set an INITIAL password only when the user has none yet;
                        # existing passwords are changed via 'Password Change' or
                        # the 'Forgot my password' email flow.
                        if pw_to_set and not objeto_usuario.has_usable_password():
                            objeto_usuario.set_password(pw_to_set)
                    else:
                        messages.error(request, "User to update does not exist.")
                        return redirect('frm_users')

                objeto_usuario.first_name = nombre
                objeto_usuario.email = request.POST.get('email', '').strip()
                objeto_usuario.companies = sCompanies
                objeto_usuario.delegations = sDelegations
                objeto_usuario.departments = sDepartments

                # Apply permission checkboxes.
                for campo, _ in PERMITS_LIST:
                    nombre_input = f"permiso_{campo}"
                    valor_permiso = 1 if nombre_input in request.POST else 0
                    if hasattr(objeto_usuario, campo):
                        setattr(objeto_usuario, campo, valor_permiso)

                objeto_usuario.save()

                messages.success(request, "User updated successfully.")
                return redirect('frm_users')

            except Exception:
                logger.exception("Error saving CustomUser %s", login)
                messages.error(request, "An error occurred while saving the user.")
                return redirect('frm_users')

    # Load initial data to populate the grids (GET).
    with connection.cursor() as cursor:
        cursor.execute("SELECT id_delegation, delegation FROM oees_delegations ORDER BY id_delegation")
        delegations = dict_fetchall(cursor)
        
        cursor.execute("SELECT id_company, name FROM oees_companies ORDER BY id_company")
        companies = dict_fetchall(cursor)
        
        cursor.execute("SELECT DISTINCT department FROM oees_staff WHERE (department IS NOT NULL AND department <> '') ORDER BY department ASC")
        departments = dict_fetchall(cursor)

    context = {
        'delegations': delegations,
        'companies': companies,
        'departments': departments,
        'permits': PERMITS_LIST,
    }
    return render(request, 'oe_inventory_py_web/frmUser.html', context)

def api_get_user(request):
    """AJAX API: look up CustomUser and map the fields for the JS."""
    login_buscado = request.GET.get('login', '').strip().lower()
    
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM oe_inventory_py_web_customuser WHERE LOWER(TRIM(username)) = %s", [login_buscado])
        row = cursor.fetchone()
        
        if row:
            columns = [col[0] for col in cursor.description]
            user_data = dict(zip(columns, row))

            # Whether the user already has a usable password (Django stores an
            # unusable password as a hash starting with '!'). Used by the form to
            # decide if the admin may set an INITIAL password.
            raw_hash = user_data.get('password') or ''
            has_password = bool(raw_hash) and not raw_hash.startswith('!')

            # Never expose sensitive auth columns to the browser.
            for sensitive in ('password', 'last_login', 'is_superuser', 'is_staff', 'date_joined'):
                user_data.pop(sensitive, None)
            # The password field is display-only on the form; send a mask, never the hash.
            user_data['password'] = '********'
            user_data['has_password'] = has_password

            # Map 'username' to 'login' for the login input.
            user_data['login'] = user_data.get('username')

            # Map 'first_name' to 'nombre' so the JS can paint it on screen.
            user_data['nombre'] = user_data.get('first_name', '')

            return JsonResponse({'success': True, 'exists': True, 'data': user_data})
        else:
            return JsonResponse({'success': True, 'exists': False, 'login': login_buscado})


@login_required
def mdi_home_view(request):
    return render(request, 'oe_inventory_py_web/base_mdi.html')


# ==========================================================================
# Staff screen (web port of frmStaff.py)
# ==========================================================================

def _csv_values(raw):
    """Split a comma-separated permission string into a clean list."""
    return [v.strip() for v in (raw or '').split(',') if v.strip()]


def scope_staff_queryset(user, qs):
    """Restrict a staff queryset to the user's allowed companies/delegations/departments.

    The three scopes are optional and applied as successive AND filters, each
    one narrowing the result only when it is set:
      - No companies assigned   -> all companies; otherwise only the assigned ones.
      - If delegations are set   -> further narrows the result (on top of companies).
      - If departments are set    -> narrows it further still.
    A user with no scope at all sees everyone.
    """
    companies = _csv_values(getattr(user, 'companies', '') or '')
    delegations = _csv_values(getattr(user, 'delegations', '') or '')
    departments = _csv_values(getattr(user, 'departments', '') or '')
    if companies:
        qs = qs.filter(company_id__in=companies)
    if delegations:
        qs = qs.filter(delegation_id__in=delegations)
    if departments:
        qs = qs.filter(department__in=departments)
    return qs


def _is_reader(user):
    return getattr(user, 'reader', 0) == 1


def _fmt_date(value):
    try:
        return value.strftime('%Y-%m-%d')
    except AttributeError:
        return str(value)[:10] if value else ''


def staff_assigned_items(staff_id):
    """Return (items, total_value): everything assigned to a staff member
    (devices, licenses, mobile phones, office access cards and access keys).

    Only the columns that exist in the legacy tables are selected (via .values()),
    because some auto-generated models declare columns the real tables don't have.
    """
    items = []
    total_value = 0.0

    devices = (OeesDevices.objects.filter(persone_id=staff_id)
               .values('id_device', 'serial_number', 'type', 'brand', 'model',
                       'origin', 'insert_date', 'obs', 'value')
               .order_by('-id_device'))
    for d in devices:
        value = float(d['value'] or 0)
        total_value += value
        items.append({'id': f"D{d['id_device']}", 'serial': d['serial_number'] or '', 'type': d['type'] or '',
                      'brand': d['brand'] or '', 'model': d['model'] or '', 'origin': d['origin'] or '',
                      'date': _fmt_date(d['insert_date']), 'obs': d['obs'] or '', 'value': value})

    licenses = (OeesLicenses.objects.filter(persone_id=staff_id)
                .values('id_license', 'serial_number', 'type', 'origin', 'insert_date', 'value')
                .order_by('-id_license'))
    for lic in licenses:
        value = float(lic['value'] or 0)
        total_value += value
        items.append({'id': f"L{lic['id_license']}", 'serial': lic['serial_number'] or '', 'type': lic['type'] or '',
                      'brand': '', 'model': '', 'origin': lic['origin'] or '',
                      'date': _fmt_date(lic['insert_date']), 'obs': '', 'value': value})

    phones = (OeesMobilePhones.objects.filter(persone_id=staff_id)
              .values('id_mobile_phone', 'serial_number', 'type', 'origin', 'brand',
                      'model', 'insert_date', 'obs', 'value')
              .order_by('-id_mobile_phone'))
    for ph in phones:
        value = float(ph['value'] or 0)
        total_value += value
        items.append({'id': f"P{ph['id_mobile_phone']}", 'serial': ph['serial_number'] or '', 'type': ph['type'] or '',
                      'brand': ph['brand'] or '', 'model': ph['model'] or '', 'origin': ph['origin'] or '',
                      'date': _fmt_date(ph['insert_date']), 'obs': ph['obs'] or '', 'value': value})

    cards = (OeesAccessCards.objects.filter(id_staff_id=staff_id)
             .exclude(state_card=4).values('id_card'))
    for card in cards:
        items.append({'id': f"T{card['id_card']}", 'serial': '', 'type': 'OFFICE ACCESS CARD',
                      'brand': '', 'model': '', 'origin': '', 'date': '', 'obs': '', 'value': 0.0})

    keys = OeesAccessKeys.objects.filter(id_staff_id=staff_id).values('id_key', 'type')
    for key in keys:
        items.append({'id': f"K{key['id_key']}", 'serial': '', 'type': key['type'] or 'ACCESS KEY',
                      'brand': '', 'model': '', 'origin': '', 'date': '', 'obs': '', 'value': 0.0})

    return items, total_value


@login_required
def frm_staff_view(request):
    user = request.user

    if request.method == 'POST':
        if _is_reader(user):
            messages.error(request, "You have a reader profile and can't modify data.")
            return redirect('frm_staff')
        action = request.POST.get('action', '')
        handlers = {
            'save': _staff_save,
            'terminate': _staff_terminate,
            'release': _staff_release,
            'upload_doc': _staff_upload_doc,
        }
        handler = handlers.get(action)
        if handler:
            return handler(request)
        return redirect('frm_staff')

    # ---- GET ----
    staff_id = request.GET.get('staff', '').strip()
    if request.GET.get('export') == 'excel' and staff_id.isdigit():
        return _staff_export_excel(staff_id)

    companies = OeesCompanies.objects.all().order_by('name')
    delegations = OeesDelegations.objects.all().order_by('delegation')
    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT department FROM oees_staff "
                       "WHERE department IS NOT NULL AND department <> '' ORDER BY department ASC")
        departments = [row[0] for row in cursor.fetchall()]

    scoped = scope_staff_queryset(user, OeesStaff.objects.select_related('company', 'delegation'))

    context = {
        'companies': companies,
        'delegations': delegations,
        'departments': departments,
        'staff_list': scoped.order_by('-id_staff'),
        'people_count': scoped.filter(persona_fisica=1).count(),
        'active_count': scoped.filter(state=1).count(),
        'preselected_staff_id': staff_id,
        'is_reader': _is_reader(user),
    }
    return render(request, 'oe_inventory_py_web/frmStaff.html', context)


def _staff_save(request):
    user = request.user
    code = request.POST.get('code', '').strip()
    name = request.POST.get('name', '').strip()
    if not name:
        messages.error(request, "Name is required.")
        return redirect('frm_staff')

    department = request.POST.get('department', '').strip()
    company_id = request.POST.get('company') or None
    delegation_id = request.POST.get('delegation') or None
    email = request.POST.get('email', '').strip()
    f_inc = request.POST.get('fecha_incorporacion', '').strip()
    person = 1 if request.POST.get('persona_fisica') else 0
    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    try:
        if not code:
            OeesStaff.objects.create(
                name=name, department=department, company_id=company_id,
                delegation_id=delegation_id, email=email, persona_fisica=person,
                fecha_incorporacion=f_inc, state=1, obs='',
                notes=f"{datetime.now()} - Created by {user.username}",
            )
        else:
            staff = OeesStaff.objects.filter(id_staff=code).first()
            if not staff:
                messages.error(request, "Staff to update does not exist.")
                return redirect('frm_staff')
            existing = request.POST.get('notes', '').strip() or (staff.notes or '')
            staff.name = name
            staff.department = department
            staff.company_id = company_id
            staff.delegation_id = delegation_id
            staff.email = email
            staff.persona_fisica = person
            staff.fecha_incorporacion = f_inc
            staff.notes = f"{now} - Modified by {user.username}\n{existing}".strip()
            staff.save()
        messages.success(request, "Staff saved successfully.")
    except Exception:
        logger.exception("Error saving staff (code=%s)", code)
        messages.error(request, "An error occurred while saving the staff member.")
    return redirect('frm_staff')


def _staff_terminate(request):
    code = request.POST.get('code', '').strip()
    staff = OeesStaff.objects.filter(id_staff=code).first()
    if not staff:
        messages.error(request, "Staff member not found.")
        return redirect('frm_staff')

    # Snapshot of everything the person had BEFORE freeing anything (the PDF lists
    # it all, marking the returned ones).
    items, _total = staff_assigned_items(staff.id_staff)
    returned_ids = {x for x in request.POST.get('returned_items', '').split(',') if x}
    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    # 1. Free each returned item in its own table (back to stock).
    freed = 0
    for it in items:
        it['returned'] = it['id'] in returned_ids
        if it['returned']:
            try:
                _release_assigned_item(
                    it['id'], it['serial'],
                    f"{datetime.now().strftime('%d/%m/%Y')} - Returned on termination of {staff.name} by {request.user.username}",
                )
                freed += 1
            except Exception:
                logger.exception("Error releasing item %s on termination", it['id'])

    # 2. Generate Terminate.pdf (same layout + RETURNED column) and store it as a
    #    document of this staff member, so it can be printed and signed.
    try:
        staff_data = {
            'name': staff.name or '',
            'dep': staff.department or '',
            'comp': staff.company.name if staff.company else '',
            'deleg': staff.delegation.delegation if staff.delegation else '',
            'fecha_i': staff.fecha_incorporacion or '',
        }
        pdf = build_staff_inventory_pdf(
            staff_data, items,
            generated_by=(request.user.get_full_name() or request.user.username),
            show_returned=True,
        )
        key = f"staff_docs/{code}/Terminate.pdf"
        if default_storage.exists(key):
            default_storage.delete(key)
        default_storage.save(key, ContentFile(pdf))
        OeesDocs.objects.get_or_create(
            id_staff_id=code, doc_name='Terminate',
            defaults={'notes': f"{datetime.now()} - Termination document by {request.user.username}"},
        )
    except Exception:
        logger.exception("Error generating Terminate.pdf for staff %s", code)

    # 3. Mark the contract as terminated.
    existing = request.POST.get('notes', '').strip() or (staff.notes or '')
    staff.notes = (f"{now} - Contract terminated by {request.user.username} "
                   f"({freed} item(s) returned)\n{existing}").strip()
    staff.fecha_baja = datetime.now().strftime('%Y-%m-%d')
    staff.state = 0
    staff.save()
    messages.success(
        request,
        f"Contract terminated. {freed} item(s) returned to stock. 'Terminate' document generated.",
    )
    return redirect(f"{reverse('frm_staff')}?staff={code}")


def _release_assigned_item(item_id, serial, note_text):
    """Free one assigned item (back to stock) in its own table.

    item_id is the prefixed code from staff_assigned_items:
    D=device, L=license, P=phone (freed by serial); T=access card, K=key (by id).
    """
    prefix, real_id = item_id[0], item_id[1:]
    prepend_note = Concat(Value(note_text + "\n"), Coalesce('notes', Value('')))
    if prefix == 'D':
        OeesDevices.objects.filter(serial_number=serial).update(persone=None, notes=prepend_note)
    elif prefix == 'L':
        OeesLicenses.objects.filter(serial_number=serial).update(persone=None, notes=prepend_note)
    elif prefix == 'P':
        OeesMobilePhones.objects.filter(serial_number=serial).update(persone=None, notes=prepend_note)
    elif prefix == 'T':
        OeesAccessCards.objects.filter(id_card=real_id).update(id_staff=None, notes=prepend_note)
    elif prefix == 'K':
        OeesAccessKeys.objects.filter(id_key=real_id).update(id_staff=None, notes=prepend_note)


def _staff_release(request):
    code = request.POST.get('code', '').strip()
    item_id = request.POST.get('item_id', '').strip()
    serial = request.POST.get('serial', '').strip()
    name = request.POST.get('name', '').strip()
    if not item_id:
        return redirect('frm_staff')

    note = f"{datetime.now().strftime('%d/%m/%Y')} - Unassigned from {name} by {request.user.username}"
    try:
        _release_assigned_item(item_id, serial, note)
        messages.success(request, "Item unassigned successfully.")
        # Document the remaining inventory for physical persons.
        staff = OeesStaff.objects.filter(id_staff=code).first()
        if staff and staff.persona_fisica:
            _save_staff_inventory_doc(staff, request.user, 'Unassign')
    except Exception:
        logger.exception("Error releasing item %s", item_id)
        messages.error(request, "An error occurred while unassigning the item.")
    return redirect(f"{reverse('frm_staff')}?staff={code}")


def _sanitize_filename(name):
    """Keep only safe characters for an on-disk file name (no path traversal)."""
    cleaned = ''.join(c for c in name if c.isalnum() or c in (' ', '-', '_')).strip()
    return cleaned or 'document'


def _staff_upload_doc(request):
    code = request.POST.get('code', '').strip()
    doc_name = request.POST.get('doc_name', '').strip()
    upload = request.FILES.get('doc_file')
    if not code.isdigit():
        messages.error(request, "Load a staff member before adding documents.")
        return redirect('frm_staff')
    if not upload or not doc_name:
        messages.error(request, "Select a file and provide a document name.")
        return redirect(f"{reverse('frm_staff')}?staff={code}")
    if not (upload.name.lower().endswith('.pdf') or upload.content_type == 'application/pdf'):
        messages.error(request, "Only PDF documents are accepted.")
        return redirect(f"{reverse('frm_staff')}?staff={code}")

    safe_name = _sanitize_filename(doc_name)
    try:
        # Save through Django's storage backend so it works both on the local
        # filesystem (dev) and on S3 (production). Overwrite a same-named doc.
        key = f"staff_docs/{code}/{safe_name}.pdf"
        if default_storage.exists(key):
            default_storage.delete(key)
        default_storage.save(key, upload)
        OeesDocs.objects.create(
            id_staff_id=code, doc_name=safe_name,
            notes=f"{datetime.now()} - Created by {request.user.username}",
        )
        messages.success(request, "Document saved successfully.")
    except Exception:
        logger.exception("Error saving staff doc for %s", code)
        messages.error(request, "An error occurred while saving the document.")
    return redirect(f"{reverse('frm_staff')}?staff={code}")


def _staff_export_excel(staff_id):
    items, _ = staff_assigned_items(staff_id)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="staff_{staff_id}_inventory.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assigned items"
    ws.append(['ID', 'Serial', 'Type', 'Brand', 'Model', 'Origin', 'Date', 'Obs', 'Value'])
    for it in items:
        ws.append([it['id'], it['serial'], it['type'], it['brand'], it['model'],
                   it['origin'], it['date'], it['obs'], it['value']])
    wb.save(response)
    return response


@login_required
def api_get_staff(request):
    """AJAX API: look up a staff member (within the user's scope) plus assigned items and docs."""
    staff_id = request.GET.get('id', '').strip()
    if not staff_id:
        return JsonResponse({'success': False, 'error': 'No staff id provided.'}, status=400)

    staff = scope_staff_queryset(
        request.user, OeesStaff.objects.select_related('company', 'delegation')
    ).filter(id_staff=staff_id).first()
    if not staff:
        return JsonResponse({'success': True, 'exists': False, 'id': staff_id})

    items, total_value = staff_assigned_items(staff.id_staff)
    docs = list(
        OeesDocs.objects.filter(id_staff_id=staff.id_staff)
        .order_by('-id_doc').values_list('doc_name', flat=True)
    )
    data = {
        'id': staff.id_staff, 'name': staff.name or '', 'department': staff.department or '',
        'company_id': staff.company_id or '', 'delegation_id': staff.delegation_id or '',
        'email': staff.email or '', 'fecha_incorporacion': staff.fecha_incorporacion or '',
        'fecha_baja': staff.fecha_baja or '', 'persona_fisica': 1 if staff.persona_fisica else 0,
        'notes': staff.notes or '', 'items': items, 'total_value': total_value,
        'devices_count': len(items), 'docs': docs,
    }
    return JsonResponse({'success': True, 'exists': True, 'data': data})


def _staff_report_payload(user, staff_id):
    """Build (staff, staff_data, items) for the PDF report, honouring the user's scope."""
    staff = scope_staff_queryset(
        user, OeesStaff.objects.select_related('company', 'delegation')
    ).filter(id_staff=staff_id).first()
    if not staff:
        return None, None, None
    items, _ = staff_assigned_items(staff.id_staff)
    staff_data = {
        'name': staff.name or '',
        'dep': staff.department or '',
        'comp': staff.company.name if staff.company else '',
        'deleg': staff.delegation.delegation if staff.delegation else '',
        'fecha_i': staff.fecha_incorporacion or '',
    }
    return staff, staff_data, items


@login_required
def staff_report(request):
    """Generate and display (inline) the staff inventory PDF."""
    staff, staff_data, items = _staff_report_payload(request.user, request.GET.get('staff', '').strip())
    if not staff:
        raise Http404("Staff member not found.")
    pdf = build_staff_inventory_pdf(
        staff_data, items, generated_by=(request.user.get_full_name() or request.user.username)
    )
    filename = f"Inventory_{(staff.name or 'staff').replace(' ', '_')}.pdf"
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
def staff_email_report(request):
    """Generate the staff inventory PDF and email it to the People department."""
    if request.method != 'POST':
        return redirect('frm_staff')
    staff_id = request.POST.get('staff', '').strip()
    staff, staff_data, items = _staff_report_payload(request.user, staff_id)
    if not staff:
        messages.error(request, "Staff member not found.")
        return redirect('frm_staff')

    recipient = getattr(settings, 'STAFF_REPORT_RECIPIENT', '')
    if not recipient or not getattr(settings, 'RESEND_API_KEY', ''):
        messages.error(request, "Email is not configured. Set RESEND_API_KEY and STAFF_REPORT_RECIPIENT in .env.")
        return redirect(f"{reverse('frm_staff')}?staff={staff_id}")

    try:
        pdf = build_staff_inventory_pdf(
            staff_data, items, generated_by=(request.user.get_full_name() or request.user.username)
        )
        email = EmailMessage(
            subject=f"Personal Inventory - {staff.name}",
            body=(f"Hello,\n\nPlease find attached the personal inventory report for {staff.name} "
                  "generated by the IT team.\n\nBest regards,\nIT Team."),
            to=[recipient],
        )
        email.attach(f"Inventory_{(staff.name or 'staff').replace(' ', '_')}.pdf", pdf, 'application/pdf')
        email.send()

        now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        staff.notes = f"{now} - Inventory sent to People Dept. by {request.user.username}\n{staff.notes or ''}".strip()
        staff.save()
        messages.success(request, "Email sent successfully.")
    except Exception:
        logger.exception("Error sending staff inventory email for %s", staff_id)
        messages.error(request, "Error sending the email. Check the email configuration.")
    return redirect(f"{reverse('frm_staff')}?staff={staff_id}")


@login_required
@xframe_options_sameorigin
def staff_doc(request, staff_id, doc_name):
    """Serve a previously uploaded staff document (PDF).

    Exempted from the global X-Frame-Options: DENY so the PDF can be shown in
    the same-origin <iframe> preview on the Staff screen.
    """
    # Served through Django's storage backend (local filesystem or S3). The
    # name is sanitized so the storage key matches what was saved and to avoid
    # path traversal; staff_id is an int from the URL converter.
    key = f"staff_docs/{staff_id}/{_sanitize_filename(doc_name)}.pdf"
    if not default_storage.exists(key):
        raise Http404("Document not found.")
    return FileResponse(default_storage.open(key, 'rb'), content_type='application/pdf')


# ==========================================================================
# Licenses screen (web port of frmLicenses.py)
# ==========================================================================

def _licenses_export_excel():
    grid = OeesLicenses.objects.select_related('company', 'persone').order_by('-id_license')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="licenses_inventory.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Licenses"
    ws.append(['Serial Number', 'Company', 'Type', 'Origin', 'Insert Date', 'Person', 'Obs', 'Value', 'Bill Number'])
    for lic in grid:
        ws.append([
            lic.serial_number, lic.company.name if lic.company else '', lic.type, lic.origin or '',
            lic.insert_date.strftime('%Y-%m-%d') if lic.insert_date else '',
            lic.persone.name if lic.persone else '', lic.obs or '', lic.value,
            lic.bill_number or '',
        ])
    wb.save(response)
    return response


@login_required
def frm_licenses_view(request):
    user = request.user

    if request.method == 'POST' and request.POST.get('action') == 'save':
        if _is_reader(user):
            messages.error(request, "You have a reader profile and can't modify data.")
            return redirect('frm_licenses')

        code = request.POST.get('serial_number', '').strip()
        if not code:
            messages.error(request, "Serial Number is required.")
            return redirect('frm_licenses')

        company_id = request.POST.get('company') or None
        ltype = request.POST.get('type', '').strip()
        origin = request.POST.get('origin', '').strip()
        obs = request.POST.get('obs', '').strip()
        bill = request.POST.get('bill_number', '').strip()
        try:
            value = float(request.POST.get('value', '0'))
        except ValueError:
            value = 0.0
        insert_date = None
        date_str = request.POST.get('insert_date', '').strip()
        if date_str:
            try:
                insert_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                insert_date = None
        now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

        try:
            lic, created = OeesLicenses.objects.get_or_create(
                serial_number=code, defaults={'type': ltype, 'value': value, 'obs': obs},
            )
            lic.company_id = company_id
            lic.type = ltype
            lic.origin = origin
            lic.obs = obs
            lic.bill_number = bill
            lic.value = value
            if insert_date:
                lic.insert_date = insert_date
            if created:
                lic.notes = f"{datetime.now()} - Created by {user.username}"
            else:
                existing = request.POST.get('notes', '').strip() or (lic.notes or '')
                lic.notes = f"{now} - Modified by {user.username}\n{existing}".strip()
            lic.save()
            messages.success(request, "License saved successfully.")
        except Exception:
            logger.exception("Error saving license %s", code)
            messages.error(request, "An error occurred while saving the license.")
        return redirect('frm_licenses')

    # ---- GET ----
    if request.GET.get('export') == 'excel':
        return _licenses_export_excel()

    companies = OeesCompanies.objects.all().order_by('name')
    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT type FROM oees_licenses "
                       "WHERE type IS NOT NULL AND type <> '' ORDER BY type ASC")
        type_options = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT origin FROM oees_licenses "
                       "WHERE origin IS NOT NULL AND origin <> '' ORDER BY origin ASC")
        origin_options = [row[0] for row in cursor.fetchall()]

    grid = OeesLicenses.objects.select_related('company', 'persone').order_by('-id_license')

    # Summary by type: Purchased = licenses of that type; Expired = those assigned
    # to the person "LICENCIAS CADUCADAS"; In Use = Purchased - Expired.
    summary_rows = (
        OeesLicenses.objects.values('type')
        .annotate(
            purchased=Count('id_license'),
            expired=Count('id_license', filter=Q(persone__name='LICENCIAS CADUCADAS')),
        )
        .order_by('type')
    )
    license_summary = [{
        'type': r['type'] or '',
        'purchased': r['purchased'] or 0,
        'expired': r['expired'] or 0,
        'in_use': (r['purchased'] or 0) - (r['expired'] or 0),
    } for r in summary_rows]

    context = {
        'companies': companies,
        'type_options': type_options,
        'origin_options': origin_options,
        'grid_data': grid,
        'license_summary': license_summary,
        'total_licenses': grid.count(),
        'total_value': grid.aggregate(total=Sum('value'))['total'] or 0,
        'is_reader': _is_reader(user),
    }
    return render(request, 'oe_inventory_py_web/frmLicenses.html', context)


@login_required
def api_get_license(request):
    """AJAX API: look up a license by serial number to fill the form."""
    serial = request.GET.get('serial_number', '').strip()
    if not serial:
        return JsonResponse({'success': False, 'error': 'Serial number not provided.'}, status=400)

    lic = OeesLicenses.objects.select_related('company', 'persone').filter(serial_number=serial).first()
    if not lic:
        return JsonResponse({'success': True, 'exists': False, 'serial': serial})

    data = {
        'serial': lic.serial_number or '',
        'company_id': lic.company_id or '',
        'type': lic.type or '',
        'origin': lic.origin or '',
        'date': lic.insert_date.strftime('%Y-%m-%d') if lic.insert_date else '',
        'value': str(lic.value) if lic.value is not None else '0',
        'obs': lic.obs or '',
        'bill': lic.bill_number or '',
        'person': lic.persone.name if lic.persone else 'Unassigned',
        'notes': lic.notes or '',
    }
    return JsonResponse({'success': True, 'exists': True, 'data': data})


# ==========================================================================
# Mobile Phones screen (web port of frmPhones.py)
# ==========================================================================

def _phone_under_repair(serial):
    return OeesUnderRepair.objects.filter(
        serial_number=serial, type='M', date_in__isnull=True
    ).exists()


def _phones_grid_rows():
    """Mobile phones for the grid, selecting only real columns (joined names included)."""
    return (OeesMobilePhones.objects
            .exclude(serial_number='Personal Mobile').exclude(type='DEVICE')
            .values('serial_number', 'company_id', 'company__name', 'brand', 'model', 'origin',
                    'insert_date', 'persone__name', 'id_line__number', 'imei', 'obs', 'value',
                    'bill_number', 'notes')
            .order_by('-id_mobile_phone'))


def _phones_export_excel():
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="phones_inventory.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mobile Phones"
    ws.append(['Serial Number', 'Company', 'Brand', 'Model', 'Origin', 'Insert Date',
               'Person', 'Number', 'IMEI', 'Obs', 'Value', 'Bill Number'])
    for r in _phones_grid_rows():
        ws.append([
            r['serial_number'], r['company__name'] or '', r['brand'] or '', r['model'] or '',
            r['origin'] or '', r['insert_date'].strftime('%Y-%m-%d') if r['insert_date'] else '',
            r['persone__name'] or '', r['id_line__number'] or '', r['imei'] or '',
            r['obs'] or '', r['value'], r['bill_number'] or '',
        ])
    wb.save(response)
    return response


@login_required
def frm_phones_view(request):
    user = request.user

    if request.method == 'POST':
        if _is_reader(user):
            messages.error(request, "You have a reader profile and can't modify data.")
            return redirect('frm_phones')
        action = request.POST.get('action', '')
        if action == 'save':
            return _phone_save(request)
        if action == 'support':
            return _phone_support(request)
        if action == 'release':
            return _phone_release(request)
        return redirect('frm_phones')

    # ---- GET ----
    if request.GET.get('export') == 'excel':
        return _phones_export_excel()

    companies = OeesCompanies.objects.all().order_by('name')
    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT origin FROM oees_mobile_phones "
                       "WHERE origin IS NOT NULL AND origin <> '' ORDER BY origin ASC")
        origin_options = [row[0] for row in cursor.fetchall()]

    grid_data = []
    total_value = 0.0
    for r in _phones_grid_rows():
        value = float(r['value'] or 0)
        total_value += value
        grid_data.append({
            'serial': r['serial_number'], 'company_id': r['company_id'] or '',
            'company': r['company__name'] or '', 'brand': r['brand'] or '', 'model': r['model'] or '',
            'origin': r['origin'] or '', 'date_obj': r['insert_date'], 'person': r['persone__name'] or '',
            'number': r['id_line__number'] or '', 'imei': r['imei'] or '', 'obs': r['obs'] or '',
            'value': value, 'bill': r['bill_number'] or '', 'notes': r['notes'] or '',
        })

    context = {
        'companies': companies,
        'origin_options': origin_options,
        'grid_data': grid_data,
        'total_phones': len(grid_data),
        'total_value': total_value,
        'preselected_sn': request.GET.get('sn', '').strip(),
        'is_reader': _is_reader(user),
    }
    return render(request, 'oe_inventory_py_web/frmPhones.html', context)


def _phone_save(request):
    user = request.user
    code = request.POST.get('serial_number', '').strip()
    if not code:
        messages.error(request, "Serial Number is required.")
        return redirect('frm_phones')

    company_id = request.POST.get('company') or None
    brand = request.POST.get('brand', '').strip()
    model = request.POST.get('model', '').strip()
    origin = request.POST.get('origin', '').strip()
    imei = request.POST.get('imei', '').strip()
    obs = request.POST.get('obs', '').strip()
    bill = request.POST.get('bill_number', '').strip()
    try:
        value = float(request.POST.get('value', '0'))
    except ValueError:
        value = 0.0
    insert_date = None
    date_str = request.POST.get('insert_date', '').strip()
    if date_str:
        try:
            insert_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            insert_date = None
    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    try:
        if OeesMobilePhones.objects.filter(serial_number=code).exists():
            existing = request.POST.get('notes', '').strip()
            notes = f"{now} - Modified by {user.username}\n{existing}".strip()
            # Partial UPDATE of known columns only (avoids any model/table mismatch).
            OeesMobilePhones.objects.filter(serial_number=code).update(
                company_id=company_id, brand=brand, model=model, origin=origin,
                insert_date=insert_date, value=value, imei=imei, obs=obs,
                notes=notes, bill_number=bill,
            )
        else:
            notes = f"{datetime.now()} - Created by {user.username}"
            with connection.cursor() as cur:
                cur.execute(
                    "INSERT INTO oees_mobile_phones "
                    "(serial_number, company, brand, model, origin, insert_date, value, imei, obs, notes, bill_number, type) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    [code, company_id, brand, model, origin, insert_date, value, imei, obs, notes, bill, ''],
                )
        messages.success(request, "Phone saved successfully.")
    except Exception:
        logger.exception("Error saving phone %s", code)
        messages.error(request, "An error occurred while saving the phone.")
    return redirect('frm_phones')


def _phone_support(request):
    serial = request.POST.get('serial_number', '').strip()
    if not serial:
        messages.error(request, "Please load a phone first.")
        return redirect('frm_phones')
    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    try:
        active = OeesUnderRepair.objects.filter(
            serial_number=serial, type='M', date_in__isnull=True
        ).first()
        if active:
            active.date_in = datetime.now().date()
            active.save()
            messages.success(request, f"Phone '{serial}' received from technical support.")
        else:
            OeesUnderRepair.objects.create(
                type='M', serial_number=serial, date_out=datetime.now().date(),
                destiny='Technical Service', value=0.0,
                notes=f"{now} - Sent to maintenance from the web panel by {request.user.username}",
            )
            messages.success(request, f"Phone '{serial}' sent to technical support.")
    except Exception:
        logger.exception("Error managing phone support for %s", serial)
        messages.error(request, "An error occurred managing technical support.")
    return redirect(f"{reverse('frm_phones')}?sn={serial}")


def _phone_release(request):
    serial = request.POST.get('serial_number', '').strip()
    person = request.POST.get('person', '').strip()
    if not serial:
        return redirect('frm_phones')
    if not person or person == 'Unassigned':
        messages.warning(request, "This phone is not assigned to anyone.")
        return redirect(f"{reverse('frm_phones')}?sn={serial}")
    # Who has it now (to document their remaining inventory after unassigning).
    rec = OeesMobilePhones.objects.filter(serial_number=serial).values('persone_id').first()
    staff_id = rec['persone_id'] if rec else None
    note = f"{datetime.now().strftime('%d/%m/%Y')} - Unassigned from {person} by {request.user.username}"
    try:
        OeesMobilePhones.objects.filter(serial_number=serial).update(
            persone=None, notes=Concat(Value(note + "\n"), Coalesce('notes', Value(''))),
        )
        messages.success(request, "Phone unassigned successfully.")
        # Document the remaining inventory for physical persons.
        if staff_id:
            staff = OeesStaff.objects.filter(id_staff=staff_id, persona_fisica=1).first()
            if staff:
                _save_staff_inventory_doc(staff, request.user, 'Unassign')
    except Exception:
        logger.exception("Error releasing phone %s", serial)
        messages.error(request, "An error occurred while unassigning the phone.")
    return redirect(f"{reverse('frm_phones')}?sn={serial}")


@login_required
def api_get_phone(request):
    """AJAX API: look up a mobile phone by serial number to fill the form."""
    serial = request.GET.get('serial_number', '').strip()
    if not serial:
        return JsonResponse({'success': False, 'error': 'Serial number not provided.'}, status=400)

    ph = (OeesMobilePhones.objects.filter(serial_number=serial)
          .values('company_id', 'brand', 'model', 'origin', 'insert_date', 'value',
                  'imei', 'obs', 'bill_number', 'persone__name', 'id_line__number', 'notes')
          .first())
    if not ph:
        return JsonResponse({'success': True, 'exists': False, 'serial': serial})

    data = {
        'serial': serial,
        'company_id': ph['company_id'] or '',
        'brand': ph['brand'] or '',
        'model': ph['model'] or '',
        'origin': ph['origin'] or '',
        'date': ph['insert_date'].strftime('%Y-%m-%d') if ph['insert_date'] else '',
        'value': str(ph['value']) if ph['value'] is not None else '0',
        'imei': ph['imei'] or '',
        'obs': ph['obs'] or '',
        'bill': ph['bill_number'] or '',
        'person': ph['persone__name'] or 'Unassigned',
        'number': ph['id_line__number'] or '',
        'notes': ph['notes'] or '',
        'under_repair': _phone_under_repair(serial),
    }
    return JsonResponse({'success': True, 'exists': True, 'data': data})


# ==========================================================================
# Fiber Lines screen (web port of frmFiberLines.py)
# ==========================================================================

def _fiber_export_excel():
    grid = OeesFiberLines.objects.select_related('id_delegation').order_by('-id_fiber_line')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="fiber_lines.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fiber Lines"
    ws.append(['ID', 'Description', 'Provider', 'Delegation', 'Order', 'Service Code', 'Access',
               'Router', 'Addressing', 'WIFI1', 'WIFI2', 'Active', 'Start Date', 'Down Date', 'Fixed IP'])
    for fl in grid:
        ws.append([
            fl.id_fiber_line, fl.description, fl.proveedor,
            fl.id_delegation.delegation if fl.id_delegation else '', fl.orden, fl.codigo_servicio,
            fl.acceso, fl.router, fl.direccionamiento, fl.wifi1, fl.wifi2,
            'Yes' if fl.estado else 'No',
            fl.fecha_inicio.strftime('%Y-%m-%d') if fl.fecha_inicio else '',
            fl.fecha_baja.strftime('%Y-%m-%d') if fl.fecha_baja else '', fl.ip_fija or '',
        ])
    wb.save(response)
    return response


def _fiber_incidences_export_excel(fiber_id):
    inc = OeesFiberLinesIncidences.objects.filter(id_fiber_line_id=fiber_id).order_by('-id_incidence')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="fiber_{fiber_id}_incidences.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Incidences"
    ws.append(['ID', 'Open Date', 'Open Description', 'Close Date', 'Close Description', 'Working Code'])
    for it in inc:
        ws.append([it.id_incidence, it.open_date or '', it.open_description or '',
                   it.close_date or '', it.close_description or '', it.working_code or ''])
    wb.save(response)
    return response


@login_required
def frm_fiber_view(request):
    user = request.user

    if request.method == 'POST':
        if _is_reader(user):
            messages.error(request, "You have a reader profile and can't modify data.")
            return redirect('frm_fiber')
        action = request.POST.get('action', '')
        if action == 'save':
            return _fiber_save(request)
        if action == 'save_incidence':
            return _fiber_save_incidence(request)
        return redirect('frm_fiber')

    # ---- GET ----
    export = request.GET.get('export')
    if export == 'excel':
        return _fiber_export_excel()
    if export == 'incidences':
        fid = request.GET.get('fiber', '').strip()
        if fid.isdigit():
            return _fiber_incidences_export_excel(fid)
        return redirect('frm_fiber')

    delegations = OeesDelegations.objects.all().order_by('id_delegation')
    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT proveedor FROM oees_fiber_lines "
                       "WHERE proveedor IS NOT NULL AND proveedor <> '' ORDER BY proveedor ASC")
        providers = [row[0] for row in cursor.fetchall()]

    context = {
        'delegations': delegations,
        'providers': providers,
        'fiber_list': OeesFiberLines.objects.select_related('id_delegation').order_by('-id_fiber_line'),
        'preselected_id': request.GET.get('fiber', '').strip(),
        'is_reader': _is_reader(user),
    }
    return render(request, 'oe_inventory_py_web/frmFiberLines.html', context)


def _fiber_save(request):
    user = request.user
    code = request.POST.get('code', '').strip()
    description = request.POST.get('description', '').strip()
    if not description:
        messages.error(request, "The Description field is required.")
        return redirect('frm_fiber')

    try:
        start_date = datetime.strptime(request.POST.get('start_date', '').strip(), '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, "A valid Start Date is required.")
        return redirect('frm_fiber')

    down_date = None
    down_str = request.POST.get('down_date', '').strip()
    if down_str:
        try:
            down_date = datetime.strptime(down_str, '%Y-%m-%d').date()
        except ValueError:
            down_date = None

    fields = {
        'description': description,
        'id_delegation_id': request.POST.get('delegation') or None,
        'proveedor': request.POST.get('provider', '').strip(),
        'orden': request.POST.get('order', '').strip(),
        'codigo_servicio': request.POST.get('service_code', '').strip(),
        'acceso': request.POST.get('access', '').strip(),
        'router': request.POST.get('router', '').strip(),
        'direccionamiento': request.POST.get('addressing', '').strip(),
        'wifi1': request.POST.get('wifi1', '').strip(),
        'wifi2': request.POST.get('wifi2', '').strip(),
        'estado': 1 if request.POST.get('estado') else 0,
        'fecha_inicio': start_date,
        'fecha_baja': down_date,
        'ip_fija': request.POST.get('ip_fixed', '').strip(),
    }
    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    try:
        if not code:
            obj = OeesFiberLines.objects.create(notes=f"{now} - Created by {user.username}", **fields)
            code = str(obj.id_fiber_line)
        else:
            fl = OeesFiberLines.objects.filter(id_fiber_line=code).first()
            if not fl:
                messages.error(request, "Fiber line to update does not exist.")
                return redirect('frm_fiber')
            existing = request.POST.get('notes', '').strip() or (fl.notes or '')
            for key, val in fields.items():
                setattr(fl, key, val)
            fl.notes = f"{now} - Modified by {user.username}\n{existing}".strip()
            fl.save()
        messages.success(request, "Fiber line saved successfully.")
    except Exception:
        logger.exception("Error saving fiber line %s", code)
        messages.error(request, "An error occurred while saving the fiber line.")
        return redirect('frm_fiber')
    return redirect(f"{reverse('frm_fiber')}?fiber={code}")


def _fiber_save_incidence(request):
    code = request.POST.get('code', '').strip()
    if not code.isdigit():
        messages.error(request, "Load a fiber line before adding incidences.")
        return redirect('frm_fiber')

    working = request.POST.get('working_code', '').strip()
    open_date = request.POST.get('open_date', '').strip()
    open_desc = request.POST.get('open_description', '').strip()
    close_date = request.POST.get('close_date', '').strip()
    close_desc = request.POST.get('close_description', '').strip()
    back = f"{reverse('frm_fiber')}?fiber={code}"

    if not working:
        messages.error(request, "You have to indicate a valid Working Order.")
        return redirect(back)
    if open_date and not open_desc:
        messages.error(request, "You have to indicate a valid Open Description.")
        return redirect(back)
    if close_date and not close_desc:
        messages.error(request, "You have to indicate a valid Close Description.")
        return redirect(back)

    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    try:
        OeesFiberLinesIncidences.objects.create(
            id_fiber_line_id=code, working_code=working,
            open_date=open_date or None, close_date=close_date or None,
            open_description=open_desc, close_description=close_desc,
            notes=f"{now} - Created by {request.user.username}",
        )
        messages.success(request, "Incidence saved successfully.")
    except Exception:
        logger.exception("Error saving incidence for fiber %s", code)
        messages.error(request, "An error occurred while saving the incidence.")
    return redirect(back)


@login_required
def api_get_fiber(request):
    """AJAX API: look up a fiber line plus its incidences and working codes."""
    fid = request.GET.get('id', '').strip()
    if not fid.isdigit():
        return JsonResponse({'success': False, 'error': 'No fiber id provided.'}, status=400)

    fl = OeesFiberLines.objects.select_related('id_delegation').filter(id_fiber_line=fid).first()
    if not fl:
        return JsonResponse({'success': True, 'exists': False, 'id': fid})

    incidences = list(
        OeesFiberLinesIncidences.objects.filter(id_fiber_line_id=fid).order_by('-id_incidence')
        .values('id_incidence', 'open_date', 'open_description', 'close_date',
                'close_description', 'working_code')
    )
    working_codes = list(
        OeesFiberLinesIncidences.objects.filter(id_fiber_line_id=fid)
        .exclude(working_code='').exclude(working_code__isnull=True)
        .values_list('working_code', flat=True).distinct()
    )
    data = {
        'id': fl.id_fiber_line,
        'description': fl.description or '',
        'delegation_id': fl.id_delegation_id or '',
        'provider': fl.proveedor or '',
        'order': fl.orden or '',
        'service_code': fl.codigo_servicio or '',
        'access': fl.acceso or '',
        'router': fl.router or '',
        'addressing': fl.direccionamiento or '',
        'wifi1': fl.wifi1 or '',
        'wifi2': fl.wifi2 or '',
        'active': 1 if fl.estado else 0,
        'start_date': fl.fecha_inicio.strftime('%Y-%m-%d') if fl.fecha_inicio else '',
        'down_date': fl.fecha_baja.strftime('%Y-%m-%d') if fl.fecha_baja else '',
        'ip_fixed': fl.ip_fija or '',
        'notes': fl.notes or '',
        'incidences': incidences,
        'working_codes': working_codes,
    }
    return JsonResponse({'success': True, 'exists': True, 'data': data})


# ==========================================================================
# Printers screen (web port of frmPrinters.py)
# ==========================================================================

def _printers_export_excel():
    grid = OeesPrinters.objects.select_related('id_delegation').order_by('-id_printer')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="printers.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Printers"
    ws.append(['Serial Number', 'Description', 'Provider', 'Delegation', 'MPS',
               'Start Date', 'Down Date', 'Fee', 'Fixed IP'])
    for pr in grid:
        ws.append([
            pr.serial_number, pr.description, pr.proveedor or '',
            pr.id_delegation.delegation if pr.id_delegation else '', pr.mps or '',
            pr.fecha_inicio.strftime('%Y-%m-%d') if pr.fecha_inicio else '',
            pr.fecha_baja.strftime('%Y-%m-%d') if pr.fecha_baja else '',
            pr.fee if pr.fee is not None else '', pr.ip or '',
        ])
    wb.save(response)
    return response


@login_required
def frm_printers_view(request):
    user = request.user

    if request.method == 'POST' and request.POST.get('action') == 'save':
        if _is_reader(user):
            messages.error(request, "You have a reader profile and can't modify data.")
            return redirect('frm_printers')
        return _printer_save(request)

    # ---- GET ----
    if request.GET.get('export') == 'excel':
        return _printers_export_excel()

    delegations = OeesDelegations.objects.all().order_by('id_delegation')
    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT proveedor FROM oees_printers "
                       "WHERE proveedor IS NOT NULL AND proveedor <> '' ORDER BY proveedor ASC")
        providers = [row[0] for row in cursor.fetchall()]

    context = {
        'delegations': delegations,
        'providers': providers,
        'printers_list': OeesPrinters.objects.select_related('id_delegation').order_by('-id_printer'),
        'preselected_sn': request.GET.get('sn', '').strip(),
        'is_reader': _is_reader(user),
        # Total fixed monthly cost across all printers (shown atop the list).
        'fee_total': OeesPrinters.objects.aggregate(t=Sum('fee'))['t'] or 0,
    }
    return render(request, 'oe_inventory_py_web/frmPrinters.html', context)


def _printer_save(request):
    user = request.user
    code = request.POST.get('serial_number', '').strip()
    if not code:
        messages.error(request, "Serial Number is required.")
        return redirect('frm_printers')
    description = request.POST.get('description', '').strip()
    if not description:
        messages.error(request, "The Description field is required.")
        return redirect('frm_printers')

    try:
        start_date = datetime.strptime(request.POST.get('start_date', '').strip(), '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, "A valid Start Date is required.")
        return redirect('frm_printers')

    down_date = None
    down_str = request.POST.get('down_date', '').strip()
    if down_str:
        try:
            down_date = datetime.strptime(down_str, '%Y-%m-%d').date()
        except ValueError:
            down_date = None

    fee = None
    fee_str = request.POST.get('fee', '').strip()
    if fee_str:
        try:
            fee = float(fee_str.replace(',', '.'))
        except ValueError:
            messages.error(request, "The Monthly fee must be a number.")
            return redirect('frm_printers')

    # Contract number: NULL when left blank. Page costs: default to 0.
    contract_number = request.POST.get('contract_number', '').strip() or None

    def _page_cost(field_name):
        raw = request.POST.get(field_name, '').strip()
        if not raw:
            return 0
        try:
            return float(raw.replace(',', '.'))
        except ValueError:
            return 0

    fields = {
        'description': description,
        'id_delegation_id': request.POST.get('delegation') or None,
        'proveedor': request.POST.get('provider', '').strip(),
        'mps': request.POST.get('mps', '').strip(),
        'fecha_inicio': start_date,
        'fecha_baja': down_date,
        'fee': fee,
        'contract_number': contract_number,
        'bw_page_cost': _page_cost('bw_page_cost'),
        'color_page_cost': _page_cost('color_page_cost'),
        'user': request.POST.get('user', '').strip(),
        'password': request.POST.get('password', '').strip(),
        'ip': request.POST.get('ip', '').strip(),
    }
    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    try:
        printer = OeesPrinters.objects.filter(serial_number=code).first()
        if printer:
            existing = request.POST.get('notes', '').strip() or (printer.notes or '')
            for key, val in fields.items():
                setattr(printer, key, val)
            printer.notes = f"{now} - Modified by {user.username}\n{existing}".strip()
            printer.save()
        else:
            OeesPrinters.objects.create(
                serial_number=code, notes=f"{now} - Created by {user.username}", **fields,
            )
        messages.success(request, "Printer saved successfully.")
    except Exception:
        logger.exception("Error saving printer %s", code)
        messages.error(request, "An error occurred while saving the printer.")
        return redirect('frm_printers')
    return redirect(f"{reverse('frm_printers')}?sn={code}")


@login_required
def api_get_printer(request):
    """AJAX API: look up a printer by serial number to fill the form."""
    serial = request.GET.get('serial_number', '').strip()
    if not serial:
        return JsonResponse({'success': False, 'error': 'Serial number not provided.'}, status=400)

    pr = OeesPrinters.objects.select_related('id_delegation').filter(serial_number=serial).first()
    if not pr:
        return JsonResponse({'success': True, 'exists': False, 'serial': serial})

    data = {
        'serial': pr.serial_number or '',
        'description': pr.description or '',
        'delegation_id': pr.id_delegation_id or '',
        'provider': pr.proveedor or '',
        'mps': pr.mps or '',
        'start_date': pr.fecha_inicio.strftime('%Y-%m-%d') if pr.fecha_inicio else '',
        'down_date': pr.fecha_baja.strftime('%Y-%m-%d') if pr.fecha_baja else '',
        'fee': str(pr.fee) if pr.fee is not None else '',
        'contract_number': pr.contract_number or '',
        'bw_page_cost': str(pr.bw_page_cost) if pr.bw_page_cost is not None else '0',
        'color_page_cost': str(pr.color_page_cost) if pr.color_page_cost is not None else '0',
        'user': pr.user or '',
        'password': pr.password or '',
        'ip': pr.ip or '',
        'notes': pr.notes or '',
    }
    return JsonResponse({'success': True, 'exists': True, 'data': data})


# ==========================================================================
# Allocations screen (web port of frmAllocations.py)
# ==========================================================================

def _unassigned(qs):
    """Items not assigned to any staff member (persone NULL or 0)."""
    return qs.filter(Q(persone__isnull=True) | Q(persone_id=0))


@login_required
def frm_allocations_view(request):
    user = request.user

    if request.method == 'POST':
        if _is_reader(user):
            messages.error(request, "You have a reader profile and can't modify data.")
            return redirect('frm_allocations')
        action = request.POST.get('action')
        if action == 'generate_doc':
            return _allocations_generate_doc(request)
        if action == 'auto_doc':
            return _allocations_auto_doc(request)
        return _allocations_assign(request)

    device_types = list(_unassigned(OeesDevices.objects).exclude(type='').exclude(type__isnull=True)
                        .values_list('type', flat=True).distinct().order_by('type'))
    device_brands = list(_unassigned(OeesDevices.objects).exclude(brand='').exclude(brand__isnull=True)
                         .values_list('brand', flat=True).distinct().order_by('brand'))
    license_types = list(_unassigned(OeesLicenses.objects).exclude(type='').exclude(type__isnull=True)
                         .values_list('type', flat=True).distinct().order_by('type'))

    staff_sel = request.GET.get('staff', '').strip()
    context = {
        'staff_list': OeesStaff.objects.filter(state=1).order_by('name').values('id_staff', 'name'),
        'device_types': device_types,
        'device_brands': device_brands,
        'license_types': license_types,
        'preselected_staff': int(staff_sel) if staff_sel.isdigit() else None,
        'is_reader': _is_reader(user),
    }
    return render(request, 'oe_inventory_py_web/frmAllocations.html', context)


def _save_staff_inventory_doc(staff, user, prefix):
    """Generate the staff inventory PDF (same layout as the Staff 'generate
    document' button) and save it as <prefix>-<date>-<time>.pdf in the person's
    documents. Non-fatal on error.
    """
    try:
        items, _total = staff_assigned_items(staff.id_staff)
        staff_data = {
            'name': staff.name or '',
            'dep': staff.department or '',
            'comp': staff.company.name if staff.company else '',
            'deleg': staff.delegation.delegation if staff.delegation else '',
            'fecha_i': staff.fecha_incorporacion or '',
        }
        pdf = build_staff_inventory_pdf(
            staff_data, items, generated_by=(user.get_full_name() or user.username))
        doc_name = f"{prefix}-{datetime.now().strftime('%d-%m-%Y-%H-%M-%S')}"
        # save() may append a suffix if the name exists; use the real saved name
        # so the OeesDocs record points to the actual file.
        saved_path = default_storage.save(f"staff_docs/{staff.id_staff}/{doc_name}.pdf", ContentFile(pdf))
        actual_name = saved_path.rsplit('/', 1)[-1]
        if actual_name.lower().endswith('.pdf'):
            actual_name = actual_name[:-4]
        OeesDocs.objects.create(
            id_staff_id=staff.id_staff, doc_name=actual_name,
            notes=f"{datetime.now()} - {prefix} document by {user.username}",
        )
    except Exception:
        logger.exception("Error generating %s doc for staff %s", prefix, staff.id_staff)


def _generate_allocation_doc(staff, user):
    """Inventory document generated after assigning items to a physical person."""
    _save_staff_inventory_doc(staff, user, 'Allocation')


def _pending_alloc_list(session):
    """Return the list of staff ids with assignments pending to be documented."""
    pend = session.get('pending_alloc_staff') or []
    return pend if isinstance(pend, list) else [pend]


def _add_pending_alloc(session, staff_id):
    pend = _pending_alloc_list(session)
    if staff_id not in pend:
        pend.append(staff_id)
    session['pending_alloc_staff'] = pend


def _remove_pending_alloc(session, staff_id):
    session['pending_alloc_staff'] = [p for p in _pending_alloc_list(session) if p != staff_id]


def _allocations_generate_doc(request):
    """Generate ONE allocation document for the selected staff, covering all the
    items currently assigned to them (groups several assignments into one PDF).
    Only for physical persons.
    """
    staff_id = request.POST.get('staff', '').strip()
    if not staff_id.isdigit():
        messages.warning(request, "Select a staff member first.")
        return redirect('frm_allocations')
    staff = OeesStaff.objects.filter(id_staff=staff_id).first()
    if not staff:
        messages.error(request, "Selected staff member not found.")
        return redirect('frm_allocations')
    if not staff.persona_fisica:
        messages.warning(request, "The allocation document is only generated for physical persons.")
        return redirect(f"{reverse('frm_allocations')}?staff={staff_id}")
    _generate_allocation_doc(staff, request.user)
    _remove_pending_alloc(request.session, staff.id_staff)
    messages.success(request, f"Allocation document generated for {staff.name}.")
    return redirect(f"{reverse('frm_allocations')}?staff={staff_id}")


def _allocations_auto_doc(request):
    """Safety net: when the user leaves frmAllocations with assignments not yet
    documented, the browser sends a beacon here and we generate ONE allocation
    document per pending staff member automatically. Idempotent: clears the list.
    """
    pending = _pending_alloc_list(request.session)
    request.session['pending_alloc_staff'] = []
    for sid in pending:
        staff = OeesStaff.objects.filter(id_staff=sid, persona_fisica=1).first()
        if staff:
            _generate_allocation_doc(staff, request.user)
    return HttpResponse(status=204)


def _allocations_assign(request):
    action = request.POST.get('action', '')
    config = {
        'assign_device': (OeesDevices, request.POST.get('device_serial', '').strip(), 'device'),
        'assign_license': (OeesLicenses, request.POST.get('license_serial', '').strip(), 'license'),
        'assign_phone': (OeesMobilePhones, request.POST.get('phone_serial', '').strip(), 'mobile phone'),
    }
    if action not in config:
        return redirect('frm_allocations')
    model, serial, label = config[action]

    staff_id = request.POST.get('staff', '').strip()
    if not staff_id.isdigit():
        messages.warning(request, "You must select a staff member first.")
        return redirect('frm_allocations')
    staff = OeesStaff.objects.filter(id_staff=staff_id).first()
    if not staff:
        messages.error(request, "Selected staff member not found.")
        return redirect('frm_allocations')
    if not serial:
        messages.warning(request, f"Select a {label} to assign.")
        return redirect('frm_allocations')

    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    note = f"{now} - Assigned to {staff.name} by {request.user.username}"
    try:
        updated = model.objects.filter(serial_number=serial).update(
            persone_id=staff.id_staff,
            notes=Concat(Value(note + "\n"), Coalesce('notes', Value(''))),
        )
        if updated:
            messages.success(request, f"{label.capitalize()} '{serial}' assigned to {staff.name}.")
            # Remember there are assignments pending to be documented (physical
            # persons only), so the doc is auto-generated if the user leaves
            # without pressing 'Generate document'.
            if staff.persona_fisica:
                _add_pending_alloc(request.session, staff.id_staff)
        else:
            messages.error(request, f"The selected {label} was not found.")
    except Exception:
        logger.exception("Error assigning %s %s", label, serial)
        messages.error(request, "An error occurred during the assignment.")
    # Keep the staff selected so several items can be assigned before generating
    # a single allocation document.
    return redirect(f"{reverse('frm_allocations')}?staff={staff_id}")


@login_required
def api_allocations_search(request):
    """Return the serial numbers of unassigned items for the allocation combos."""
    kind = request.GET.get('kind', '')
    if kind == 'devices':
        qs = _unassigned(OeesDevices.objects)
        device_type = request.GET.get('type', '').strip()
        brand = request.GET.get('brand', '').strip()
        if device_type:
            qs = qs.filter(type=device_type)
        if brand:
            qs = qs.filter(brand=brand)
    elif kind == 'licenses':
        qs = _unassigned(OeesLicenses.objects)
        license_type = request.GET.get('type', '').strip()
        if license_type:
            qs = qs.filter(type=license_type)
    elif kind == 'phones':
        qs = _unassigned(OeesMobilePhones.objects).exclude(serial_number__icontains='Personal Mobile')
    else:
        return JsonResponse({'success': False, 'error': 'Unknown kind.'}, status=400)

    serials = list(qs.exclude(serial_number='').values_list('serial_number', flat=True)
                   .order_by('serial_number')[:500])
    return JsonResponse({'success': True, 'serials': serials})


# ==========================================================================
# Incorporations screen (web port of frmIncorporations.py)
# ==========================================================================

def _chk(request, name):
    return 1 if request.POST.get(name) else 0


def _is_remote_delegation(delegation):
    return bool(delegation and ((delegation.delegation or '').strip().upper() == 'REMOTE'
                                or delegation.id_delegation == 11))


def _incorporation_rows(qs):
    rows = []
    for r in qs.select_related('company', 'delegation'):
        rows.append({
            'id': r.id,
            'name': r.name,
            'company': r.company.name if r.company else '',
            'department': r.department or '',
            'delegation': r.delegation.delegation if r.delegation else '',
            'date': r.insert_date,
            'flags': [r.win, r.mba, r.mbp, r.phone, r.screen, r.mouse, r.keyboard,
                      r.cordedh, r.cordlessh, r.usbchub, r.pdf, r.acad, r.send, r.receive],
        })
    return rows


def _incorporations_export_excel(tab):
    qs = OeesIncorporations.objects.select_related('company', 'delegation')
    if tab == 'discarded':
        qs = qs.filter(descartado=1)
    elif tab == 'incorporated':
        qs = qs.filter(incorporated=1)
    else:
        tab = 'pending'
        qs = qs.filter(descartado=0, incorporated=0)
    qs = qs.order_by('-id')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="incorporations_{tab}.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tab.capitalize()
    ws.append(['ID', 'Name', 'Company', 'Department', 'Delegation', 'Date', 'WIN', 'MBA', 'MBP',
               'Phone', 'Screen', 'Mouse', 'Keyboard', 'CordedH', 'CordlessH', 'USB-C HUB',
               'PDF', 'ACAD', 'Send', 'Receive'])
    for row in _incorporation_rows(qs):
        ws.append([
            row['id'], row['name'], row['company'], row['department'], row['delegation'],
            row['date'].strftime('%Y-%m-%d') if row['date'] else '',
            *['Yes' if f else '' for f in row['flags']],
        ])
    wb.save(response)
    return response


@login_required
def frm_incorporations_view(request):
    user = request.user

    if request.method == 'POST':
        if _is_reader(user):
            messages.error(request, "You have a reader profile and can't modify data.")
            return redirect('frm_incorporations')
        action = request.POST.get('action', '')
        handlers = {
            'save': _incorporation_save,
            'send': _incorporation_send,
            'receive': _incorporation_receive,
            'complete': _incorporation_complete,
        }
        handler = handlers.get(action)
        if handler:
            return handler(request)
        return redirect('frm_incorporations')

    # ---- GET ----
    if request.GET.get('export') == 'excel':
        return _incorporations_export_excel(request.GET.get('tab', 'pending'))

    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT department FROM oees_staff "
                       "WHERE department IS NOT NULL AND department <> '' ORDER BY department ASC")
        departments = [row[0] for row in cursor.fetchall()]

    context = {
        'companies': OeesCompanies.objects.all().order_by('name'),
        'delegations': OeesDelegations.objects.all().order_by('id_delegation'),
        'departments': departments,
        'pending_rows': _incorporation_rows(
            OeesIncorporations.objects.filter(descartado=0, incorporated=0)
            .order_by('-insert_date', '-id')),
        'discarded_rows': _incorporation_rows(
            OeesIncorporations.objects.filter(descartado=1).order_by('-id')),
        'incorporated_rows': _incorporation_rows(
            OeesIncorporations.objects.filter(incorporated=1).order_by('-id')),
        'preselected_id': request.GET.get('inc', '').strip(),
        'is_reader': _is_reader(user),
    }
    return render(request, 'oe_inventory_py_web/frmIncorporations.html', context)


def _incorporation_save(request):
    user = request.user
    code = request.POST.get('code', '').strip()
    name = request.POST.get('name', '').strip()
    company_id = request.POST.get('company') or None
    department = request.POST.get('department', '').strip()
    delegation_id = request.POST.get('delegation') or None
    date_str = request.POST.get('insert_date', '').strip()

    if not (name and company_id and department and delegation_id and date_str):
        messages.error(request, "Name, Company, Department, Delegation and Date are required.")
        return redirect('frm_incorporations')
    try:
        insert_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, "A valid Date is required.")
        return redirect('frm_incorporations')

    laptop = request.POST.get('laptop', '')
    headset = request.POST.get('headset', '')
    common = {
        'name': name, 'company_id': company_id, 'department': department,
        'delegation_id': delegation_id, 'insert_date': insert_date,
        'direccion': request.POST.get('direccion', '').strip(),
        'win': 1 if laptop == 'win' else 0,
        'mba': 1 if laptop == 'mba' else 0,
        'mbp': 1 if laptop == 'mbp' else 0,
        'cordedh': 1 if headset == 'corded' else 0,
        'cordlessh': 1 if headset == 'cordless' else 0,
        'phone': _chk(request, 'phone'), 'screen': _chk(request, 'screen'),
        'mouse': _chk(request, 'mouse'), 'keyboard': _chk(request, 'keyboard'),
        'descartado': _chk(request, 'descartado'), 'usbchub': _chk(request, 'usbchub'),
        'pdf': _chk(request, 'pdf'), 'acad': _chk(request, 'acad'),
    }
    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    try:
        if code:
            rec = OeesIncorporations.objects.filter(id=code).first()
            if not rec:
                messages.error(request, "Incorporation to update does not exist.")
                return redirect('frm_incorporations')
            previous = request.POST.get('notes', '').strip() or (rec.notes or '')
            for key, val in common.items():
                setattr(rec, key, val)
            rec.notes = f"{now} - Modified by {user.username}\n{previous}".strip()
            rec.save()
        else:
            OeesIncorporations.objects.create(
                incorporated=0, send=0, receive=0,
                notes=f"{now} - Added by {user.username}", **common,
            )
        messages.success(request, "Incorporation saved successfully.")
    except Exception:
        logger.exception("Error saving incorporation %s", code)
        messages.error(request, "An error occurred while saving the incorporation.")
    return redirect('frm_incorporations')


def _incorporation_send(request):
    code = request.POST.get('code', '').strip()
    agency = request.POST.get('agency', '').strip()
    if not code.isdigit():
        return redirect('frm_incorporations')
    rec = OeesIncorporations.objects.select_related('delegation').filter(id=code).first()
    if not rec:
        messages.error(request, "Incorporation not found.")
        return redirect('frm_incorporations')
    if not _is_remote_delegation(rec.delegation):
        messages.info(request, "You can only send devices for the REMOTE delegation.")
        return redirect(f"{reverse('frm_incorporations')}?inc={code}")
    if not agency:
        messages.warning(request, "You need to indicate the medium used to send the devices.")
        return redirect(f"{reverse('frm_incorporations')}?inc={code}")
    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    note = f"{now} - Device sent by {agency} - {request.user.username}"
    OeesIncorporations.objects.filter(id=code).update(
        send=1, notes=Concat(Value(note + "\n"), Coalesce('notes', Value(''))))
    messages.success(request, "Sending process completed successfully.")
    return redirect(f"{reverse('frm_incorporations')}?inc={code}")


def _incorporation_receive(request):
    code = request.POST.get('code', '').strip()
    if not code.isdigit():
        return redirect('frm_incorporations')
    if not OeesIncorporations.objects.filter(id=code).exists():
        messages.error(request, "Incorporation not found.")
        return redirect('frm_incorporations')
    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    note = f"{now} - Device received by {request.user.username}"
    OeesIncorporations.objects.filter(id=code).update(
        receive=1, notes=Concat(Value(note + "\n"), Coalesce('notes', Value(''))))
    messages.success(request, "Reception process completed successfully.")
    return redirect(f"{reverse('frm_incorporations')}?inc={code}")


def _incorporation_complete(request):
    code = request.POST.get('code', '').strip()
    rec = OeesIncorporations.objects.filter(id=code).first()
    if not rec:
        messages.error(request, "Incorporation not found.")
        return redirect('frm_incorporations')
    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    try:
        OeesStaff.objects.create(
            name=rec.name, department=rec.department or '', company_id=rec.company_id,
            state=1, delegation_id=rec.delegation_id, persona_fisica=1,
            notes=f"{now} - Migrated from Incorporations by {request.user.username}",
            fecha_incorporacion=rec.insert_date.strftime('%d-%m-%Y') if rec.insert_date else '',
        )
        OeesIncorporations.objects.filter(id=code).update(
            incorporated=1, descartado=2,
            notes=Concat(Value(f"{now} - Migrated to Staff by {request.user.username}\n"),
                         Coalesce('notes', Value(''))),
        )
        messages.success(request, "Incorporation completed. Go to Allocations to assign the devices for this person.")
    except Exception:
        logger.exception("Error completing incorporation %s", code)
        messages.error(request, "An error occurred while completing the incorporation.")
    return redirect('frm_incorporations')


@login_required
def api_get_incorporation(request):
    """AJAX API: look up an incorporation record to fill the form."""
    rid = request.GET.get('id', '').strip()
    if not rid.isdigit():
        return JsonResponse({'success': False, 'error': 'No incorporation id provided.'}, status=400)

    r = OeesIncorporations.objects.select_related('company', 'delegation').filter(id=rid).first()
    if not r:
        return JsonResponse({'success': True, 'exists': False, 'id': rid})

    laptop = 'win' if r.win else ('mba' if r.mba else ('mbp' if r.mbp else ''))
    headset = 'corded' if r.cordedh else ('cordless' if r.cordlessh else '')
    data = {
        'id': r.id, 'name': r.name or '', 'company_id': r.company_id or '',
        'department': r.department or '', 'delegation_id': r.delegation_id or '',
        'date': r.insert_date.strftime('%Y-%m-%d') if r.insert_date else '',
        'direccion': r.direccion or '', 'laptop': laptop, 'headset': headset,
        'phone': 1 if r.phone else 0, 'screen': 1 if r.screen else 0, 'mouse': 1 if r.mouse else 0,
        'keyboard': 1 if r.keyboard else 0, 'descartado': 1 if r.descartado else 0,
        'usbchub': 1 if r.usbchub else 0, 'pdf': 1 if r.pdf else 0, 'acad': 1 if r.acad else 0,
        'notes': r.notes or '', 'is_remote': _is_remote_delegation(r.delegation),
    }
    return JsonResponse({'success': True, 'exists': True, 'data': data})


# ==========================================================================
# Orders screen (web port of frmOrders.py)
# ==========================================================================

def _order_rows(qs):
    return [{'id': o.id_order, 'article': o.article, 'date': o.insert_date,
             'uds': o.uds, 'processed': o.tramitado} for o in qs]


def _orders_export_excel(tab):
    qs = OeesOrders.objects.all()
    if tab == 'canceled':
        qs = qs.filter(cancelado=1)
    elif tab == 'received':
        qs = qs.filter(recibido=1)
    else:
        tab = 'pending'
        qs = qs.filter(cancelado=0, recibido=0)
    qs = qs.order_by('-id_order')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="orders_{tab}.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tab.capitalize()
    ws.append(['ID', 'Article', 'Date', 'Uds', 'Processed'])
    for o in qs:
        ws.append([o.id_order, o.article, o.insert_date.strftime('%Y-%m-%d') if o.insert_date else '',
                   o.uds, 'Yes' if o.tramitado else 'No'])
    wb.save(response)
    return response


@login_required
def frm_orders_view(request):
    user = request.user

    if request.method == 'POST':
        if _is_reader(user):
            messages.error(request, "You have a reader profile and can't modify data.")
            return redirect('frm_orders')
        action = request.POST.get('action', '')
        handlers = {
            'save': _order_save,
            'cancel': _order_cancel,
            'process': _order_process,
            'receive': _order_receive,
        }
        handler = handlers.get(action)
        if handler:
            return handler(request)
        return redirect('frm_orders')

    # ---- GET ----
    if request.GET.get('export') == 'excel':
        return _orders_export_excel(request.GET.get('tab', 'pending'))

    context = {
        'pending_rows': _order_rows(
            OeesOrders.objects.filter(cancelado=0, recibido=0).order_by('-id_order')),
        'canceled_rows': _order_rows(
            OeesOrders.objects.filter(cancelado=1).order_by('-id_order')),
        'received_rows': _order_rows(
            OeesOrders.objects.filter(recibido=1).order_by('-id_order')),
        'preselected_id': request.GET.get('order', '').strip(),
        'is_reader': _is_reader(user),
    }
    return render(request, 'oe_inventory_py_web/frmOrders.html', context)


def _order_save(request):
    user = request.user
    code = request.POST.get('code', '').strip()
    article = request.POST.get('article', '').strip()
    uds_str = request.POST.get('uds', '').strip()
    date_str = request.POST.get('insert_date', '').strip()

    if not article:
        messages.error(request, "The 'Article' field cannot be empty.")
        return redirect('frm_orders')
    if not uds_str.isdigit():
        messages.error(request, "The 'Uds' field must be a whole number.")
        return redirect('frm_orders')
    try:
        insert_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, "A valid Date is required.")
        return redirect('frm_orders')
    uds = int(uds_str)
    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    try:
        if code:
            order = OeesOrders.objects.filter(id_order=code).first()
            if not order:
                messages.error(request, "Order to update does not exist.")
                return redirect('frm_orders')
            previous = request.POST.get('notes', '').strip() or (order.notes or '')
            order.article = article
            order.insert_date = insert_date
            order.uds = uds
            order.notes = f"{now} - Modified by {user.username}\n{previous}".strip()
            order.save()
        else:
            OeesOrders.objects.create(
                article=article, insert_date=insert_date, uds=uds,
                tramitado=0, cancelado=0, recibido=0,
                notes=f"{now} - Created by {user.username}",
            )
        messages.success(request, "Order saved successfully.")
    except Exception:
        logger.exception("Error saving order %s", code)
        messages.error(request, "An error occurred while saving the order.")
    return redirect('frm_orders')


def _order_status_change(request, field, label, can_change):
    code = request.POST.get('code', '').strip()
    order = OeesOrders.objects.filter(id_order=code).first()
    if not order:
        messages.error(request, "No order selected.")
        return redirect('frm_orders')
    error = can_change(order)
    if error:
        messages.warning(request, error)
        return redirect(f"{reverse('frm_orders')}?order={code}")
    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    note = f"{now} - {label} by {request.user.username}"
    OeesOrders.objects.filter(id_order=code).update(
        **{field: 1}, notes=Concat(Value(note + "\n"), Coalesce('notes', Value(''))))
    messages.success(request, f"Order {label.lower()}.")
    return redirect(f"{reverse('frm_orders')}?order={code}")


def _order_cancel(request):
    return _order_status_change(
        request, 'cancelado', 'Canceled',
        lambda o: "This order is processed. It cannot be canceled." if o.tramitado == 1 else None)


def _order_process(request):
    return _order_status_change(
        request, 'tramitado', 'Processed',
        lambda o: "This order is already processed." if o.tramitado == 1 else None)


def _order_receive(request):
    return _order_status_change(
        request, 'recibido', 'Received',
        lambda o: "You cannot receive an order that has not been processed." if o.tramitado == 0 else None)


@login_required
def api_get_order(request):
    """AJAX API: look up an order to fill the form."""
    oid = request.GET.get('id', '').strip()
    if not oid.isdigit():
        return JsonResponse({'success': False, 'error': 'No order id provided.'}, status=400)

    o = OeesOrders.objects.filter(id_order=oid).first()
    if not o:
        return JsonResponse({'success': True, 'exists': False, 'id': oid})

    data = {
        'id': o.id_order, 'article': o.article or '', 'uds': o.uds,
        'date': o.insert_date.strftime('%Y-%m-%d') if o.insert_date else '',
        'notes': o.notes or '',
        'tramitado': o.tramitado, 'cancelado': o.cancelado, 'recibido': o.recibido,
    }
    return JsonResponse({'success': True, 'exists': True, 'data': data})


# ==========================================================================
# Mobile Lines (SIM cards) screen (web port of frmMobileLines.py)
# ==========================================================================

_LINES_GRID_SQL = (
    "SELECT a.number, b.name as company_name, a.origin, a.pin, a.puk, a.pin2, a.puk2, a.imei, "
    "a.insert_date, c.serial_number as phone_serial, d.name as staff_name, a.extension, "
    "a.esim, a.M2M, a.fecha_baja, a.obs "
    "FROM oees_mobile_lines a "
    "LEFT JOIN oees_companies b ON a.company = b.id_company "
    "LEFT JOIN oees_mobile_phones c ON a.mobile = c.id_mobile_phone "
    "LEFT JOIN oees_staff d ON c.persone = d.id_staff "
    "ORDER BY a.id_mobile_line DESC"
)


def _lines_summary():
    queries = [
        ('asig', "SELECT origin, COUNT(*) FROM oees_mobile_lines WHERE mobile<>'' AND esim=0 AND m2m=0 AND fecha_baja IS NULL GROUP BY origin", ''),
        ('asig', "SELECT origin, COUNT(*) FROM oees_mobile_lines WHERE mobile<>'' AND esim=1 AND fecha_baja IS NULL GROUP BY origin", ' eSIM'),
        ('asig', "SELECT origin, COUNT(*) FROM oees_mobile_lines WHERE mobile<>'' AND m2m=1 AND fecha_baja IS NULL GROUP BY origin", ' M2M'),
        ('free', "SELECT origin, COUNT(*) FROM oees_mobile_lines WHERE mobile='' AND esim=0 AND m2m=0 AND fecha_baja IS NULL GROUP BY origin", ''),
        ('free', "SELECT origin, COUNT(*) FROM oees_mobile_lines WHERE mobile='' AND esim=1 AND fecha_baja IS NULL GROUP BY origin", ' eSIM'),
        ('free', "SELECT origin, COUNT(*) FROM oees_mobile_lines WHERE mobile='' AND m2m=1 AND fecha_baja IS NULL GROUP BY origin", ' M2M'),
        ('baja', "SELECT origin, COUNT(*) FROM oees_mobile_lines WHERE fecha_baja IS NOT NULL GROUP BY origin", ''),
    ]
    summary = {}
    with connection.cursor() as cur:
        for col, sql, suffix in queries:
            cur.execute(sql)
            for origin, cnt in cur.fetchall():
                key = (origin or '') + suffix
                bucket = summary.setdefault(key, {'asig': 0, 'free': 0, 'baja': 0})
                bucket[col] += int(cnt or 0)

    rows = []
    grand = {'asig': 0, 'free': 0, 'baja': 0}
    for key in sorted(summary):
        v = summary[key]
        total = v['asig'] + v['free'] + v['baja']
        rows.append({'type': key, 'asig': v['asig'], 'free': v['free'], 'baja': v['baja'], 'total': total})
        for k in grand:
            grand[k] += v[k]
    grand['total'] = grand['asig'] + grand['free'] + grand['baja']
    return rows, grand


def _lines_grid_rows():
    with connection.cursor() as cur:
        cur.execute(_LINES_GRID_SQL)
        raw = dict_fetchall(cur)
    rows = []
    for r in raw:
        rows.append({
            'number': r['number'], 'company': r['company_name'] or '', 'origin': r['origin'] or '',
            'pin': r['pin'] or '', 'puk': r['puk'] or '', 'pin2': r['pin2'] or '', 'puk2': r['puk2'] or '',
            'imei': r['imei'] or '', 'date': r['insert_date'], 'phone': r['phone_serial'] or '',
            'person': r['staff_name'] or '', 'extension': r['extension'] or '',
            'esim': r['esim'], 'm2m': r['M2M'], 'baja': bool(r['fecha_baja']), 'obs': r['obs'] or '',
        })
    return rows


def _lines_export_excel():
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="mobile_lines.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mobile Lines"
    ws.append(['Number', 'Company', 'Origin', 'PIN', 'PUK', 'PIN2', 'PUK2', 'IMEI', 'Insert Date',
               'Mobile', 'Person', 'Ext', 'eSIM', 'M2M', 'Cancelled', 'Obs'])
    for r in _lines_grid_rows():
        ws.append([r['number'], r['company'], r['origin'], r['pin'], r['puk'], r['pin2'], r['puk2'],
                   r['imei'], r['date'].strftime('%Y-%m-%d') if r['date'] else '', r['phone'], r['person'],
                   r['extension'], 'Yes' if r['esim'] else '', 'Yes' if r['m2m'] else '',
                   'Yes' if r['baja'] else '', r['obs']])
    wb.save(response)
    return response


@login_required
def frm_mobile_lines_view(request):
    user = request.user

    if request.method == 'POST':
        if _is_reader(user):
            messages.error(request, "You have a reader profile and can't modify data.")
            return redirect('frm_mobile_lines')
        action = request.POST.get('action', '')
        handlers = {
            'save': _line_save,
            'assign_mobile': _line_assign_mobile,
            'assign_esim': _line_assign_esim,
            'assign_m2m': _line_assign_m2m,
            'release': _line_release,
            'cancel': _line_cancel,
        }
        handler = handlers.get(action)
        if handler:
            return handler(request)
        return redirect('frm_mobile_lines')

    # ---- GET ----
    if request.GET.get('export') == 'excel':
        return _lines_export_excel()

    with connection.cursor() as cur:
        cur.execute("SELECT DISTINCT origin FROM oees_mobile_lines "
                    "WHERE origin IS NOT NULL AND origin <> '' ORDER BY origin ASC")
        origins = [row[0] for row in cur.fetchall()]

    # Phones available to receive a SIM: those WITHOUT a SIM (id_line) assigned.
    stock_phones = list(
        OeesMobilePhones.objects.filter(Q(id_line__isnull=True) | Q(id_line=0))
        .exclude(serial_number='Personal Mobile').order_by('serial_number')
        .values_list('serial_number', flat=True)
    )
    # Active physical staff (for eSIM assignment).
    staff = list(OeesStaff.objects.filter(state=1, persona_fisica=1)
                 .order_by('name').values('id_staff', 'name'))
    # M2M-capable devices not already linked.
    with connection.cursor() as cur:
        cur.execute(
            "SELECT serial_number FROM oees_devices WHERE mobile_line = 1 AND serial_number NOT IN "
            "(SELECT b.serial_number FROM oees_mobile_lines a "
            "INNER JOIN oees_mobile_phones b ON a.mobile = b.id_mobile_phone) "
            "ORDER BY serial_number ASC"
        )
        m2m_devices = [row[0] for row in cur.fetchall()]

    summary_rows, summary_total = _lines_summary()

    context = {
        'companies': OeesCompanies.objects.all().order_by('name'),
        'origins': origins,
        'stock_phones': stock_phones,
        'staff': staff,
        'm2m_devices': m2m_devices,
        'summary_rows': summary_rows,
        'summary_total': summary_total,
        'grid_rows': _lines_grid_rows(),
        'preselected_number': request.GET.get('number', '').strip(),
        'is_reader': _is_reader(user),
    }
    return render(request, 'oe_inventory_py_web/frmMobileLines.html', context)


def _line_redirect(code):
    if code:
        return redirect(f"{reverse('frm_mobile_lines')}?number={code}")
    return redirect('frm_mobile_lines')


def _prepend(text, existing):
    return f"{text}\n{existing or ''}"


def _line_save(request):
    user = request.user.username
    code = request.POST.get('number', '').strip()
    if not code:
        messages.error(request, "Number is required.")
        return redirect('frm_mobile_lines')

    company_id = request.POST.get('company') or None
    origin = request.POST.get('origin', '').strip()
    date_str = request.POST.get('insert_date', '').strip()
    insert_date = None
    if date_str:
        try:
            insert_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "A valid Insert Date is required (or leave it empty).")
            return _line_redirect(code)
    pin = request.POST.get('pin', '').strip()
    pin2 = request.POST.get('pin2', '').strip()
    puk = request.POST.get('puk', '').strip()
    puk2 = request.POST.get('puk2', '').strip()
    imei = request.POST.get('card', '').strip()
    extension = request.POST.get('extension', '').strip()
    obs = request.POST.get('obs', '').strip()
    esim = 1 if request.POST.get('esim') else 0
    m2m = 1 if request.POST.get('m2m') else 0
    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    try:
        with connection.cursor() as cur:
            cur.execute("SELECT esim, M2M, mobile FROM oees_mobile_lines WHERE number = %s", [code])
            existing = cur.fetchone()

            if not existing:
                cur.execute(
                    "INSERT INTO oees_mobile_lines "
                    "(number, company, imei, pin, puk, origin, insert_date, notes, pin2, puk2, extension, obs, esim, m2m) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    [code, company_id, imei, pin, puk, origin, insert_date,
                     f"{now} - Created by {user}", pin2, puk2, extension, obs, esim, m2m],
                )
            else:
                was_esim, was_m2m, mobile_id = existing
                # If eSIM or M2M was unchecked, free the linked phone first.
                if (was_esim == 1 and esim == 0) or (was_m2m == 1 and m2m == 0):
                    if mobile_id:
                        cur.execute("SELECT notes FROM oees_mobile_phones WHERE id_mobile_phone = %s", [mobile_id])
                        prow = cur.fetchone()
                        cur.execute(
                            "UPDATE oees_mobile_phones SET id_line = NULL, notes = %s WHERE id_mobile_phone = %s",
                            [_prepend(f"{now} - Unlinked by {user}", prow[0] if prow else ''), mobile_id],
                        )
                ui_notes = request.POST.get('notes', '').strip()
                cur.execute(
                    "UPDATE oees_mobile_lines SET company = %s, origin = %s, insert_date = %s, notes = %s, "
                    "imei = %s, pin = %s, puk = %s, pin2 = %s, puk2 = %s, extension = %s, obs = %s, "
                    "esim = %s, m2m = %s WHERE number = %s",
                    [company_id, origin, insert_date, f"{now} - Modified by {user}\n{ui_notes}",
                     imei, pin, puk, pin2, puk2, extension, obs, esim, m2m, code],
                )
        messages.success(request, "Mobile line saved successfully.")
    except Exception:
        logger.exception("Error saving mobile line %s", code)
        messages.error(request, "An error occurred while saving the mobile line.")
    return _line_redirect(code)


def _free_phones_with_line(cur, id_line, note):
    """Set id_line = NULL on any phone currently holding this line, prepending a note."""
    cur.execute("SELECT id_mobile_phone, notes FROM oees_mobile_phones WHERE id_line = %s", [id_line])
    for phone_id, phone_notes in cur.fetchall():
        cur.execute("UPDATE oees_mobile_phones SET id_line = NULL, notes = %s WHERE id_mobile_phone = %s",
                    [_prepend(note, phone_notes), phone_id])


def _line_assign_mobile(request):
    user = request.user.username
    code = request.POST.get('number', '').strip()
    mobile_serial = request.POST.get('mobile_serial', '').strip()
    if not code or not mobile_serial:
        return _line_redirect(code)
    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    try:
        with connection.cursor() as cur:
            cur.execute("SELECT id_mobile_line, notes FROM oees_mobile_lines WHERE number = %s", [code])
            r = cur.fetchone()
            if not r:
                messages.error(request, "Mobile line not found.")
                return _line_redirect(code)
            id_line, ui_notes = r[0], (r[1] or '')
            cur.execute(
                "SELECT a.id_mobile_phone, a.id_line, a.notes, b.number, b.notes "
                "FROM oees_mobile_phones a LEFT JOIN oees_mobile_lines b ON a.id_line = b.id_mobile_line "
                "WHERE a.serial_number = %s", [mobile_serial])
            p = cur.fetchone()
            if not p:
                messages.error(request, "Selected phone not found.")
                return _line_redirect(code)
            phone_id, phone_line, phone_notes, old_card, old_line_notes = p

            if not phone_line:
                _free_phones_with_line(cur, id_line, f"{now} - Removed card nº {code} by {user}")
                cur.execute("UPDATE oees_mobile_phones SET id_line = %s, notes = %s WHERE serial_number = %s",
                            [id_line, _prepend(f"{now} - Assigned SIM Card nº {code} by {user}", phone_notes), mobile_serial])
                cur.execute("UPDATE oees_mobile_lines SET mobile = %s, notes = %s WHERE id_mobile_line = %s",
                            [phone_id, _prepend(f"{now} - Assigned card to mobile phone {mobile_serial} by {user}", ui_notes), id_line])
            else:
                cur.execute("UPDATE oees_mobile_phones SET id_line = NULL, notes = %s WHERE serial_number = %s",
                            [_prepend(f"{now} - Removed SIM Card nº {old_card or ''} by {user}", phone_notes), mobile_serial])
                cur.execute("UPDATE oees_mobile_lines SET mobile = '', notes = %s WHERE id_mobile_line = %s",
                            [_prepend(f"{now} - Removed SIM Card from phone {mobile_serial} by {user}", old_line_notes), phone_line])
                cur.execute("UPDATE oees_mobile_phones SET id_line = %s, notes = %s WHERE serial_number = %s",
                            [id_line, _prepend(f"{now} - Assigned SIM Card nº {code} by {user}", ''), mobile_serial])
                cur.execute("UPDATE oees_mobile_lines SET mobile = %s, notes = %s WHERE id_mobile_line = %s",
                            [phone_id, _prepend(f"{now} - Assigned SIM Card to phone {mobile_serial} by {user}", ui_notes), id_line])
        messages.success(request, "Assignment completed successfully.")
    except Exception:
        logger.exception("Error assigning mobile line %s to phone %s", code, mobile_serial)
        messages.error(request, "An error occurred during the assignment.")
    return _line_redirect(code)


def _line_assign_esim(request):
    user = request.user.username
    code = request.POST.get('number', '').strip()
    staff_id = request.POST.get('staff', '').strip()
    if not code or not staff_id.isdigit():
        return _line_redirect(code)
    staff = OeesStaff.objects.filter(id_staff=staff_id).first()
    if not staff:
        messages.error(request, "Selected staff member not found.")
        return _line_redirect(code)
    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    today = datetime.now().strftime('%Y-%m-%d')
    try:
        with connection.cursor() as cur:
            cur.execute("SELECT id_mobile_line, notes FROM oees_mobile_lines WHERE number = %s", [code])
            r = cur.fetchone()
            if not r:
                messages.error(request, "Mobile line not found.")
                return _line_redirect(code)
            id_line, ui_notes = r[0], (r[1] or '')
            _free_phones_with_line(cur, id_line, f"{now} - Removed card nº {code} by {user}")

            cur.execute("SELECT COUNT(*) FROM oees_mobile_phones WHERE persone = %s AND serial_number = 'Personal Mobile'",
                        [staff_id])
            if cur.fetchone()[0] == 0:
                cur.execute(
                    "INSERT INTO oees_mobile_phones (serial_number, origin, insert_date, company, type, notes, value, persone, id_line) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    ['Personal Mobile', '', today, 1, 'MOBILE', f"{now} - Assigned card nº {code} by {user}", 0, staff_id, id_line],
                )
            cur.execute("SELECT id_mobile_phone FROM oees_mobile_phones WHERE serial_number = 'Personal Mobile' AND persone = %s",
                        [staff_id])
            mrow = cur.fetchone()
            mobile_id = mrow[0] if mrow else 0
            cur.execute("UPDATE oees_mobile_lines SET mobile = %s, esim = 1, notes = %s WHERE number = %s",
                        [mobile_id, _prepend(f"{now} - Assigned card as eSIM to {staff.name} by {user}", ui_notes), code])
        messages.success(request, "eSIM assignment completed successfully.")
    except Exception:
        logger.exception("Error assigning eSIM for line %s", code)
        messages.error(request, "An error occurred during the eSIM assignment.")
    return _line_redirect(code)


def _line_assign_m2m(request):
    user = request.user.username
    code = request.POST.get('number', '').strip()
    device = request.POST.get('device', '').strip()
    if not code or not device:
        return _line_redirect(code)
    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    today = datetime.now().strftime('%Y-%m-%d')
    try:
        with connection.cursor() as cur:
            cur.execute("SELECT id_mobile_line, notes FROM oees_mobile_lines WHERE number = %s", [code])
            r = cur.fetchone()
            if not r:
                messages.error(request, "Mobile line not found.")
                return _line_redirect(code)
            id_line, ui_notes = r[0], (r[1] or '')
            _free_phones_with_line(cur, id_line, f"{now} - Removed card nº {code} by {user}")

            cur.execute(
                "INSERT INTO oees_mobile_phones (serial_number, origin, insert_date, company, type, notes, value, id_line) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                [device, '', today, 1, 'DEVICE', f"{now} - Assigned card nº {code} by {user}", 0, id_line],
            )
            cur.execute("SELECT id_mobile_phone FROM oees_mobile_phones WHERE id_line = %s", [id_line])
            mrow = cur.fetchone()
            mobile_id = mrow[0] if mrow else 0
            cur.execute("UPDATE oees_mobile_lines SET mobile = %s, m2m = 1, notes = %s WHERE id_mobile_line = %s",
                        [mobile_id, _prepend(f"{now} - SIM card assigned to the M2M device {device} by {user}", ui_notes), id_line])
        messages.success(request, "M2M assignment completed successfully.")
    except Exception:
        logger.exception("Error assigning M2M for line %s", code)
        messages.error(request, "An error occurred during the M2M assignment.")
    return _line_redirect(code)


def _line_release(request):
    user = request.user.username
    code = request.POST.get('number', '').strip()
    if not code:
        return _line_redirect(code)
    now = datetime.now().strftime('%d/%m/%Y')
    try:
        with connection.cursor() as cur:
            cur.execute("SELECT mobile, notes FROM oees_mobile_lines WHERE number = %s", [code])
            r = cur.fetchone()
            if not r:
                messages.error(request, "Mobile line not found.")
                return _line_redirect(code)
            mobile_id, notes = r
            if not mobile_id:
                messages.warning(request, "This line is not assigned to a phone.")
                return _line_redirect(code)
            cur.execute("UPDATE oees_mobile_lines SET mobile = '', notes = %s WHERE number = %s",
                        [_prepend(f"{now} - Unassigned from phone by {user}", notes), code])
        messages.success(request, "Line unassigned successfully.")
    except Exception:
        logger.exception("Error releasing line %s", code)
        messages.error(request, "An error occurred while unassigning the line.")
    return _line_redirect(code)


def _line_cancel(request):
    user = request.user.username
    code = request.POST.get('number', '').strip()
    if not code:
        return _line_redirect(code)
    today = datetime.now().strftime('%Y-%m-%d')
    today_disp = datetime.now().strftime('%d/%m/%Y')
    try:
        with connection.cursor() as cur:
            cur.execute("SELECT mobile, fecha_baja, company FROM oees_mobile_lines WHERE number = %s", [code])
            r = cur.fetchone()
            if not r:
                messages.error(request, "Mobile line not found.")
                return _line_redirect(code)
            mobile_id, fecha_baja, company_id = r
            if mobile_id:
                messages.warning(request, "You cannot cancel a line assigned to a phone.")
                return _line_redirect(code)
            if fecha_baja:
                messages.warning(request, "This line has already been cancelled.")
                return _line_redirect(code)
            company_name = ''
            if company_id:
                cur.execute("SELECT name FROM oees_companies WHERE id_company = %s", [company_id])
                crow = cur.fetchone()
                company_name = crow[0] if crow else ''
            cur.execute("UPDATE oees_mobile_lines SET fecha_baja = %s, notes = %s WHERE number = %s",
                        [today, f"{today_disp} - CANCELLED with {company_name} by {user}", code])
        messages.success(request, "Line cancelled with the provider.")
    except Exception:
        logger.exception("Error cancelling line %s", code)
        messages.error(request, "An error occurred while cancelling the line.")
    return _line_redirect(code)


@login_required
def api_get_line(request):
    """AJAX API: look up a mobile line by number to fill the form."""
    number = request.GET.get('number', '').strip()
    if not number:
        return JsonResponse({'success': False, 'error': 'Number not provided.'}, status=400)

    with connection.cursor() as cur:
        cur.execute(
            "SELECT a.origin, b.name as company_name, a.company, a.insert_date, a.pin, a.pin2, a.puk, "
            "a.puk2, a.imei, a.extension, a.obs, c.serial_number, d.name as staff_name, a.fecha_baja, "
            "a.esim, a.M2M, a.notes "
            "FROM oees_mobile_lines a "
            "LEFT JOIN oees_companies b ON a.company = b.id_company "
            "LEFT JOIN oees_mobile_phones c ON a.mobile = c.id_mobile_phone "
            "LEFT JOIN oees_staff d ON c.persone = d.id_staff "
            "WHERE a.number = %s", [number])
        row = dict_fetchall(cur)

    if not row:
        return JsonResponse({'success': True, 'exists': False, 'number': number})

    r = row[0]
    esim = 1 if r['esim'] else 0
    m2m = 1 if r['M2M'] else 0
    mode = 'esim' if esim else ('m2m' if m2m else 'normal')
    insert_date = r['insert_date']
    baja = r['fecha_baja']
    data = {
        'number': number,
        'company_id': r['company'] or '',
        'origin': r['origin'] or '',
        'date': insert_date.strftime('%Y-%m-%d') if hasattr(insert_date, 'strftime') else (insert_date or ''),
        'pin': r['pin'] or '', 'pin2': r['pin2'] or '', 'puk': r['puk'] or '', 'puk2': r['puk2'] or '',
        'card': r['imei'] or '', 'extension': r['extension'] or '', 'obs': r['obs'] or '',
        'phone_serial': r['serial_number'] or '', 'person': r['staff_name'] or '',
        'esim': esim, 'm2m': m2m, 'mode': mode,
        'has_phone': bool(r['serial_number']),
        'baja': (baja.strftime('%d-%m-%Y') if hasattr(baja, 'strftime') else (str(baja) if baja else '')),
        'notes': r['notes'] or '',
    }
    return JsonResponse({'success': True, 'exists': True, 'data': data})


# ==========================================================================
# Availability screen (web port of frmAvailability.py)
# ==========================================================================

_AVAILABILITY_KEYS = [
    "LAPTOP WIN", "LAPTOP MBA", "LAPTOP MBP", "KEYBOARD", "MOUSE", "SCREEN",
    "CORDED HEADSET", "CORDLESS HEADSET", "USBC HUB", "PHONE",
]


def _availability_rows():
    data = {}

    def bucket(article):
        return data.setdefault(article, {'stock': 0, 'needs': 0, 'orders': 0})

    with connection.cursor() as cur:
        # Stock: unassigned devices grouped by type.
        cur.execute("SELECT COUNT(*), type FROM oees_devices "
                    "WHERE (persone IS NULL OR persone = '') GROUP BY type")
        for qty, article in cur.fetchall():
            if article:
                bucket(article)['stock'] = int(qty or 0)

        # Stock: unassigned mobile phones.
        cur.execute("SELECT COUNT(*) FROM oees_mobile_phones WHERE (persone IS NULL OR persone = '')")
        phones = cur.fetchone()[0] or 0
        if phones > 0:
            bucket('PHONE')['stock'] = int(phones)

        # Needs: pending incorporations.
        cur.execute(
            "SELECT SUM(win), SUM(mba), SUM(mbp), SUM(keyboard), SUM(mouse), SUM(screen), "
            "SUM(cordedh), SUM(cordlessh), SUM(usbchub), SUM(phone) "
            "FROM oees_incorporations WHERE incorporated = 0 AND descartado = 0"
        )
        s = cur.fetchone() or [0] * 10
        needs_map = {
            "LAPTOP WIN": s[0], "LAPTOP MBA": s[1], "LAPTOP MBP": s[2], "KEYBOARD": s[3],
            "MOUSE": s[4], "SCREEN": s[5], "CORDED HEADSET": s[6], "CORDLESS HEADSET": s[7],
            "USBC HUB": s[8], "PHONE": s[9],
        }
        for article, qty in needs_map.items():
            qty = int(qty or 0)
            if qty > 0:
                bucket(article)['needs'] = qty

        # Orders: pending orders matched by article text.
        cur.execute("SELECT article, uds FROM oees_orders WHERE recibido = 0 AND cancelado = 0")
        for article, uds in cur.fetchall():
            uds = int(uds or 0)
            article_upper = (article or '').upper()
            for key in _AVAILABILITY_KEYS:
                if key in article_upper:
                    bucket(key)['orders'] += uds
                    break

    rows = []
    for article in sorted(data):
        stock = data[article]['stock']
        needs = data[article]['needs']
        orders = data[article]['orders']
        disp = stock - needs + orders
        rows.append({'article': article, 'stock': stock, 'needs': needs,
                     'orders': orders, 'disp': disp, 'positive': disp >= 0})
    return rows


@login_required
def frm_availability_view(request):
    rows = _availability_rows()

    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="availability.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Availability"
        ws.append(['Article', 'Stock', 'Needs', 'Orders', 'Disp'])
        for r in rows:
            ws.append([r['article'], r['stock'], r['needs'], r['orders'], r['disp']])
        wb.save(response)
        return response

    return render(request, 'oe_inventory_py_web/frmAvailability.html', {'rows': rows})


# ==========================================================================
# Not Returned screen: items still assigned to staff who have already left
# (oees_staff.fecha_baja is set) — i.e. material not handed back.
# ==========================================================================

def _parse_flexible_date(value):
    """Parse a date stored as a string in mixed formats (legacy data has both
    English yyyy-mm-dd and Spanish dd-mm-yyyy / dd/mm/yyyy). Returns a date or None.
    """
    if not value:
        return None
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(value.strip(), fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _with_person_subtotals(rows):
    """Insert a highlighted subtotal row after each person's group (the rows
    are already ordered by person). Subtotal rows carry sub=True."""
    out = []
    if not rows:
        return out
    current = rows[0]['person']
    running = 0.0
    for r in rows:
        if r['person'] != current:
            out.append({'sub': True, 'person': f'Total {current}', 'termination_date': '',
                        'aging_days': None, 'category': '', 'serial': '', 'description': '',
                        'value': running})
            current = r['person']
            running = 0.0
        out.append(r)
        running += r['value'] or 0
    out.append({'sub': True, 'person': f'Total {current}', 'termination_date': '',
                'aging_days': None, 'category': '', 'serial': '', 'description': '', 'value': running})
    return out


def _not_returned_rows():
    """Items still assigned to terminated staff, grouped by person.

    Covers devices, licenses, phones, access cards and access keys. Returns a
    flat list ordered by person, each row with the termination date (normalised
    to yyyy-mm-dd), the days elapsed since (aging), the serial/code, a
    description and the value (if any).
    """
    today = datetime.now().date()
    terminated = (OeesStaff.objects
                  .exclude(fecha_baja__isnull=True).exclude(fecha_baja='')
                  .order_by('name'))

    rows = []
    for staff in terminated:
        parsed = _parse_flexible_date(staff.fecha_baja)
        term_date = parsed.strftime('%Y-%m-%d') if parsed else (staff.fecha_baja or '')
        aging = (today - parsed).days if parsed else None

        def add(category, serial, description, value):
            rows.append({
                'sub': False,
                'person': staff.name or '',
                'termination_date': term_date,
                'aging_days': aging,
                'category': category,
                'serial': serial or '',
                'description': description or '',
                'value': value,
            })

        for d in OeesDevices.objects.filter(persone_id=staff.id_staff):
            desc = ' '.join(x for x in [d.type, d.brand, d.model] if x).strip()
            add('Device', d.serial_number, desc, d.value)
        for lic in OeesLicenses.objects.filter(persone_id=staff.id_staff):
            add('License', lic.serial_number, lic.type, lic.value)
        for p in OeesMobilePhones.objects.filter(persone_id=staff.id_staff):
            desc = ' '.join(x for x in [p.type, p.brand, p.model] if x).strip()
            add('Phone', p.serial_number, desc, p.value)
        for c in OeesAccessCards.objects.filter(id_staff_id=staff.id_staff).exclude(state_card=4):
            add('Access Card', c.ac_max, 'Office Access Card', None)
        for k in OeesAccessKeys.objects.filter(id_staff_id=staff.id_staff):
            add('Access Key', '', k.type or 'Access Key', None)

    return rows


def _not_returned_export_excel(rows):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="not_returned.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Not Returned"
    ws.append(['Person', 'Termination Date', 'Days', 'Category', 'Serial / Code', 'Description', 'Value (€)'])
    for r in rows:
        ws.append([
            r['person'], r['termination_date'],
            r['aging_days'] if r['aging_days'] is not None else '',
            r['category'], r['serial'], r['description'],
            r['value'] if r['value'] is not None else '',
        ])
    wb.save(response)
    return response


@login_required
def frm_not_returned_view(request):
    rows = _not_returned_rows()
    show_subtotals = request.GET.get('subtotals') == 'on'

    total_value = sum(r['value'] for r in rows if r['value'])
    total_items = len(rows)
    display_rows = _with_person_subtotals(rows) if show_subtotals else rows

    if request.GET.get('export') == 'excel':
        return _not_returned_export_excel(display_rows)

    return render(request, 'oe_inventory_py_web/frmNotReturned.html', {
        'rows': display_rows,
        'total_items': total_items,
        'total_value': total_value,
        'show_subtotals': show_subtotals,
    })


# ==========================================================================
# Omada screen: per-site overview pulled from the TP-Link Omada Open API.
# ==========================================================================

@login_required
def frm_omada_view(request):
    from . import omada
    configured = omada.omada_configured()
    rows, error = [], None
    if configured:
        try:
            rows = omada.site_overview()
        except Exception:
            logger.exception("Omada API error")
            error = ("Could not reach the Omada controller. Check the API "
                     "credentials and that the controller is reachable.")
    return render(request, 'oe_inventory_py_web/frmOmada.html', {
        'rows': rows,
        'configured': configured,
        'error': error,
    })


@login_required
def frm_net_overview_view(request):
    import json
    from . import nebula, status_cache
    # The page first loads a spinner shell; the data arrives via this same view
    # with ?partial=1 (AJAX). The data is served from the background-refreshed
    # cache (status_cache) so this request never waits on the slow Nebula API —
    # that synchronous call behind gunicorn's 30 s timeout caused 502s on the
    # server. While the first compute is still running we return HTTP 202 and
    # the shell retries.
    if not request.GET.get('partial'):
        return render(request, 'oe_inventory_py_web/frmNetOverview.html', {})

    configured = nebula.nebula_configured()
    if not configured:
        return render(request, 'oe_inventory_py_web/_net_overview_content.html',
                      {'rows': [], 'configured': False, 'error': None})

    rows, error = status_cache.get_net_overview()
    if rows is None and not error:
        # Cold start: the background compute hasn't produced data yet.
        return HttpResponse(status=202)  # the shell keeps the spinner and retries

    rows = rows or []
    # Serialise each site's offline devices for the alerts pop-up and its
    # topology for the map modal.
    for r in rows:
        r['alert_json'] = json.dumps(r.get('alert_list') or [])
        r['topology_json'] = json.dumps(r.get('topology') or {})
    return render(request, 'oe_inventory_py_web/_net_overview_content.html', {
        'rows': rows,
        'configured': True,
        'error': error,
    })


@login_required
def api_footer_counts(request):
    """Live footer counters (pending orders, pending cards, online users), so the
    status bar refreshes periodically without reloading the page. Pending counts
    come from the background-refreshed cache; online users are counted live."""
    from . import status_cache
    from .context_processors import online_user_ids
    status = status_cache.get_status()
    ids = online_user_ids()
    ids.add(str(request.user.pk))  # the requester is online by definition
    return JsonResponse({
        'total_orders': status.get('total_orders') or 0,
        'total_cards': status.get('total_cards') or 0,
        'online_users': len(ids),
        'last_update': status_cache.last_update_str(),
    })


@login_required
def api_net_alerts(request):
    """Total number of active network alerts (offline + outdated devices across
    all sites), for the footer 'Net Alerts' badge. Polled periodically by the
    footer JS and served instantly from the background-refreshed cache.

    Only users with the net_overview permission may query it. Returns
    ``{'alerts': N, 'ok': True}`` when a figure is available; ``ok=False`` while
    it is still unknown (cold start / Nebula unreachable) so the badge hides."""
    if not getattr(request.user, 'net_overview', 0):
        return JsonResponse({'alerts': 0, 'ok': False}, status=403)

    # Read the background-refreshed cache (never call Nebula in the request path).
    from . import status_cache
    status = status_cache.get_status()
    n = status.get('net_alerts')
    ad = status.get('anydesk_alerts')
    vr = status.get('video_rooms_alerts')
    return JsonResponse({
        'alerts': n or 0, 'ok': n is not None,
        'anydesk_alerts': ad or 0, 'anydesk_ok': ad is not None,
        'video_rooms_alerts': vr or 0, 'video_rooms_ok': vr is not None,
    })


@login_required
def frm_remote_machines_view(request):
    """AnyDesk screen: shows each remote machine from oees_anydesk as a card with
    its description and an online/offline status dot. The accessibility check runs
    in the background process (status_cache); this screen just displays the latest
    result. Gated by the net_overview permission.

    Status per machine: the last background-check result when available; otherwise
    (e.g. before the API key is set) it falls back to whether last_connection is
    set, so the design can still be tuned with green/red dots."""
    if not getattr(request.user, 'net_overview', 0):
        return redirect('mdi_home')

    from . import anydesk, status_cache
    configured = anydesk.anydesk_configured()
    status_map = status_cache.get_anydesk_status()   # code(str) -> online(bool)

    machines, available = [], True
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM oees_anydesk ORDER BY 1")
            cols = [c[0] for c in cursor.description]

            def idx(col):
                return cols.index(col) if col else None

            # Tolerate schema naming variants (the live column is `last_connetion`).
            i_code = idx(anydesk.pick_column(cols, 'code'))
            i_last = idx(anydesk.pick_column(cols, 'last_connection', prefix='last_conn'))
            i_desc = idx(anydesk.pick_column(cols, 'description', 'name'))

            for r in cursor.fetchall():
                code = str(r[i_code]).strip() if i_code is not None else ''
                last = r[i_last] if i_last is not None else None
                desc = (r[i_desc] if i_desc is not None else None) or code
                # Real status when the background check knows it; otherwise fall
                # back to last_connection so the dots still render for design.
                if code in status_map:
                    online = status_map[code]
                else:
                    online = last is not None
                machines.append({
                    'code': code, 'description': desc,
                    'last_connection': last, 'online': online,
                })
    except Exception:
        logger.warning("oees_anydesk table not available in this environment")
        available = False

    return render(request, 'oe_inventory_py_web/frmRemoteMachines.html', {
        'machines': machines,
        'available': available,
        'configured': configured,
    })


def _booking_incidences(bookings):
    """English incidence lines for future bookings whose organizer email is, in
    oees_staff, deactivated (has a fecha_baja) or not found at all. Bookings are
    grouped per organizer email; the count is the number of future bookings."""
    from collections import Counter
    counts = Counter()
    for b in bookings:
        email = (b.get('organizer_email') or '').strip().lower()
        if email:
            counts[email] += 1

    lines = []
    if not counts:
        return lines
    with connection.cursor() as cur:
        for email in sorted(counts):
            n = counts[email]
            cur.execute(
                "SELECT name, fecha_baja FROM oees_staff "
                "WHERE LOWER(TRIM(email)) = %s ORDER BY id_staff DESC LIMIT 1", [email])
            row = cur.fetchone()
            if not row:
                lines.append(f"The email {email} is not found in the users table.")
            elif row[1]:  # fecha_baja set -> the organizer is on leave
                plural = 's' if n != 1 else ''
                lines.append(f"User {row[0]}, deactivated since {row[1]}, "
                             f"has {n} future booking{plural}.")
    return lines


def _low_occupancy_meetings():
    """Meetings whose effective occupancy was <= 50% of their duration, from
    oees_meeting_room. Each row: date, start time, title, organizer, % occupied."""
    from django.db.models import F
    from .models import OeesMeetingRoom
    rows = []
    qs = (OeesMeetingRoom.objects
          .filter(duration__gt=0, occupied__lte=F('duration') / 2.0)
          .order_by('start_time'))
    for m in qs:
        rows.append({
            'date': m.start_time.strftime('%d-%m-%Y') if m.start_time else '—',
            'time': m.start_time.strftime('%H:%M') if m.start_time else '—',
            # End time when known; the template falls back to the duration if not.
            'end': m.end_time.strftime('%H:%M') if m.end_time else None,
            'duration': m.duration,
            'title': m.description or '—',
            'organizer': m.org_email or '—',
            'pct': round(m.occupied / m.duration * 100),
        })
    return rows


def _organizer_no_show_ranking():
    """Ranking of organizers by number of meetings that effectively didn't take
    place (occupied <= 10 minutes), most first. Organizer names are resolved from
    oees_staff by email; unknown emails are shown as-is."""
    from django.db.models import Count, Sum
    from .models import OeesMeetingRoom
    agg = list(OeesMeetingRoom.objects
               .filter(occupied__lte=10)
               .exclude(org_email__isnull=True).exclude(org_email='')
               .values('org_email')
               .annotate(n=Count('id'),
                         total_duration=Sum('duration'), total_occupied=Sum('occupied'))
               .order_by('-n', 'org_email'))
    if not agg:
        return []
    name_by_email = {}
    emails = [a['org_email'].strip().lower() for a in agg if a['org_email']]
    with connection.cursor() as cur:
        ph = ','.join(['%s'] * len(emails))
        cur.execute(f"SELECT LOWER(TRIM(email)), name FROM oees_staff "
                    f"WHERE LOWER(TRIM(email)) IN ({ph})", emails)
        for em, nm in cur.fetchall():
            name_by_email[em] = nm
    rows = []
    for a in agg:
        em = (a['org_email'] or '').strip()
        rows.append({'organizer': name_by_email.get(em.lower()) or em or '—',
                     'count': a['n'],
                     'total_duration': a['total_duration'] or 0,
                     'total_occupied': a['total_occupied'] or 0})
    return rows


def _demo_no_show_ranking():
    """Sample organizer no-show ranking for the design preview (no real data yet)."""
    return [
        {'organizer': 'Ana García', 'count': 7, 'total_duration': 420, 'total_occupied': 25},
        {'organizer': 'Luis Pérez', 'count': 5, 'total_duration': 300, 'total_occupied': 15},
        {'organizer': 'Marta Ruiz', 'count': 4, 'total_duration': 240, 'total_occupied': 20},
        {'organizer': 'Javier Soler', 'count': 2, 'total_duration': 90, 'total_occupied': 5},
        {'organizer': 'Carlos Méndez', 'count': 1, 'total_duration': 60, 'total_occupied': 0},
    ]


def _demo_low_occupancy():
    """Sample low-occupancy meetings for the design preview (no real data yet)."""
    # occupied & duration are multiples of 5 (5-minute sampling); pct = occ/dur.
    return [
        {'date': '24-06-2026', 'time': '09:00', 'end': '10:00', 'duration': 60,
         'title': 'Comité de Dirección', 'organizer': 'ana.garcia@octopusenergy.es', 'pct': 25},
        {'date': '24-06-2026', 'time': '11:30', 'end': '12:10', 'duration': 40,
         'title': 'Sprint Review', 'organizer': 'marta.ruiz@octopusenergy.es', 'pct': 50},
        {'date': '24-06-2026', 'time': '13:00', 'end': '13:20', 'duration': 20,
         'title': '1:1 RRHH', 'organizer': 'former.employee@octopusenergy.es', 'pct': 25},
        {'date': '25-06-2026', 'time': '10:00', 'end': '11:00', 'duration': 60,
         'title': 'Onboarding Q3', 'organizer': 'luis.perez@octopusenergy.es', 'pct': 50},
        {'date': '25-06-2026', 'time': '16:00', 'end': '16:50', 'duration': 50,
         'title': 'Revisión Presupuesto', 'organizer': 'javier.soler@octopusenergy.es', 'pct': 40},
    ]


def _demo_bookings():
    """Sample future bookings for the design preview: a couple of real
    deactivated staff (so the message shows live data) plus an unknown email."""
    bookings = []
    with connection.cursor() as cur:
        cur.execute(
            "SELECT email FROM oees_staff WHERE fecha_baja IS NOT NULL AND fecha_baja <> '' "
            "AND email IS NOT NULL AND email <> '' ORDER BY id_staff DESC LIMIT 3")
        for i, (email,) in enumerate(cur.fetchall()):
            for _ in range(i + 1):   # 1, 2, 3 future bookings -> varied counts
                bookings.append({'organizer_email': email})
    bookings.append({'organizer_email': 'former.employee@example.com'})  # not in staff
    return bookings


@login_required
def frm_video_rooms_view(request):
    """Videoconference rooms status from the Logitech Sync Cloud API (mTLS).
    Read-only overview, gated by the net_overview permission. Shows a
    "not configured" notice until the client certificate/key are in place.
    The right panel lists future-booking incidences (organizer on leave or
    unknown email) cross-referenced against oees_staff."""
    if not getattr(request.user, 'net_overview', 0):
        return redirect('mdi_home')

    from . import logitech
    configured = logitech.logitech_configured()
    rooms, error, demo = [], None, False
    booking_incidences, low_occupancy, no_show_ranking = [], [], []
    if configured:
        try:
            rooms = logitech.rooms_overview()
        except Exception:
            logger.exception("Logitech Sync API error")
            error = ("Could not reach the Logitech Sync Cloud API. Check the "
                     "certificate, private key and base URL.")
        if not error:
            try:
                booking_incidences = _booking_incidences(logitech.future_bookings())
            except Exception:
                logger.warning("Logitech bookings unavailable")
            low_occupancy = _low_occupancy_meetings()
            no_show_ranking = _organizer_no_show_ranking()
    else:
        # No certificate/license yet: show sample rooms + incidences for design.
        rooms = logitech.demo_rooms()
        demo = True
        booking_incidences = _booking_incidences(_demo_bookings())
        low_occupancy = _demo_low_occupancy()
        no_show_ranking = _demo_no_show_ranking()
    return render(request, 'oe_inventory_py_web/frmVideoRooms.html', {
        'rooms': rooms,
        'configured': configured,
        'error': error,
        'demo': demo,
        'booking_incidences': booking_incidences,
        'low_occupancy': low_occupancy,
        'no_show_ranking': no_show_ranking,
    })


# ==========================================================================
# Under Repair screen (web port of frmUnderRepair.py)
# ==========================================================================

def _under_repair_rows(repaired):
    cond = "a.date_in IS NOT NULL" if repaired else "a.date_in IS NULL"
    sql = (
        "SELECT a.id_under_repair, a.serial_number, b.brand, b.model, a.date_out, a.date_in, a.destiny, a.value "
        "FROM oees_under_repair a INNER JOIN oees_devices b ON a.serial_number = b.serial_number "
        f"WHERE {cond} AND a.type = 'D' "
        "UNION "
        "SELECT a.id_under_repair, a.serial_number, b.brand, b.model, a.date_out, a.date_in, a.destiny, a.value "
        "FROM oees_under_repair a INNER JOIN oees_mobile_phones b ON a.serial_number = b.serial_number "
        f"WHERE {cond} AND a.type = 'M' "
        "ORDER BY id_under_repair DESC"
    )
    with connection.cursor() as cur:
        cur.execute(sql)
        raw = dict_fetchall(cur)
    rows = []
    for r in raw:
        model = f"{r['brand'] or ''} {r['model'] or ''}".strip()
        rows.append({
            'id': r['id_under_repair'], 'serial': r['serial_number'], 'model': model,
            'date_out': r['date_out'], 'date_in': r['date_in'], 'destiny': r['destiny'] or '',
            'value': r['value'],
        })
    return rows


def _under_repair_export_excel(tab):
    repaired = (tab == 'repaired')
    rows = _under_repair_rows(repaired)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="under_repair_{"repaired" if repaired else "pending"}.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Repaired" if repaired else "Pending"
    if repaired:
        ws.append(['ID', 'Serial Number', 'Model', 'Date Out', 'Date In', 'Destiny', 'Value'])
        for r in rows:
            ws.append([r['id'], r['serial'], r['model'],
                       r['date_out'].strftime('%Y-%m-%d') if r['date_out'] else '',
                       r['date_in'].strftime('%Y-%m-%d') if r['date_in'] else '',
                       r['destiny'], r['value']])
    else:
        ws.append(['ID', 'Serial Number', 'Model', 'Date Out', 'Destiny'])
        for r in rows:
            ws.append([r['id'], r['serial'], r['model'],
                       r['date_out'].strftime('%Y-%m-%d') if r['date_out'] else '', r['destiny']])
    wb.save(response)
    return response


@login_required
def frm_under_repair_view(request):
    user = request.user

    if request.method == 'POST' and request.POST.get('action') == 'receive':
        if _is_reader(user):
            messages.error(request, "You have a reader profile and can't modify data.")
            return redirect('frm_under_repair')

        repair_id = request.POST.get('id', '').strip()
        try:
            value = float(request.POST.get('value', '0').replace(',', '.'))
        except ValueError:
            messages.error(request, "You have entered a non-numeric value.")
            return redirect('frm_under_repair')

        ur = OeesUnderRepair.objects.filter(id_under_repair=repair_id, date_in__isnull=True).first()
        if not ur:
            messages.error(request, "Repair record not found or already received.")
            return redirect('frm_under_repair')
        now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        try:
            ur.date_in = datetime.now().date()
            ur.value = value
            ur.notes = f"{now} - Received from maintenance by {user.username}\n{ur.notes or ''}".strip()
            ur.save()
            messages.success(request, "Device received from maintenance.")
        except Exception:
            logger.exception("Error receiving repair %s", repair_id)
            messages.error(request, "An error occurred while receiving the device.")
        return redirect('frm_under_repair')

    # ---- GET ----
    if request.GET.get('export') == 'excel':
        return _under_repair_export_excel(request.GET.get('tab', 'pending'))

    repaired_rows = _under_repair_rows(repaired=True)
    repaired_total = sum(float(r['value'] or 0) for r in repaired_rows)
    context = {
        'pending_rows': _under_repair_rows(repaired=False),
        'repaired_rows': repaired_rows,
        'repaired_total': repaired_total,
        'is_reader': _is_reader(user),
    }
    return render(request, 'oe_inventory_py_web/frmUnderRepair.html', context)


# ==========================================================================
# Distribution Invoices screen (web port of frmDistributionInvoices.py)
# ==========================================================================

def _dist_invoice_rows(bill, subtotals):
    base = (
        "SELECT a.serial_number, {item} AS item, a.value, b.name AS user_name, "
        "d.delegation, b.department, c.name AS company "
        "FROM {table} a "
        "LEFT JOIN oees_staff b ON a.persone = b.id_staff "
        "LEFT JOIN oees_companies c ON b.company = c.id_company "
        "LEFT JOIN oees_delegations d ON b.delegation = d.id_delegation "
        "WHERE a.bill_number = %s"
    )
    queries = [
        base.format(item='a.model', table='oees_devices'),
        base.format(item='a.type', table='oees_licenses'),
        base.format(item='a.type', table='oees_mobile_phones'),
    ]

    grouped = {}
    with connection.cursor() as cur:
        for sql in queries:
            cur.execute(sql, [bill])
            for row in dict_fetchall(cur):
                company = row['company'] or ''
                delegation = row['delegation'] or ''
                department = row['department'] or ''
                (grouped.setdefault(company, {})
                        .setdefault(delegation, {})
                        .setdefault(department, [])
                        .append(row))

    rows = []
    for company in sorted(grouped):
        company_total = 0.0
        for delegation in sorted(grouped[company]):
            delegation_total = 0.0
            for department in sorted(grouped[company][delegation]):
                dept_total = 0.0
                for item in grouped[company][delegation][department]:
                    try:
                        value = float(item['value'] or 0)
                    except (TypeError, ValueError):
                        value = 0.0
                    dept_total += value
                    delegation_total += value
                    company_total += value
                    rows.append({
                        'company': company, 'delegation': delegation, 'department': department,
                        'user': item['user_name'] or '', 'serial': item['serial_number'] or '',
                        'model': item['item'] or '', 'value': f"{value:,.2f}", 'sub': False,
                    })
                if subtotals:
                    rows.append({'company': company, 'delegation': delegation,
                                 'department': f"Total {department}", 'user': '', 'serial': '',
                                 'model': '', 'value': f"{dept_total:,.2f}", 'sub': True})
            if subtotals:
                rows.append({'company': company, 'delegation': f"Total {delegation}",
                             'department': '', 'user': '', 'serial': '', 'model': '',
                             'value': f"{delegation_total:,.2f}", 'sub': True})
        if subtotals:
            rows.append({'company': f"Total {company}", 'delegation': '', 'department': '',
                         'user': '', 'serial': '', 'model': '', 'value': f"{company_total:,.2f}", 'sub': True})
    return rows


@login_required
def frm_dist_invoices_view(request):
    bill = request.GET.get('bill', '').strip()
    subtotals = request.GET.get('subtotals') == 'on'
    rows = _dist_invoice_rows(bill, subtotals) if bill else []

    if request.GET.get('export') == 'excel' and bill:
        if _is_reader(request.user):
            messages.error(request, "You have a reader profile and can't export data.")
            return redirect('frm_dist_invoices')
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="invoice_{bill}.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Distribution"
        ws.append(['Company', 'Delegation', 'Department', 'User', 'Serial Number', 'Model', 'Value'])
        for r in rows:
            ws.append([r['company'], r['delegation'], r['department'], r['user'],
                       r['serial'], r['model'], r['value']])
        wb.save(response)
        return response

    context = {
        'bill': bill,
        'subtotals': subtotals,
        'rows': rows,
        'searched': bool(bill),
    }
    return render(request, 'oe_inventory_py_web/frmDistributionInvoices.html', context)


# ==========================================================================
# Password Change screen (web port of frmPasswordChange.py)
# ==========================================================================

@login_required
def frm_password_change_view(request):
    """Let the logged-in user change their own password (Django-managed auth)."""
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            # Keep the user logged in after changing the password.
            update_session_auth_hash(request, user)
            messages.success(request, "Password changed successfully.")
            return redirect('frm_password_change')
        for field_errors in form.errors.values():
            for error in field_errors:
                messages.error(request, error)

    return render(request, 'oe_inventory_py_web/frmPasswordChange.html')


# ==========================================================================
# Delegations screen (web port of frmDelegations.py)
# ==========================================================================

def _geocode_delegation(direccion, cpostal, poblacion, provincia_name):
    """Geocode a delegation address to (lat, lng) via OpenStreetMap Nominatim.

    Uses the standard library only. Returns (None, None) on any failure so a
    save never breaks if the address can't be resolved or there's no internet.
    """
    import json as _json
    import urllib.parse
    import urllib.request

    if not (direccion or poblacion or cpostal):
        return None, None
    parts = [p for p in (direccion, cpostal, poblacion, provincia_name, 'España') if p]
    url = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode({
        'q': ", ".join(parts), 'format': 'json', 'limit': '1', 'countrycodes': 'es',
    })
    try:
        req = urllib.request.Request(
            url, headers={'User-Agent': 'OE-Inventory/1.0 (internal asset management app)'})
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = _json.loads(resp.read().decode('utf-8'))
        if data:
            return float(data[0]['lat']), float(data[0]['lon'])
    except Exception:
        logger.warning("Geocoding failed for delegation address: %r", ", ".join(parts))
    return None, None


@login_required
def frm_delegations_view(request):
    user = request.user

    # Force re-geocoding of the loaded delegation using the address on screen.
    if request.method == 'POST' and request.POST.get('action') == 'geocode':
        if _is_reader(user):
            messages.error(request, "You have a reader profile and can't modify data.")
            return redirect('frm_delegations')
        code = request.POST.get('code', '').strip()
        deleg = OeesDelegations.objects.filter(id_delegation=code).first() if code.isdigit() else None
        if not deleg:
            messages.error(request, "Load and save a delegation first, then geolocate it.")
            return redirect('frm_delegations')
        prov_raw = request.POST.get('provincia', '').strip()
        prov_id = int(prov_raw) if prov_raw.isdigit() else None
        prov_name = (OeesProvinces.objects.filter(id_province=prov_id)
                     .values_list('province', flat=True).first()) if prov_id else (
            deleg.provincia.province if deleg.provincia else '')
        lat, lng = _geocode_delegation(
            request.POST.get('direccion', '').strip() or (deleg.direccion or ''),
            request.POST.get('cpostal', '').strip() or (deleg.cpostal or ''),
            request.POST.get('poblacion', '').strip() or (deleg.poblacion or ''),
            prov_name or '',
        )
        if lat is not None:
            deleg.latitude, deleg.longitude = lat, lng
            deleg.save(update_fields=['latitude', 'longitude'])
            messages.success(request, f"Delegation located on the map ({lat:.5f}, {lng:.5f}).")
        else:
            messages.warning(request, "Could not locate this address. Check it and try again.")
        return redirect(f"{reverse('frm_delegations')}?delegation={code}")

    if request.method == 'POST' and request.POST.get('action') == 'save':
        if _is_reader(user):
            messages.error(request, "You have a reader profile and can't modify data.")
            return redirect('frm_delegations')
        code = request.POST.get('code', '').strip()
        name = request.POST.get('delegation', '').strip()
        if not name:
            messages.error(request, "The Delegation field is required.")
            return redirect('frm_delegations')

        prov_raw = request.POST.get('provincia', '').strip()
        prov_id = int(prov_raw) if prov_raw.isdigit() else None
        fields = {
            'direccion': request.POST.get('direccion', '').strip(),
            'cpostal': request.POST.get('cpostal', '').strip(),
            'poblacion': request.POST.get('poblacion', '').strip(),
        }
        # Geocode the address so it can be pinned on the map. On failure we keep
        # whatever coordinates the delegation already had (don't wipe them).
        prov_name = (OeesProvinces.objects.filter(id_province=prov_id)
                     .values_list('province', flat=True).first()) if prov_id else ''
        lat, lng = _geocode_delegation(fields['direccion'], fields['cpostal'],
                                       fields['poblacion'], prov_name or '')
        now = datetime.now().strftime('%d/%m/%Y %H:%M')
        try:
            existing = OeesDelegations.objects.filter(id_delegation=code).first() if code.isdigit() else None
            if existing:
                prev = request.POST.get('notes', '').strip() or (existing.notes or '')
                existing.delegation = name
                for key, val in fields.items():
                    setattr(existing, key, val)
                existing.provincia_id = prov_id
                if lat is not None:
                    existing.latitude, existing.longitude = lat, lng
                existing.notes = f"{now} - Modified by {user.username}\n{prev}".strip()
                existing.save()
            else:
                OeesDelegations.objects.create(
                    delegation=name, notes=f"{now} - Created by {user.username}",
                    provincia_id=prov_id, latitude=lat, longitude=lng, **fields,
                )
            messages.success(request, "Delegation saved successfully.")
        except Exception:
            logger.exception("Error saving delegation %s", code)
            messages.error(request, "An error occurred while saving the delegation.")
        return redirect('frm_delegations')

    # ---- GET ----
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="delegations.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Delegations"
        ws.append(['ID', 'Delegation', 'Address', 'Post Code', 'Town', 'Province'])
        for d in OeesDelegations.objects.select_related('provincia').order_by('id_delegation'):
            ws.append([d.id_delegation, d.delegation, d.direccion or '', d.cpostal or '',
                       d.poblacion or '', d.provincia.province if d.provincia else ''])
        wb.save(response)
        return response

    # Points for the map: only delegations that already have coordinates.
    map_points = [
        {
            'name': d.delegation,
            'lat': d.latitude,
            'lng': d.longitude,
            'activo': d.activo,
            'address': ", ".join([p for p in [d.direccion, d.poblacion] if p]),
        }
        for d in OeesDelegations.objects.filter(latitude__isnull=False, longitude__isnull=False)
    ]

    context = {
        'grid_rows': OeesDelegations.objects.select_related('provincia').order_by('id_delegation'),
        'provinces': OeesProvinces.objects.all().order_by('province'),
        'preselected_id': request.GET.get('delegation', '').strip(),
        'is_reader': _is_reader(user),
        'map_points': map_points,
    }
    return render(request, 'oe_inventory_py_web/frmDelegations.html', context)


@login_required
def api_get_delegation(request):
    """AJAX API: look up a delegation by id to fill the form."""
    did = request.GET.get('id', '').strip()
    if not did.isdigit():
        return JsonResponse({'success': False, 'error': 'No delegation id provided.'}, status=400)
    d = OeesDelegations.objects.filter(id_delegation=did).first()
    if not d:
        return JsonResponse({'success': True, 'exists': False, 'id': did})
    data = {
        'id': d.id_delegation, 'delegation': d.delegation or '', 'direccion': d.direccion or '',
        'cpostal': d.cpostal or '', 'poblacion': d.poblacion or '', 'provincia': d.provincia_id or '',
        'notes': d.notes or '',
    }
    return JsonResponse({'success': True, 'exists': True, 'data': data})


# ==========================================================================
# Access Cards screen (web port of frmAccessCards.py)
# ==========================================================================

def _access_cards_rows():
    sql = (
        "SELECT a.id_card, a.ac_max, a.fermax_mif, a.pin_card, b.name AS staff_name, "
        "c.description AS state_desc, a.obs "
        "FROM oees_access_cards a "
        "LEFT JOIN oees_staff b ON a.id_staff = b.id_staff "
        "LEFT JOIN oees_access_cards_states c ON a.state_card = c.id_state "
        "ORDER BY a.id_card DESC"
    )
    with connection.cursor() as cur:
        cur.execute(sql)
        raw = dict_fetchall(cur)
    return [{
        'id': r['id_card'], 'card': r['ac_max'], 'fermax': r['fermax_mif'] or '',
        'pin': r['pin_card'] or '', 'staff': r['staff_name'] or '', 'state': r['state_desc'] or '',
        'obs': r['obs'] or '', 'lost': (r['state_desc'] or '').upper() == 'LOST',
    } for r in raw]


@login_required
def frm_access_cards_view(request):
    user = request.user

    if request.method == 'POST':
        if _is_reader(user):
            messages.error(request, "You have a reader profile and can't modify data.")
            return redirect('frm_access_cards')
        action = request.POST.get('action', '')
        if action == 'save':
            return _card_save(request)
        if action == 'release':
            return _card_release(request)
        return redirect('frm_access_cards')

    if request.GET.get('export') == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="access_cards.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Access Cards"
        ws.append(['ID', 'Card', 'Fermax MIF', 'PIN', 'Staff', 'State', 'Obs'])
        for r in _access_cards_rows():
            ws.append([r['id'], r['card'], r['fermax'], r['pin'], r['staff'], r['state'], r['obs']])
        wb.save(response)
        return response

    context = {
        'grid_rows': _access_cards_rows(),
        'staff': list(OeesStaff.objects.filter(state=1).order_by('name').values('id_staff', 'name')),
        'states': list(OeesAccessCardsStates.objects.order_by('description').values('id_state', 'description')),
        'preselected_card': request.GET.get('card', '').strip(),
        'is_reader': _is_reader(user),
    }
    return render(request, 'oe_inventory_py_web/frmAccessCards.html', context)


def _card_save(request):
    user = request.user.username
    code = request.POST.get('card', '').strip()
    fermax = request.POST.get('fermax', '').strip()
    if not code or not fermax:
        messages.error(request, "Card code and Fermax MIF are required.")
        return redirect('frm_access_cards')

    pin = request.POST.get('pin', '').strip()
    obs = request.POST.get('obs', '').strip()
    staff_id = request.POST.get('staff') or None
    state_id = request.POST.get('state') or None

    now = datetime.now().strftime('%d/%m/%Y %H:%M')
    try:
        card = OeesAccessCards.objects.filter(ac_max=code).first()
        if card:
            # Block edits only if the card is ALREADY lost (check its CURRENT
            # state, not the new one) — marking a card AS lost must be allowed.
            current = OeesAccessCardsStates.objects.filter(id_state=card.state_card).first()
            if current and (current.description or '').strip().upper() == 'LOST':
                messages.error(request, "This card is in LOST state and can't be modified.")
                return redirect(f"{reverse('frm_access_cards')}?card={code}")
            prev = request.POST.get('notes', '').strip() or (card.notes or '')
            card.fermax_mif = fermax
            card.pin_card = pin
            card.state_card = int(state_id) if state_id else 0
            card.obs = obs
            card.id_staff_id = staff_id
            card.notes = f"{now} - Modified by {user}\n{prev}".strip()
            card.save()
        else:
            OeesAccessCards.objects.create(
                ac_max=code, fermax_mif=fermax, pin_card=pin,
                state_card=int(state_id) if state_id else 0, obs=obs, id_staff_id=staff_id,
                notes=f"{now} - Created by {user}",
            )
        # The PIN is now in use: remove it from the available-PINs pool so it
        # isn't offered again for the next card.
        if pin:
            OeesAccessCardsPins.objects.filter(pin=pin).delete()
        messages.success(request, "Card saved successfully.")
    except Exception:
        logger.exception("Error saving access card %s", code)
        messages.error(request, "An error occurred while saving the card.")
    return redirect(f"{reverse('frm_access_cards')}?card={code}")


def _card_release(request):
    user = request.user.username
    code = request.POST.get('card', '').strip()
    card = OeesAccessCards.objects.filter(ac_max=code).first()
    if not card:
        messages.error(request, "Card not found.")
        return redirect('frm_access_cards')
    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    note = f"{now} - Converted to a VISITOR CARD by {user}"
    try:
        visitor = OeesAccessCardsVisitors.objects.create(
            name='', ac_max=card.ac_max, fermax_mif=card.fermax_mif,
            state_card_id=1, obs=card.obs or '', notes=note, pin_card='',
        )
        OeesAccessCardsVisitorsNotes.objects.create(id_visitors_card_id=visitor.id_card, notes=note)
        card.delete()
        messages.success(request, "Card converted to a visitor card.")
    except Exception:
        logger.exception("Error converting card %s to visitor", code)
        messages.error(request, "An error occurred while converting the card.")
    return redirect('frm_access_cards')


@login_required
def api_get_card(request):
    code = request.GET.get('card', '').strip()
    if not code:
        return JsonResponse({'success': False, 'error': 'No card code provided.'}, status=400)
    with connection.cursor() as cur:
        cur.execute(
            "SELECT a.fermax_mif, a.pin_card, a.obs, a.notes, a.id_staff, a.state_card, "
            "b.name AS staff_name, c.description AS state_desc "
            "FROM oees_access_cards a "
            "LEFT JOIN oees_staff b ON a.id_staff = b.id_staff "
            "LEFT JOIN oees_access_cards_states c ON a.state_card = c.id_state "
            "WHERE a.ac_max = %s", [code])
        row = dict_fetchall(cur)
    if not row:
        return JsonResponse({'success': True, 'exists': False, 'card': code})
    r = row[0]
    data = {
        'card': code, 'fermax': r['fermax_mif'] or '', 'pin': r['pin_card'] or '',
        'obs': r['obs'] or '', 'notes': r['notes'] or '', 'staff_id': r['id_staff'] or '',
        'state_id': r['state_card'] or '', 'staff_name': r['staff_name'] or '',
        'state': r['state_desc'] or '',
    }
    return JsonResponse({'success': True, 'exists': True, 'data': data})


@login_required
def api_card_pin(request):
    # Pick a RANDOM available PIN (so they aren't handed out consecutively).
    pin = OeesAccessCardsPins.objects.order_by('?').values_list('pin', flat=True).first()
    return JsonResponse({'success': True, 'pin': pin or ''})


# ==========================================================================
# Access Keys screen (web port of frmAccessKeys.py)
# ==========================================================================

@login_required
def frm_access_keys_view(request):
    user = request.user

    if request.method == 'POST' and request.POST.get('action') == 'save':
        if _is_reader(user):
            messages.error(request, "You have a reader profile and can't modify data.")
            return redirect('frm_access_keys')
        return _key_save(request)

    if request.GET.get('export') == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="access_keys.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Access Keys"
        ws.append(['ID', 'Company', 'Type', 'Staff', 'Insert Date'])
        for k in (OeesAccessKeys.objects.select_related('id_company', 'id_staff').order_by('-id_key')):
            ws.append([k.id_key, k.id_company.name if k.id_company else '', k.type or '',
                       k.id_staff.name if k.id_staff else '',
                       k.insert_date.strftime('%Y-%m-%d') if k.insert_date else ''])
        wb.save(response)
        return response

    with connection.cursor() as cur:
        cur.execute("SELECT DISTINCT type FROM oees_access_keys WHERE type IS NOT NULL AND type <> '' ORDER BY type")
        types = [row[0] for row in cur.fetchall()]
        cur.execute(
            "SELECT a.id_key, a.type, a.insert_date, c.name AS company_name, s.name AS staff_name "
            "FROM oees_access_keys a "
            "LEFT JOIN oees_companies c ON a.id_company = c.id_company "
            "LEFT JOIN oees_staff s ON a.id_staff = s.id_staff "
            "ORDER BY a.id_key DESC")
        grid = dict_fetchall(cur)

    context = {
        'grid_rows': grid,
        'companies': OeesCompanies.objects.all().order_by('name'),
        'staff': list(OeesStaff.objects.all().order_by('name').values('id_staff', 'name')),
        'types': types,
        'preselected_id': request.GET.get('key', '').strip(),
        'is_reader': _is_reader(user),
    }
    return render(request, 'oe_inventory_py_web/frmAccessKeys.html', context)


def _key_save(request):
    user = request.user.username
    code = request.POST.get('code', '').strip()
    key_type = request.POST.get('type', '').strip()
    if not key_type:
        messages.error(request, "You have to indicate a valid Type.")
        return redirect('frm_access_keys')
    try:
        insert_date = datetime.strptime(request.POST.get('insert_date', '').strip(), '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, "A valid Insert Date is required.")
        return redirect('frm_access_keys')

    company_id = request.POST.get('company') or None
    staff_id = request.POST.get('staff') or None
    now = datetime.now().strftime('%d/%m/%Y %H:%M')

    try:
        key = OeesAccessKeys.objects.filter(id_key=code).first() if code.isdigit() else None
        if key:
            prev = request.POST.get('notes', '').strip() or (key.notes or '')
            key.type = key_type
            key.insert_date = insert_date
            key.id_company_id = company_id
            key.id_staff_id = staff_id
            key.notes = f"{now} - Modified by {user}\n{prev}".strip()
            key.save()
        else:
            OeesAccessKeys.objects.create(
                type=key_type, insert_date=insert_date, id_company_id=company_id,
                id_staff_id=staff_id, notes=f"{now} - Created by {user}",
            )
        messages.success(request, "Access key saved successfully.")
    except Exception:
        logger.exception("Error saving access key %s", code)
        messages.error(request, "An error occurred while saving the access key.")
    return redirect('frm_access_keys')


@login_required
def api_get_key(request):
    kid = request.GET.get('id', '').strip()
    if not kid.isdigit():
        return JsonResponse({'success': False, 'error': 'No key id provided.'}, status=400)
    with connection.cursor() as cur:
        cur.execute(
            "SELECT a.id_key, a.type, a.insert_date, a.notes, a.id_company, a.id_staff, "
            "c.name AS company_name, s.name AS staff_name "
            "FROM oees_access_keys a "
            "LEFT JOIN oees_companies c ON a.id_company = c.id_company "
            "LEFT JOIN oees_staff s ON a.id_staff = s.id_staff "
            "WHERE a.id_key = %s", [kid])
        row = dict_fetchall(cur)
    if not row:
        return JsonResponse({'success': True, 'exists': False, 'id': kid})
    r = row[0]
    date_value = r['insert_date']
    data = {
        'id': r['id_key'], 'type': r['type'] or '',
        'date': date_value.strftime('%Y-%m-%d') if hasattr(date_value, 'strftime') else (date_value or ''),
        'company_id': r['id_company'] or '', 'staff_id': r['id_staff'] or '',
        'company_name': r['company_name'] or '', 'staff_name': r['staff_name'] or '',
        'notes': r['notes'] or '',
    }
    return JsonResponse({'success': True, 'exists': True, 'data': data})


# ==========================================================================
# Visitors Access Cards screen (web port of frmVisitorsAccessCards.py)
# ==========================================================================

def _visitor_cards_rows():
    sql = (
        "SELECT a.id_card, a.ac_max, a.fermax_mif, a.name, a.obs, c.description AS state_desc "
        "FROM oees_access_cards_visitors a "
        "LEFT JOIN oees_access_cards_states c ON a.state_card = c.id_state "
        "ORDER BY a.id_card DESC"
    )
    with connection.cursor() as cur:
        cur.execute(sql)
        raw = dict_fetchall(cur)
    return [{
        'id': r['id_card'], 'card': r['ac_max'], 'fermax': r['fermax_mif'] or '',
        'user': r['name'] or '', 'state': r['state_desc'] or '', 'obs': r['obs'] or '',
        'lost': (r['state_desc'] or '').upper() == 'LOST',
    } for r in raw]


@login_required
def frm_visitor_cards_view(request):
    user = request.user

    if request.method == 'POST' and request.POST.get('action') == 'save':
        if _is_reader(user):
            messages.error(request, "You have a reader profile and can't modify data.")
            return redirect('frm_visitor_cards')
        return _visitor_save(request)

    if request.GET.get('export') == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="visitor_cards.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visitor Cards"
        ws.append(['ID', 'Card', 'Fermax MIF', 'User', 'State', 'Obs'])
        for r in _visitor_cards_rows():
            ws.append([r['id'], r['card'], r['fermax'], r['user'], r['state'], r['obs']])
        wb.save(response)
        return response

    context = {
        'grid_rows': _visitor_cards_rows(),
        'states': list(OeesAccessCardsStates.objects.order_by('description').values('id_state', 'description')),
        'preselected_card': request.GET.get('card', '').strip(),
        'is_reader': _is_reader(user),
    }
    return render(request, 'oe_inventory_py_web/frmVisitorsAccessCards.html', context)


def _visitor_save(request):
    user = request.user.username
    code = request.POST.get('card', '').strip()
    fermax = request.POST.get('fermax', '').strip()
    if not code or not fermax:
        messages.error(request, "The Card and Fermax MIF fields are mandatory.")
        return redirect('frm_visitor_cards')

    name = request.POST.get('user', '').strip()
    obs = request.POST.get('obs', '').strip()
    state_id = request.POST.get('state') or None
    now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    state_desc = ''
    if state_id:
        st = OeesAccessCardsStates.objects.filter(id_state=state_id).first()
        state_desc = st.description if st else ''

    try:
        visitor = OeesAccessCardsVisitors.objects.filter(ac_max=code).first()
        if visitor:
            prev = visitor.notes or ''
            visitor.fermax_mif = fermax
            visitor.name = name
            visitor.state_card_id = state_id
            visitor.obs = obs
            visitor.notes = f"{now} - Modified by {user}\n{prev}".strip()
            visitor.save()
            short = f"{now} - Modified by {user}. {'Assigned to ' + name if name else 'No user assigned'}. State: {state_desc}"
        else:
            visitor = OeesAccessCardsVisitors.objects.create(
                ac_max=code, fermax_mif=fermax, name=name, state_card_id=state_id,
                obs=obs, pin_card='', notes=f"{now} - Saved by {user}",
            )
            short = f"{now} - Saved by {user}. {'Assigned to ' + name if name else 'No user assigned'}. State: {state_desc}"
        OeesAccessCardsVisitorsNotes.objects.create(id_visitors_card_id=visitor.id_card, notes=short)
        messages.success(request, "Visitor card saved successfully.")
    except Exception:
        logger.exception("Error saving visitor card %s", code)
        messages.error(request, "An error occurred while saving the visitor card.")
    return redirect(f"{reverse('frm_visitor_cards')}?card={code}")


@login_required
def api_get_visitor_card(request):
    code = request.GET.get('card', '').strip()
    if not code:
        return JsonResponse({'success': False, 'error': 'No card code provided.'}, status=400)
    with connection.cursor() as cur:
        cur.execute(
            "SELECT a.id_card, a.fermax_mif, a.name, a.obs, a.notes, a.state_card, c.description AS state_desc "
            "FROM oees_access_cards_visitors a "
            "LEFT JOIN oees_access_cards_states c ON a.state_card = c.id_state "
            "WHERE a.ac_max = %s", [code])
        row = dict_fetchall(cur)
    if not row:
        return JsonResponse({'success': True, 'exists': False, 'card': code})
    r = row[0]
    with connection.cursor() as cur:
        cur.execute("SELECT id_line, notes FROM oees_access_cards_visitors_notes "
                    "WHERE id_visitors_card = %s ORDER BY id_line DESC", [r['id_card']])
        notes = [{'id': nid, 'notes': txt} for nid, txt in cur.fetchall()]
    data = {
        'card': code, 'fermax': r['fermax_mif'] or '', 'user': r['name'] or '',
        'obs': r['obs'] or '', 'notes': r['notes'] or '', 'state_id': r['state_card'] or '',
        'state': r['state_desc'] or '', 'history': notes,
    }
    return JsonResponse({'success': True, 'exists': True, 'data': data})
