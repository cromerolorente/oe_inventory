from datetime import date

from django.contrib.auth import get_user_model
from django.contrib.messages import get_messages
from django.test import TestCase, override_settings
from django.urls import reverse

from .models import (
    OeesAccessCards, OeesAccessCardsStates, OeesAccessCardsVisitors, OeesAccessKeys,
    OeesCompanies, OeesDelegations, OeesDevices, OeesFiberLines, OeesFiberLinesIncidences,
    OeesIncorporations, OeesLicenses, OeesMobileLines, OeesMobilePhones, OeesOrders,
    OeesPrinters, OeesStaff, OeesUnderRepair,
)


class PublicPagesTests(TestCase):
    """Smoke tests that don't depend on pre-existing data."""

    def test_login_page_renders(self):
        response = self.client.get(reverse('login'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/login.html')

    def test_home_requires_login(self):
        response = self.client.get(reverse('mdi_home'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_devices_requires_login(self):
        response = self.client.get(reverse('frm_devices'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)


class AuthenticatedFlowTests(TestCase):
    """Tests covering an authenticated session."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(
            username='tester', password='pass12345', devices=1
        )

    def test_home_renders_when_logged_in(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('mdi_home'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/base_mdi.html')

    def test_api_get_user_never_returns_password_hash(self):
        """Security regression: the password hash must never reach the client."""
        response = self.client.get(reverse('api_get_user'), {'login': 'tester'})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['exists'])
        self.assertEqual(data['data']['password'], '********')
        self.assertNotIn(self.user.password, response.content.decode())


class StaffScreenTests(TestCase):
    """Smoke tests for the Staff screen."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='ituser', password='pass12345', staff=1)
        self.staff = OeesStaff.objects.create(
            name='Jane Tester', persona_fisica=1, notes='', state=1, department='IT'
        )

    def test_staff_requires_login(self):
        response = self.client.get(reverse('frm_staff'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_staff_page_renders(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_staff'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/frmStaff.html')

    def test_staff_list_shows_assigned_phone_number(self):
        from oe_inventory_py_web.models import OeesMobileLines, OeesMobilePhones
        line = OeesMobileLines.objects.create(
            number='600123456', imei='', pin='', puk='', pin2='', puk2='',
            extension='', esim=0, m2m=0, obs='')
        OeesMobilePhones.objects.create(
            serial_number='PH-1', type='MOBILE', value=0.0,
            persone=self.staff, id_line=line)
        self.client.force_login(self.user)
        resp = self.client.get(reverse('frm_staff'))
        self.assertContains(resp, '<th>Phone</th>')     # new column header
        self.assertContains(resp, '600123456')          # the assigned line number

    def test_api_get_staff_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_staff'), {'id': self.staff.id_staff})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['exists'])
        self.assertEqual(data['data']['name'], 'Jane Tester')

    def test_api_get_staff_not_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_staff'), {'id': 999999})
        self.assertFalse(response.json()['exists'])

    def test_staff_report_returns_pdf(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('staff_report'), {'staff': self.staff.id_staff})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')

    def test_release_generates_unassign_doc(self):
        import tempfile
        from datetime import date
        from django.test import override_settings
        from oe_inventory_py_web.models import OeesDevices, OeesDocs
        User = get_user_model()
        admin = User.objects.create_user(username='rel_u', password='pass12345', staff=1, reader=0)
        dev = OeesDevices.objects.create(
            serial_number='DEV-R1', type='LAPTOP', brand='Dell', model='X',
            origin='New', insert_date=date(2026, 1, 1), value=0.0, persone=self.staff)
        with tempfile.TemporaryDirectory() as tmp, override_settings(MEDIA_ROOT=tmp):
            self.client.force_login(admin)
            resp = self.client.post(reverse('frm_staff'), {
                'action': 'release', 'code': str(self.staff.id_staff),
                'item_id': f'D{dev.id_device}', 'serial': 'DEV-R1', 'name': self.staff.name,
            })
            self.assertEqual(resp.status_code, 302)
            dev.refresh_from_db()
            self.assertIsNone(dev.persone_id)              # item unassigned
            self.assertTrue(OeesDocs.objects.filter(       # Unassign document generated
                id_staff_id=self.staff.id_staff, doc_name__startswith='Unassign-').exists())

    def test_terminate_returns_items_and_generates_document(self):
        import tempfile
        from datetime import date
        from django.test import override_settings
        from oe_inventory_py_web.models import OeesDevices, OeesDocs
        User = get_user_model()
        admin = User.objects.create_user(username='term_u', password='pass12345', staff=1, reader=0)
        dev = OeesDevices.objects.create(
            serial_number='DEV-T1', type='LAPTOP', brand='Dell', model='X',
            origin='New', insert_date=date(2026, 1, 1), value=0.0, persone=self.staff)

        with tempfile.TemporaryDirectory() as tmp, override_settings(MEDIA_ROOT=tmp):
            self.client.force_login(admin)
            resp = self.client.post(reverse('frm_staff'), {
                'action': 'terminate', 'code': str(self.staff.id_staff),
                'returned_items': f'D{dev.id_device}',
            })
            self.assertEqual(resp.status_code, 302)

            dev.refresh_from_db()
            self.assertIsNone(dev.persone_id)              # item freed back to stock
            self.staff.refresh_from_db()
            self.assertEqual(self.staff.state, 0)          # contract terminated
            self.assertTrue(self.staff.fecha_baja)
            self.assertTrue(OeesDocs.objects.filter(       # Terminate document saved
                id_staff_id=self.staff.id_staff, doc_name='Terminate').exists())

    def test_scope_staff_applies_optional_and_filters(self):
        # The three scopes are optional and combined with AND (each one narrows
        # further only when set). A user with no scope sees everyone; with a
        # company AND a department set, only staff matching BOTH are shown.
        from .views import scope_staff_queryset
        from .models import OeesCompanies

        comp = OeesCompanies.objects.create(name='ACME')
        # self.staff = Jane, department='IT', no company.
        OeesStaff.objects.create(name='MatchBoth', persona_fisica=1, notes='', state=1,
                                 department='IT', company_id=comp.id_company)
        OeesStaff.objects.create(name='OnlyCompany', persona_fisica=1, notes='', state=1,
                                 department='SALES', company_id=comp.id_company)
        User = get_user_model()

        # No scope -> sees everyone.
        no_scope = User.objects.create_user(username='no_scope', password='pass12345')
        self.assertEqual(scope_staff_queryset(no_scope, OeesStaff.objects.all()).count(), 3)

        # Company AND department set -> only staff matching both.
        scoped = User.objects.create_user(
            username='scoped_user', password='pass12345',
            departments='IT', companies=str(comp.id_company),
        )
        names = set(
            scope_staff_queryset(scoped, OeesStaff.objects.all()).values_list('name', flat=True)
        )
        self.assertEqual(names, {'MatchBoth'})  # Jane (no company) and OnlyCompany (SALES) excluded

    def test_scope_staff_without_company_uses_other_dimensions(self):
        # A user may have NO company assigned but still have delegations and/or
        # departments: the missing dimension simply doesn't filter, the others do.
        from .views import scope_staff_queryset
        from .models import OeesDelegations

        deleg = OeesDelegations.objects.create(delegation='MADRID', notes='')
        OeesStaff.objects.create(name='DelegDept', persona_fisica=1, notes='', state=1,
                                 department='IT', delegation_id=deleg.id_delegation)
        OeesStaff.objects.create(name='SameDelegOtherDept', persona_fisica=1, notes='', state=1,
                                 department='HR', delegation_id=deleg.id_delegation)
        User = get_user_model()
        u = User.objects.create_user(
            username='no_company_user', password='pass12345',
            delegations=str(deleg.id_delegation), departments='IT',  # no companies
        )
        names = set(
            scope_staff_queryset(u, OeesStaff.objects.all()).values_list('name', flat=True)
        )
        self.assertEqual(names, {'DelegDept'})  # delegation AND department; company not required


class LicensesScreenTests(TestCase):
    """Smoke tests for the Licenses screen."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='lic_user', password='pass12345', licenses=1, reader=0)
        self.lic = OeesLicenses.objects.create(serial_number='LIC-1', type='Office', value=99.0, obs='')

    def test_licenses_requires_login(self):
        response = self.client.get(reverse('frm_licenses'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_licenses_page_renders(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_licenses'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/frmLicenses.html')

    def test_license_summary_by_type(self):
        from oe_inventory_py_web.models import OeesStaff
        expired = OeesStaff.objects.create(name='LICENCIAS CADUCADAS', persona_fisica=1, notes='', state=1)
        # self.lic = LIC-1 type 'Office' (unassigned). Add one Office assigned to
        # the expired person, and one of another type.
        OeesLicenses.objects.create(serial_number='LIC-2', type='Office', value=0, obs='', persone=expired)
        OeesLicenses.objects.create(serial_number='LIC-3', type='Visio', value=0, obs='')

        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_licenses'))
        summary = {s['type']: s for s in response.context['license_summary']}

        self.assertEqual(summary['Office']['purchased'], 2)
        self.assertEqual(summary['Office']['expired'], 1)
        self.assertEqual(summary['Office']['in_use'], 1)
        self.assertEqual(summary['Visio']['purchased'], 1)
        self.assertEqual(summary['Visio']['expired'], 0)
        self.assertEqual(summary['Visio']['in_use'], 1)

    def test_api_get_license_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_license'), {'serial_number': 'LIC-1'})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['exists'])
        self.assertEqual(data['data']['type'], 'Office')

    def test_api_get_license_not_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_license'), {'serial_number': 'NOPE'})
        self.assertFalse(response.json()['exists'])

    def test_save_creates_license(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_licenses'), {
            'action': 'save', 'serial_number': 'LIC-NEW', 'type': 'CAD',
            'origin': 'Vendor', 'value': '150.00', 'obs': 'test', 'insert_date': '2026-01-10',
        })
        self.assertEqual(response.status_code, 302)
        self.assertTrue(OeesLicenses.objects.filter(serial_number='LIC-NEW').exists())


class PhonesScreenTests(TestCase):
    """Smoke tests for the Mobile Phones screen."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='ph_user', password='pass12345', phones=1, reader=0)
        self.phone = OeesMobilePhones.objects.create(serial_number='PH-1', type='', value=50.0, brand='Apple')

    def test_phones_requires_login(self):
        response = self.client.get(reverse('frm_phones'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_phones_page_renders(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_phones'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/frmPhones.html')

    def test_excel_export_includes_active_column(self):
        import io
        import openpyxl
        from oe_inventory_py_web.models import OeesStaff, OeesMobilePhones
        p = OeesStaff.objects.create(name='Active P', persona_fisica=1, notes='', state=1)
        OeesMobilePhones.objects.create(serial_number='PH-X', type='', value=0.0, persone=p)
        self.client.force_login(self.user)
        resp = self.client.get(reverse('frm_phones'), {'export': 'excel'})
        ws = openpyxl.load_workbook(io.BytesIO(resp.content)).active
        header = [c.value for c in ws[1]]
        self.assertIn('Active', header)
        self.assertEqual(header.index('Active'), header.index('Person') + 1)  # after Person
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        px = next(r for r in rows if r[0] == 'PH-X')
        self.assertEqual(px[header.index('Active')], 'Yes')

    def test_total_phones_assigned_counts_phones_with_a_line(self):
        from oe_inventory_py_web.models import OeesMobileLines, OeesMobilePhones
        line = OeesMobileLines.objects.create(
            number='600000001', imei='', pin='', puk='', pin2='', puk2='',
            extension='', esim=0, m2m=0, obs='')
        OeesMobilePhones.objects.create(serial_number='PH-WL', type='', value=0.0, id_line=line)
        OeesMobilePhones.objects.create(serial_number='PH-NL', type='', value=0.0)
        self.client.force_login(self.user)
        resp = self.client.get(reverse('frm_phones'))
        self.assertEqual(resp.context['total_phones'], 3)            # PH-1, PH-WL, PH-NL
        self.assertEqual(resp.context['total_phones_assigned'], 1)   # only PH-WL has a line
        self.assertContains(resp, 'Total Phones Assigned')

    def test_grid_active_column_reflects_person_state(self):
        from oe_inventory_py_web.models import OeesStaff, OeesMobilePhones
        active_p = OeesStaff.objects.create(name='Active P', persona_fisica=1, notes='', state=1)
        inactive_p = OeesStaff.objects.create(name='Inactive P', persona_fisica=1, notes='', state=0)
        OeesMobilePhones.objects.create(serial_number='PH-A', type='', value=0.0, persone=active_p)
        OeesMobilePhones.objects.create(serial_number='PH-I', type='', value=0.0, persone=inactive_p)
        self.client.force_login(self.user)
        resp = self.client.get(reverse('frm_phones'))
        by_serial = {g['serial']: g for g in resp.context['grid_data']}
        self.assertTrue(by_serial['PH-A']['active'])     # person state == 1
        self.assertFalse(by_serial['PH-I']['active'])    # person state == 0
        self.assertFalse(by_serial['PH-1']['active'])    # no person assigned
        self.assertContains(resp, '<th class="text-center">Active</th>')

    def test_phone_release_generates_unassign_doc(self):
        import tempfile
        from django.test import override_settings
        from oe_inventory_py_web.models import OeesStaff, OeesDocs
        owner = OeesStaff.objects.create(name='Phone Owner', persona_fisica=1, notes='', state=1)
        self.phone.persone = owner
        self.phone.save(update_fields=['persone'])
        self.client.force_login(self.user)
        with tempfile.TemporaryDirectory() as tmp, override_settings(MEDIA_ROOT=tmp):
            resp = self.client.post(reverse('frm_phones'), {
                'action': 'release', 'serial_number': 'PH-1', 'person': 'Phone Owner',
            })
            self.assertEqual(resp.status_code, 302)
            self.phone.refresh_from_db()
            self.assertIsNone(self.phone.persone_id)        # phone unassigned
            self.assertTrue(OeesDocs.objects.filter(        # Unassign document generated
                id_staff_id=owner.id_staff, doc_name__startswith='Unassign-').exists())

    def test_api_get_phone_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_phone'), {'serial_number': 'PH-1'})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['exists'])
        self.assertEqual(data['data']['brand'], 'Apple')

    def test_api_get_phone_not_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_phone'), {'serial_number': 'NOPE'})
        self.assertFalse(response.json()['exists'])

    def test_save_creates_phone(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_phones'), {
            'action': 'save', 'serial_number': 'PH-NEW', 'brand': 'Samsung',
            'model': 'S24', 'origin': 'Vendor', 'value': '300.00', 'imei': '123', 'insert_date': '2026-02-01',
        })
        self.assertEqual(response.status_code, 302)
        self.assertTrue(OeesMobilePhones.objects.filter(serial_number='PH-NEW').exists())

    def test_support_creates_under_repair(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_phones'), {'action': 'support', 'serial_number': 'PH-1'})
        self.assertEqual(response.status_code, 302)
        from .models import OeesUnderRepair
        self.assertTrue(OeesUnderRepair.objects.filter(serial_number='PH-1', type='M', date_in__isnull=True).exists())


class FiberScreenTests(TestCase):
    """Smoke tests for the Fiber Lines screen."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='fb_user', password='pass12345', fiber_lines=1, reader=0)
        self.fiber = OeesFiberLines.objects.create(
            description='Fiber A', proveedor='ISP', orden='O1', codigo_servicio='SC1',
            acceso='acc', router='rt', direccionamiento='addr', wifi1='w1', wifi2='w2',
            estado=1, fecha_inicio=date(2026, 1, 1), fee=29.95, notes='',
        )

    def test_fiber_requires_login(self):
        response = self.client.get(reverse('frm_fiber'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_fiber_page_renders(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_fiber'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/frmFiberLines.html')

    def test_api_get_fiber_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_fiber'), {'id': self.fiber.id_fiber_line})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['exists'])
        self.assertEqual(data['data']['description'], 'Fiber A')
        self.assertEqual(data['data']['fee'], 29.95)

    def test_api_get_fiber_not_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_fiber'), {'id': 999999})
        self.assertFalse(response.json()['exists'])

    def test_save_creates_fiber(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_fiber'), {
            'action': 'save', 'description': 'New Fiber', 'provider': 'ISP2', 'order': 'O2',
            'service_code': 'SC2', 'access': 'a', 'router': 'r', 'addressing': 'd',
            'wifi1': '', 'wifi2': '', 'estado': '1', 'start_date': '2026-03-01', 'ip_fixed': '',
            'fee': '19,90',
        })
        self.assertEqual(response.status_code, 302)
        saved = OeesFiberLines.objects.get(description='New Fiber')
        self.assertEqual(saved.fee, 19.90)

    def test_save_incidence(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_fiber'), {
            'action': 'save_incidence', 'code': str(self.fiber.id_fiber_line), 'working_code': 'WC1',
            'open_date': '2026-02-01', 'open_description': 'opened',
        })
        self.assertEqual(response.status_code, 302)
        self.assertTrue(OeesFiberLinesIncidences.objects.filter(
            id_fiber_line_id=self.fiber.id_fiber_line, working_code='WC1').exists())


class PrintersScreenTests(TestCase):
    """Smoke tests for the Printers screen."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='pr_user', password='pass12345', printers=1, reader=0)
        self.printer = OeesPrinters.objects.create(serial_number='PR-1', description='HP LaserJet', ip='10.0.0.5')

    def test_printers_requires_login(self):
        response = self.client.get(reverse('frm_printers'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_printers_page_renders(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_printers'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/frmPrinters.html')

    def test_api_get_printer_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_printer'), {'serial_number': 'PR-1'})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['exists'])
        self.assertEqual(data['data']['description'], 'HP LaserJet')

    def test_api_get_printer_not_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_printer'), {'serial_number': 'NOPE'})
        self.assertFalse(response.json()['exists'])

    def test_save_creates_printer(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_printers'), {
            'action': 'save', 'serial_number': 'PR-NEW', 'description': 'Canon',
            'provider': 'Vendor', 'ip': '10.0.0.9', 'start_date': '2026-04-01', 'fee': '12.50',
        })
        self.assertEqual(response.status_code, 302)
        self.assertTrue(OeesPrinters.objects.filter(serial_number='PR-NEW').exists())

    def test_save_persists_contract_and_page_costs(self):
        self.client.force_login(self.user)
        self.client.post(reverse('frm_printers'), {
            'action': 'save', 'serial_number': 'PR-CTR', 'description': 'Xerox',
            'ip': '10.0.0.10', 'start_date': '2026-04-01',
            'contract_number': 'RENT-2026-77', 'bw_page_cost': '0.012', 'color_page_cost': '0.085',
        })
        pr = OeesPrinters.objects.get(serial_number='PR-CTR')
        self.assertEqual(pr.contract_number, 'RENT-2026-77')
        self.assertEqual(pr.bw_page_cost, 0.012)
        self.assertEqual(pr.color_page_cost, 0.085)

    def test_page_costs_default_to_zero_and_contract_null_when_blank(self):
        self.client.force_login(self.user)
        self.client.post(reverse('frm_printers'), {
            'action': 'save', 'serial_number': 'PR-BLANK', 'description': 'Brother',
            'ip': '10.0.0.11', 'start_date': '2026-04-01',
        })
        pr = OeesPrinters.objects.get(serial_number='PR-BLANK')
        self.assertIsNone(pr.contract_number)
        self.assertEqual(pr.bw_page_cost, 0)
        self.assertEqual(pr.color_page_cost, 0)

    def test_api_get_printer_returns_new_fields(self):
        self.printer.contract_number = 'C-1'
        self.printer.bw_page_cost = 0.02
        self.printer.color_page_cost = 0.1
        self.printer.save()
        self.client.force_login(self.user)
        data = self.client.get(reverse('api_get_printer'), {'serial_number': 'PR-1'}).json()['data']
        self.assertEqual(data['contract_number'], 'C-1')
        self.assertEqual(data['bw_page_cost'], '0.02')
        self.assertEqual(data['color_page_cost'], '0.1')

    def test_list_shows_total_monthly_fee(self):
        OeesPrinters.objects.create(serial_number='PR-F1', description='A', ip='1', fee=10.0)
        OeesPrinters.objects.create(serial_number='PR-F2', description='B', ip='2', fee=12.5)
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_printers'))
        self.assertEqual(response.context['fee_total'], 22.5)


class AllocationsScreenTests(TestCase):
    """Smoke tests for the Allocations screen."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='al_user', password='pass12345', allocations=1, reader=0)
        self.staff = OeesStaff.objects.create(name='Joe Worker', persona_fisica=1, notes='', state=1)
        self.device = OeesDevices.objects.create(
            serial_number='DV-1', type='Laptop', brand='Dell', model='X',
            origin='New', insert_date=date(2026, 1, 1), value=0.0,
        )
        self.phone = OeesMobilePhones.objects.create(serial_number='PH-A', type='', value=0.0)

    def test_allocations_requires_login(self):
        response = self.client.get(reverse('frm_allocations'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_allocations_page_renders(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_allocations'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/frmAllocations.html')

    def test_search_available_phones(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_allocations_search'), {'kind': 'phones'})
        self.assertEqual(response.status_code, 200)
        self.assertIn('PH-A', response.json()['serials'])

    def test_assign_device_to_staff(self):
        from oe_inventory_py_web.models import OeesDocs
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_allocations'), {
            'action': 'assign_device', 'staff': str(self.staff.id_staff), 'device_serial': 'DV-1',
        })
        self.assertEqual(response.status_code, 302)
        self.device.refresh_from_db()
        self.assertEqual(self.device.persone_id, self.staff.id_staff)
        # Assignment alone no longer generates a document (grouped via the button).
        self.assertFalse(OeesDocs.objects.filter(id_staff_id=self.staff.id_staff).exists())

    def test_generate_allocation_doc_for_physical(self):
        import tempfile
        from django.test import override_settings
        from oe_inventory_py_web.models import OeesDocs
        self.client.force_login(self.user)
        with tempfile.TemporaryDirectory() as tmp, override_settings(MEDIA_ROOT=tmp):
            # Assign a couple of items first (no doc yet), then generate one document.
            self.client.post(reverse('frm_allocations'), {
                'action': 'assign_device', 'staff': str(self.staff.id_staff), 'device_serial': 'DV-1',
            })
            self.client.post(reverse('frm_allocations'), {
                'action': 'assign_phone', 'staff': str(self.staff.id_staff), 'phone_serial': 'PH-A',
            })
            self.assertFalse(OeesDocs.objects.filter(id_staff_id=self.staff.id_staff).exists())

            resp = self.client.post(reverse('frm_allocations'), {
                'action': 'generate_doc', 'staff': str(self.staff.id_staff),
            })
            self.assertEqual(resp.status_code, 302)
            docs = OeesDocs.objects.filter(id_staff_id=self.staff.id_staff, doc_name__startswith='Allocation-')
            self.assertEqual(docs.count(), 1)   # a single grouped document

    def test_generate_doc_skipped_for_non_physical(self):
        from oe_inventory_py_web.models import OeesDocs
        company = OeesStaff.objects.create(name='ACME SL', persona_fisica=0, notes='', state=1)
        self.client.force_login(self.user)
        resp = self.client.post(reverse('frm_allocations'), {
            'action': 'generate_doc', 'staff': str(company.id_staff),
        })
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(OeesDocs.objects.filter(id_staff_id=company.id_staff).exists())

    def test_auto_doc_on_leave_generates_pending(self):
        import tempfile
        from django.test import override_settings
        from oe_inventory_py_web.models import OeesDocs
        self.client.force_login(self.user)
        with tempfile.TemporaryDirectory() as tmp, override_settings(MEDIA_ROOT=tmp):
            # Assign (marks pending in session), then leave without generating.
            self.client.post(reverse('frm_allocations'), {
                'action': 'assign_device', 'staff': str(self.staff.id_staff), 'device_serial': 'DV-1',
            })
            self.assertFalse(OeesDocs.objects.filter(id_staff_id=self.staff.id_staff).exists())

            resp = self.client.post(reverse('frm_allocations'), {'action': 'auto_doc'})
            self.assertEqual(resp.status_code, 204)
            self.assertEqual(OeesDocs.objects.filter(
                id_staff_id=self.staff.id_staff, doc_name__startswith='Allocation-').count(), 1)

            # A second auto_doc (no pending left) does nothing.
            self.client.post(reverse('frm_allocations'), {'action': 'auto_doc'})
            self.assertEqual(OeesDocs.objects.filter(id_staff_id=self.staff.id_staff).count(), 1)

    def test_auto_doc_generates_one_document_per_pending_person(self):
        import tempfile
        from django.test import override_settings
        from oe_inventory_py_web.models import OeesDocs
        other = OeesStaff.objects.create(name='Ann Other', persona_fisica=1, notes='', state=1)
        self.client.force_login(self.user)
        with tempfile.TemporaryDirectory() as tmp, override_settings(MEDIA_ROOT=tmp):
            # Assign to person A, then switch and assign to person B (no doc yet).
            self.client.post(reverse('frm_allocations'), {
                'action': 'assign_device', 'staff': str(self.staff.id_staff), 'device_serial': 'DV-1',
            })
            self.client.post(reverse('frm_allocations'), {
                'action': 'assign_phone', 'staff': str(other.id_staff), 'phone_serial': 'PH-A',
            })
            # On leave, ONE document is generated per pending person.
            resp = self.client.post(reverse('frm_allocations'), {'action': 'auto_doc'})
            self.assertEqual(resp.status_code, 204)
            self.assertEqual(OeesDocs.objects.filter(
                id_staff_id=self.staff.id_staff, doc_name__startswith='Allocation-').count(), 1)
            self.assertEqual(OeesDocs.objects.filter(
                id_staff_id=other.id_staff, doc_name__startswith='Allocation-').count(), 1)


class IncorporationsScreenTests(TestCase):
    """Smoke tests for the Incorporations screen."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='inc_user', password='pass12345', incorporations=1, reader=0)
        self.company = OeesCompanies.objects.create(name='Acme')
        self.deleg = OeesDelegations.objects.create(delegation='HQ', notes='')
        self.inc = OeesIncorporations.objects.create(
            name='New Hire', department='IT', insert_date=date(2026, 1, 1),
            company_id=self.company.id_company, delegation_id=self.deleg.id_delegation,
            cordedh=0, cordlessh=0, usbchub=0, pdf=0, acad=0,
            incorporated=0, send=0, receive=0, descartado=0, sweatshirt_size='L',
            email='ada@example.com',
        )

    def test_incorporations_requires_login(self):
        response = self.client.get(reverse('frm_incorporations'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_incorporations_page_renders(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_incorporations'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/frmIncorporations.html')

    def test_api_get_incorporation_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_incorporation'), {'id': self.inc.id})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['exists'])
        self.assertEqual(data['data']['name'], 'New Hire')
        self.assertEqual(data['data']['sweatshirt_size'], 'L')
        self.assertEqual(data['data']['email'], 'ada@example.com')

    def test_api_get_incorporation_not_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_incorporation'), {'id': 999999})
        self.assertFalse(response.json()['exists'])

    def test_pending_ordered_by_date_descending(self):
        # self.inc is dated 2026-01-01; add an older and a newer one.
        common = dict(
            department='IT', company_id=self.company.id_company,
            delegation_id=self.deleg.id_delegation, cordedh=0, cordlessh=0,
            usbchub=0, pdf=0, acad=0, incorporated=0, send=0, receive=0, descartado=0)
        OeesIncorporations.objects.create(name='Older', insert_date=date(2025, 6, 1), **common)
        OeesIncorporations.objects.create(name='Newer', insert_date=date(2026, 12, 1), **common)
        self.client.force_login(self.user)
        rows = self.client.get(reverse('frm_incorporations')).context['pending_rows']
        dates = [r['date'] for r in rows]
        self.assertEqual(dates, sorted(dates, reverse=True))   # newest date first
        self.assertEqual(rows[0]['name'], 'Newer')

    def test_save_creates_incorporation(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_incorporations'), {
            'action': 'save', 'name': 'Bob New', 'company': str(self.company.id_company),
            'department': 'IT', 'delegation': str(self.deleg.id_delegation),
            'insert_date': '2026-05-01', 'laptop': 'win', 'phone': '1',
            'sweatshirt_size': 'XL', 'email': 'bob@example.com',
        })
        self.assertEqual(response.status_code, 302)
        saved = OeesIncorporations.objects.get(name='Bob New', win=1, phone=1)
        self.assertEqual(saved.sweatshirt_size, 'XL')
        self.assertEqual(saved.email, 'bob@example.com')

    def test_build_incorporation_pdf_bytes(self):
        from oe_inventory_py_web.reports import build_incorporation_form_pdf
        pdf = build_incorporation_form_pdf({
            'name': 'New Hire', 'email': 'ada@example.com', 'sweatshirt_size': 'L',
            'phone': 1, 'keyboard': 1,
        })
        self.assertTrue(pdf.startswith(b'%PDF'))
        self.assertGreater(len(pdf), 1000)

    def test_build_incorporation_pdf_without_size(self):
        # Empty sweatshirt size used to crash reportlab's choice (falsy value).
        from oe_inventory_py_web.reports import build_incorporation_form_pdf
        pdf = build_incorporation_form_pdf({
            'name': 'No Size', 'email': 'ns@example.com', 'sweatshirt_size': '',
        })
        self.assertTrue(pdf.startswith(b'%PDF'))

    def test_build_incorporation_pdf_non_latin1_chars(self):
        # Typographic chars (en-dash, curly quotes) used to crash reportlab's
        # escapePDF (KeyError) on the pure-Python path (as on AWS).
        from oe_inventory_py_web.reports import build_incorporation_form_pdf
        pdf = build_incorporation_form_pdf({
            'id': 1, 'name': 'Ana – O’Neill', 'email': 'a@b.com',
            'address': 'Calle “Sol” 3 – 2ºB', 'is_remote': True,
            'sweatshirt_size': 'M',
        })
        self.assertTrue(pdf.startswith(b'%PDF'))

    def test_apply_pdf_round_trip_updates_record(self):
        # Build the editable PDF, then read it back and apply it to the record:
        # the editable fields must be updated and an audit line prepended.
        from oe_inventory_py_web.reports import build_incorporation_form_pdf
        from oe_inventory_py_web import incorporation_mail
        rec = OeesIncorporations.objects.create(
            name='Round Trip', department='IT', insert_date=date(2026, 3, 1),
            company_id=self.company.id_company, delegation_id=self.deleg.id_delegation,
            cordedh=0, cordlessh=0, usbchub=0, pdf=1, acad=1, mouse=0, left_mouse=1,
            keyboard=0, phone=1, screen=1, incorporated=0, send=0, receive=0,
            descartado=0, sweatshirt_size='S',
        )
        pdf = build_incorporation_form_pdf({
            'id': rec.id, 'name': rec.name, 'email': 'hire@example.com',
            'usbchub': 1, 'pdf': 0, 'mouse': 1, 'left_mouse': 0, 'acad': 0, 'keyboard': 1,
            'sweatshirt_size': 'XXL',
        })
        self.assertTrue(incorporation_mail.apply_pdf(pdf, 'hire@example.com'))

        rec.refresh_from_db()
        self.assertEqual(rec.usbchub, 1)
        self.assertEqual(rec.pdf, 0)
        self.assertEqual(rec.mouse, 1)          # right mouse selected in the PDF
        self.assertEqual(rec.left_mouse, 0)     # ...so left mouse cleared
        self.assertEqual(rec.acad, 0)
        self.assertEqual(rec.keyboard, 1)
        self.assertEqual(rec.sweatshirt_size, 'XXL')
        # Phone/Screen are not in the PDF, so they must be left untouched.
        self.assertEqual(rec.phone, 1)
        self.assertEqual(rec.screen, 1)
        self.assertIn('Received preferences from hire@example.com automatically', rec.notes)
        self.assertEqual(rec.email_processed, 2)

    def test_apply_pdf_left_mouse(self):
        from oe_inventory_py_web.reports import build_incorporation_form_pdf
        from oe_inventory_py_web import incorporation_mail
        rec = OeesIncorporations.objects.create(
            name='Lefty', department='IT', insert_date=date(2026, 3, 1),
            company_id=self.company.id_company, delegation_id=self.deleg.id_delegation,
            cordedh=0, cordlessh=0, usbchub=0, pdf=0, acad=0, mouse=1, left_mouse=0,
            keyboard=0, phone=0, screen=0, incorporated=0, send=0, receive=0, descartado=0,
        )
        pdf = build_incorporation_form_pdf({
            'id': rec.id, 'name': rec.name, 'mouse': 0, 'left_mouse': 1,
        })
        self.assertTrue(incorporation_mail.apply_pdf(pdf, 'x@example.com'))
        rec.refresh_from_db()
        self.assertEqual(rec.mouse, 0)
        self.assertEqual(rec.left_mouse, 1)

    def test_save_mouse_mutually_exclusive(self):
        # If both Right and Left mouse arrive checked, only Right is kept.
        self.client.force_login(self.user)
        self.client.post(reverse('frm_incorporations'), {
            'action': 'save', 'name': 'Both Mice', 'company': str(self.company.id_company),
            'department': 'IT', 'delegation': str(self.deleg.id_delegation),
            'insert_date': '2026-05-01', 'mouse': '1', 'left_mouse': '1',
        })
        saved = OeesIncorporations.objects.get(name='Both Mice')
        self.assertEqual(saved.mouse, 1)
        self.assertEqual(saved.left_mouse, 0)

    def test_apply_pdf_unknown_id_is_ignored(self):
        from oe_inventory_py_web.reports import build_incorporation_form_pdf
        from oe_inventory_py_web import incorporation_mail
        pdf = build_incorporation_form_pdf({'id': 99999999, 'name': 'Ghost'})
        self.assertFalse(incorporation_mail.apply_pdf(pdf, 'x@example.com'))

    def test_apply_pdf_flat_pdf_ignored(self):
        # A flat PDF with no AcroForm fields (a photo/printout of the document)
        # is not an editable form -> ignored.
        import io
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from oe_inventory_py_web import incorporation_mail
        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=A4)
        c.drawString(72, 750, 'Just a scanned document, no form fields.')
        c.save()
        self.assertFalse(incorporation_mail.apply_pdf(buf.getvalue(), 'x@example.com'))

    def test_preferences_requires_email(self):
        # An incorporation without an email address cannot be sent the form.
        no_email = OeesIncorporations.objects.create(
            name='No Mail', department='IT', insert_date=date(2026, 2, 1),
            company_id=self.company.id_company, delegation_id=self.deleg.id_delegation,
            cordedh=0, cordlessh=0, usbchub=0, pdf=0, acad=0,
            incorporated=0, send=0, receive=0, descartado=0,
        )
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_incorporations'), {
            'action': 'preferences', 'code': str(no_email.id),
        })
        self.assertEqual(response.status_code, 302)
        msgs = [m.message for m in get_messages(response.wsgi_request)]
        self.assertTrue(any('email' in m.lower() for m in msgs))

    @override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
                       RESEND_API_KEY='test-key')
    def test_preferences_send_updates_notes(self):
        from django.core import mail
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_incorporations'), {
            'action': 'preferences', 'code': str(self.inc.id),
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, ['ada@example.com'])
        self.inc.refresh_from_db()
        self.assertIn('Sent preferences to ada@example.com by inc_user', self.inc.notes)
        self.assertEqual(self.inc.email_processed, 1)

    def test_save_rejects_invalid_sweatshirt_size(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_incorporations'), {
            'action': 'save', 'name': 'Bad Size', 'company': str(self.company.id_company),
            'department': 'IT', 'delegation': str(self.deleg.id_delegation),
            'insert_date': '2026-05-01', 'sweatshirt_size': 'HUGE',
        })
        self.assertEqual(response.status_code, 302)
        saved = OeesIncorporations.objects.get(name='Bad Size')
        self.assertIsNone(saved.sweatshirt_size)   # invalid size -> not stored

    def test_complete_migrates_to_staff(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_incorporations'), {
            'action': 'complete', 'code': str(self.inc.id),
        })
        self.assertEqual(response.status_code, 302)
        self.inc.refresh_from_db()
        self.assertEqual(self.inc.incorporated, 1)
        self.assertTrue(OeesStaff.objects.filter(name='New Hire').exists())


class OrdersScreenTests(TestCase):
    """Smoke tests for the Orders screen."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='or_user', password='pass12345', orders=1, reader=0)
        self.order = OeesOrders.objects.create(
            article='Cables', uds=10, insert_date=date(2026, 1, 1), notes='',
            tramitado=0, cancelado=0, recibido=0,
        )

    def test_orders_requires_login(self):
        response = self.client.get(reverse('frm_orders'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_orders_page_renders(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_orders'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/frmOrders.html')

    def test_api_get_order_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_order'), {'id': self.order.id_order})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['exists'])
        self.assertEqual(data['data']['article'], 'Cables')

    def test_api_get_order_not_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_order'), {'id': 999999})
        self.assertFalse(response.json()['exists'])

    def test_save_creates_order(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_orders'), {
            'action': 'save', 'article': 'Mice', 'uds': '5', 'insert_date': '2026-06-01',
        })
        self.assertEqual(response.status_code, 302)
        self.assertTrue(OeesOrders.objects.filter(article='Mice', uds=5).exists())

    def test_process_order(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_orders'), {'action': 'process', 'code': str(self.order.id_order)})
        self.assertEqual(response.status_code, 302)
        self.order.refresh_from_db()
        self.assertEqual(self.order.tramitado, 1)


class MobileLinesScreenTests(TestCase):
    """Smoke tests for the Mobile Lines screen."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='ml_user', password='pass12345', mobile_lines=1, reader=0)
        self.line = OeesMobileLines.objects.create(
            number='600100200', imei='', pin='', puk='', pin2='', puk2='',
            extension='', esim=0, m2m=0, obs='', origin='Vodafone',
            fee=12.5, desc_tarif='Plan 20GB',
        )

    def test_mobile_lines_requires_login(self):
        response = self.client.get(reverse('frm_mobile_lines'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_mobile_lines_page_renders(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_mobile_lines'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/frmMobileLines.html')

    def test_api_get_line_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_line'), {'number': '600100200'})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['exists'])
        self.assertEqual(data['data']['origin'], 'Vodafone')
        self.assertEqual(data['data']['fee'], 12.5)
        self.assertEqual(data['data']['desc_tarif'], 'Plan 20GB')

    def test_stock_phones_lists_phones_without_sim(self):
        # The phone dropdown for SIM assignment must list phones WITHOUT a SIM
        # (id_line), regardless of whether they have a person assigned.
        from oe_inventory_py_web.models import OeesMobilePhones, OeesStaff
        staff = OeesStaff.objects.create(name='Owner', notes='', persona_fisica=1)
        OeesMobilePhones.objects.create(serial_number='NO-SIM-NO-PERSON', type='MOBILE', value=0)
        OeesMobilePhones.objects.create(serial_number='NO-SIM-WITH-PERSON', type='MOBILE', value=0, persone=staff)
        OeesMobilePhones.objects.create(serial_number='HAS-SIM', type='MOBILE', value=0, id_line=self.line)
        self.client.force_login(self.user)
        stock = self.client.get(reverse('frm_mobile_lines')).context['stock_phones']
        self.assertIn('NO-SIM-NO-PERSON', stock)
        self.assertIn('NO-SIM-WITH-PERSON', stock)   # has a person but no SIM -> eligible
        self.assertNotIn('HAS-SIM', stock)           # already has a SIM -> excluded

    def test_grid_active_column_reflects_person_state(self):
        from oe_inventory_py_web.models import OeesStaff, OeesMobilePhones
        act = OeesStaff.objects.create(name='Act', notes='', persona_fisica=1, state=1)
        ina = OeesStaff.objects.create(name='Ina', notes='', persona_fisica=1, state=0)
        ph_a = OeesMobilePhones.objects.create(serial_number='PH-ACT', type='MOBILE', value=0, persone=act)
        ph_i = OeesMobilePhones.objects.create(serial_number='PH-INA', type='MOBILE', value=0, persone=ina)
        common = dict(imei='', pin='', puk='', pin2='', puk2='', extension='', esim=0, m2m=0, obs='')
        OeesMobileLines.objects.create(number='600000010', mobile=ph_a, **common)
        OeesMobileLines.objects.create(number='600000011', mobile=ph_i, **common)
        self.client.force_login(self.user)
        resp = self.client.get(reverse('frm_mobile_lines'))
        by_num = {r['number']: r for r in resp.context['grid_rows']}
        self.assertTrue(by_num['600000010']['active'])    # person state == 1
        self.assertFalse(by_num['600000011']['active'])   # person state == 0
        self.assertFalse(by_num['600100200']['active'])   # no phone/person
        self.assertContains(resp, '<th class="text-center">Active</th>')

    def test_api_get_line_not_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_line'), {'number': '000'})
        self.assertFalse(response.json()['exists'])

    def test_save_creates_line(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_mobile_lines'), {
            'action': 'save', 'number': '600999888', 'origin': 'Orange', 'pin': '1234',
            'fee': '9,99', 'desc_tarif': 'Plan 50GB',
        })
        self.assertEqual(response.status_code, 302)
        saved = OeesMobileLines.objects.get(number='600999888')
        self.assertEqual(saved.fee, 9.99)
        self.assertEqual(saved.desc_tarif, 'Plan 50GB')

    def test_cancel_sets_baja(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_mobile_lines'), {'action': 'cancel', 'number': '600100200'})
        self.assertEqual(response.status_code, 302)
        self.line.refresh_from_db()
        self.assertIsNotNone(self.line.fecha_baja)


class AvailabilityScreenTests(TestCase):
    """Smoke tests for the Availability screen."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='av_user', password='pass12345', disponibility=1)
        OeesDevices.objects.create(
            serial_number='AV-1', type='LAPTOP WIN', brand='Dell', model='X',
            origin='New', insert_date=date(2026, 1, 1), value=0.0,
        )

    def test_availability_requires_login(self):
        response = self.client.get(reverse('frm_availability'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_availability_page_renders(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_availability'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/frmAvailability.html')

    def test_availability_excel_export(self):
        # Only non-reader profiles can download the Excel export.
        self.user.reader = 0
        self.user.save(update_fields=['reader'])
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_availability'), {'export': 'excel'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_availability_excel_export_blocked_for_reader(self):
        # Reader profiles (reader=1) cannot download Excel exports; the
        # middleware redirects them back to the screen instead.
        self.user.reader = 1
        self.user.save(update_fields=['reader'])
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_availability'), {'export': 'excel'})
        self.assertEqual(response.status_code, 302)


class UnderRepairScreenTests(TestCase):
    """Smoke tests for the Under Repair screen."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='ur_user', password='pass12345', under_repair=1, reader=0)
        OeesDevices.objects.create(
            serial_number='UR-1', type='Laptop', brand='Dell', model='X',
            origin='New', insert_date=date(2026, 1, 1), value=0.0,
        )
        self.repair = OeesUnderRepair.objects.create(
            type='D', serial_number='UR-1', date_out=date(2026, 1, 2),
            destiny='Tech Service', notes='', value=0.0,
        )

    def test_under_repair_requires_login(self):
        response = self.client.get(reverse('frm_under_repair'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_under_repair_page_renders(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_under_repair'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/frmUnderRepair.html')

    def test_receive_sets_date_in(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_under_repair'), {
            'action': 'receive', 'id': str(self.repair.id_under_repair), 'value': '12,50',
        })
        self.assertEqual(response.status_code, 302)
        self.repair.refresh_from_db()
        self.assertIsNotNone(self.repair.date_in)
        self.assertEqual(self.repair.value, 12.5)


class DistInvoicesScreenTests(TestCase):
    """Smoke tests for the Distribution Invoices screen."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='di_user', password='pass12345', facturas=1, reader=0)
        company = OeesCompanies.objects.create(name='Acme')
        deleg = OeesDelegations.objects.create(delegation='HQ', notes='')
        self.staff = OeesStaff.objects.create(
            name='Owner', persona_fisica=1, notes='', state=1,
            company_id=company.id_company, delegation_id=deleg.id_delegation, department='IT',
        )
        OeesDevices.objects.create(
            serial_number='DI-1', type='Laptop', brand='Dell', model='XPS', origin='New',
            insert_date=date(2026, 1, 1), value=1000.0, bill_number='BILL-1', persone=self.staff,
        )

    def test_dist_invoices_requires_login(self):
        response = self.client.get(reverse('frm_dist_invoices'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_dist_invoices_page_renders(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_dist_invoices'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/frmDistributionInvoices.html')

    def test_dist_invoices_search(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_dist_invoices'), {'bill': 'BILL-1'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'DI-1')

    def test_dist_invoices_excel(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_dist_invoices'), {'bill': 'BILL-1', 'export': 'excel'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )


class PasswordChangeScreenTests(TestCase):
    """Smoke tests for the Password Change screen (Django-managed auth)."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='pwd_user', password='oldpass12345')

    def test_password_change_requires_login(self):
        response = self.client.get(reverse('frm_password_change'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_password_change_page_renders(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_password_change'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/frmPasswordChange.html')

    def test_password_change_success(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_password_change'), {
            'old_password': 'oldpass12345',
            'new_password1': 'BrandNew9988',
            'new_password2': 'BrandNew9988',
        })
        self.assertEqual(response.status_code, 302)
        self.user.refresh_from_db()
        self.assertTrue(self.user.check_password('BrandNew9988'))

    def test_password_change_wrong_old(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_password_change'), {
            'old_password': 'WRONG',
            'new_password1': 'BrandNew9988',
            'new_password2': 'BrandNew9988',
        })
        self.assertEqual(response.status_code, 200)
        self.user.refresh_from_db()
        self.assertTrue(self.user.check_password('oldpass12345'))


class DelegationsScreenTests(TestCase):
    """Smoke tests for the Delegations screen."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='dl_user', password='pass12345', delegations=1, reader=0)
        self.deleg = OeesDelegations.objects.create(delegation='Headquarters', notes='', poblacion='Madrid')

    def test_delegations_requires_login(self):
        response = self.client.get(reverse('frm_delegations'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_delegations_page_renders(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_delegations'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/frmDelegations.html')

    def test_api_get_delegation_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_delegation'), {'id': self.deleg.id_delegation})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['exists'])
        self.assertEqual(data['data']['delegation'], 'Headquarters')
        self.assertEqual(data['data']['poblacion'], 'Madrid')

    def test_save_creates_delegation(self):
        from unittest.mock import patch
        self.client.force_login(self.user)
        # Mock geocoding so the test never hits the network.
        with patch('oe_inventory_py_web.views._geocode_delegation', return_value=(40.4168, -3.7038)):
            response = self.client.post(reverse('frm_delegations'), {
                'action': 'save', 'delegation': 'Branch Office', 'direccion': 'Main St 1',
                'cpostal': '28001', 'poblacion': 'Madrid', 'provincia': 'Madrid',
            })
        self.assertEqual(response.status_code, 302)
        d = OeesDelegations.objects.get(delegation='Branch Office', poblacion='Madrid')
        self.assertEqual(d.latitude, 40.4168)   # geocoded coordinates stored
        self.assertEqual(d.longitude, -3.7038)

    def test_map_includes_geocoded_delegation(self):
        # A delegation with coordinates must appear in the map data on the page.
        self.deleg.latitude = 40.4168
        self.deleg.longitude = -3.7038
        self.deleg.save(update_fields=['latitude', 'longitude'])
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_delegations'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'delegations-map')      # map container
        self.assertContains(response, 'delegations-map-data')  # JSON data block
        self.assertContains(response, '40.4168')               # the point itself

    def test_geocode_action_updates_coordinates(self):
        from unittest.mock import patch
        self.client.force_login(self.user)
        with patch('oe_inventory_py_web.views._geocode_delegation', return_value=(41.3874, 2.1686)):
            response = self.client.post(reverse('frm_delegations'), {
                'action': 'geocode', 'code': str(self.deleg.id_delegation),
                'direccion': 'Av Diagonal 1', 'poblacion': 'Barcelona',
            })
        self.assertEqual(response.status_code, 302)
        self.deleg.refresh_from_db()
        self.assertEqual(self.deleg.latitude, 41.3874)
        self.assertEqual(self.deleg.longitude, 2.1686)


class AccessCardsScreenTests(TestCase):
    """Smoke tests for the Access Cards screen."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='ac_user', password='pass12345', access_cards=1, reader=0)
        self.state = OeesAccessCardsStates.objects.create(id_state=1, description='ACTIVE')
        self.card = OeesAccessCards.objects.create(
            ac_max='C1', fermax_mif='F1', pin_card='', state_card=1, obs='', notes='',
        )

    def test_access_cards_requires_login(self):
        response = self.client.get(reverse('frm_access_cards'))
        self.assertEqual(response.status_code, 302)

    def test_access_cards_page_renders(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_access_cards'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/frmAccessCards.html')

    def test_api_get_card_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_card'), {'card': 'C1'})
        self.assertTrue(response.json()['exists'])

    def test_save_creates_card(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_access_cards'), {
            'action': 'save', 'card': 'C2', 'fermax': 'F2', 'state': '1', 'obs': '',
        })
        self.assertEqual(response.status_code, 302)
        self.assertTrue(OeesAccessCards.objects.filter(ac_max='C2').exists())

    def test_save_card_consumes_pin_from_pool(self):
        from oe_inventory_py_web.models import OeesAccessCardsPins
        OeesAccessCardsPins.objects.create(pin='1234')
        OeesAccessCardsPins.objects.create(pin='5678')
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_access_cards'), {
            'action': 'save', 'card': 'C3', 'fermax': 'F3', 'state': '1', 'obs': '', 'pin': '1234',
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(OeesAccessCards.objects.get(ac_max='C3').pin_card, '1234')
        self.assertFalse(OeesAccessCardsPins.objects.filter(pin='1234').exists())  # consumed
        self.assertTrue(OeesAccessCardsPins.objects.filter(pin='5678').exists())   # untouched

    def test_can_mark_card_as_lost(self):
        OeesAccessCardsStates.objects.create(id_state=2, description='LOST')
        self.client.force_login(self.user)
        self.client.post(reverse('frm_access_cards'), {
            'action': 'save', 'card': 'C1', 'fermax': 'F1', 'state': '2', 'obs': '',
        })
        self.card.refresh_from_db()
        self.assertEqual(self.card.state_card, 2)   # marking AS lost is allowed

    def test_cannot_modify_a_card_already_lost(self):
        OeesAccessCardsStates.objects.create(id_state=2, description='LOST')
        self.card.state_card = 2
        self.card.save()
        self.client.force_login(self.user)
        self.client.post(reverse('frm_access_cards'), {
            'action': 'save', 'card': 'C1', 'fermax': 'CHANGED', 'state': '1', 'obs': '',
        })
        self.card.refresh_from_db()
        self.assertEqual(self.card.fermax_mif, 'F1')   # blocked: unchanged
        self.assertEqual(self.card.state_card, 2)       # still lost

    def test_api_get_card_returns_assigned_staff_name(self):
        staff = OeesStaff.objects.create(
            name='Gone Person', notes='', persona_fisica=1, state=0, fecha_baja='2026-01-01')
        self.card.id_staff = staff
        self.card.save()
        self.client.force_login(self.user)
        data = self.client.get(reverse('api_get_card'), {'card': 'C1'}).json()['data']
        self.assertEqual(data['staff_id'], staff.id_staff)
        self.assertEqual(data['staff_name'], 'Gone Person')

    def test_card_pin_api_returns_available_pin(self):
        from oe_inventory_py_web.models import OeesAccessCardsPins
        OeesAccessCardsPins.objects.create(pin='4321')
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_card_pin'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['pin'], '4321')

    def test_release_converts_to_visitor(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_access_cards'), {'action': 'release', 'card': 'C1'})
        self.assertEqual(response.status_code, 302)
        self.assertTrue(OeesAccessCardsVisitors.objects.filter(ac_max='C1').exists())
        self.assertFalse(OeesAccessCards.objects.filter(ac_max='C1').exists())


class AccessKeysScreenTests(TestCase):
    """Smoke tests for the Access Keys screen."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='ak_user', password='pass12345', access_keys=1, reader=0)
        self.key = OeesAccessKeys.objects.create(type='Master', insert_date=date(2026, 1, 1), notes='')

    def test_access_keys_requires_login(self):
        response = self.client.get(reverse('frm_access_keys'))
        self.assertEqual(response.status_code, 302)

    def test_access_keys_page_renders(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_access_keys'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/frmAccessKeys.html')

    def test_api_get_key_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_key'), {'id': self.key.id_key})
        data = response.json()
        self.assertTrue(data['exists'])
        self.assertEqual(data['data']['type'], 'Master')

    def test_save_creates_key(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_access_keys'), {
            'action': 'save', 'type': 'Backup', 'insert_date': '2026-02-01',
        })
        self.assertEqual(response.status_code, 302)
        self.assertTrue(OeesAccessKeys.objects.filter(type='Backup').exists())


class VisitorCardsScreenTests(TestCase):
    """Smoke tests for the Visitors Access Cards screen."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='vc_user', password='pass12345', visitors_cards=1, reader=0)
        self.state = OeesAccessCardsStates.objects.create(id_state=1, description='ACTIVE')
        self.card = OeesAccessCardsVisitors.objects.create(
            ac_max='V1', fermax_mif='F1', name='Visitor', state_card_id=1, obs='', pin_card='', notes='',
        )

    def test_visitor_cards_requires_login(self):
        response = self.client.get(reverse('frm_visitor_cards'))
        self.assertEqual(response.status_code, 302)

    def test_visitor_cards_page_renders(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_visitor_cards'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'oe_inventory_py_web/frmVisitorsAccessCards.html')

    def test_api_get_visitor_card_found(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_get_visitor_card'), {'card': 'V1'})
        data = response.json()
        self.assertTrue(data['exists'])
        self.assertEqual(data['data']['user'], 'Visitor')

    def test_save_creates_visitor_card(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_visitor_cards'), {
            'action': 'save', 'card': 'V2', 'fermax': 'F2', 'user': 'Jane', 'state': '1', 'obs': '',
        })
        self.assertEqual(response.status_code, 302)
        self.assertTrue(OeesAccessCardsVisitors.objects.filter(ac_max='V2').exists())


class PasswordFlowTests(TestCase):
    """Initial-password by admin (frmUser) and the 'Forgot my password' reset flow."""

    def test_login_page_has_forgot_password_link(self):
        response = self.client.get(reverse('login'))
        self.assertContains(response, reverse('password_reset'))

    def test_password_reset_form_renders(self):
        response = self.client.get(reverse('password_reset'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'registration/password_reset_form.html')

    def test_admin_with_users_permit_sets_initial_password_for_new_user(self):
        User = get_user_model()
        admin = User.objects.create_user(username='admin_u', password='pass12345', users=1)
        self.client.force_login(admin)
        response = self.client.post(reverse('frm_users'), {
            'login': 'newbie', 'nombre': 'New Bie', 'is_new': '1',
            'email': 'newbie@example.com', 'password': 'initpass123',
        })
        self.assertEqual(response.status_code, 302)
        newbie = User.objects.get(username='newbie')
        self.assertTrue(newbie.has_usable_password())
        self.assertTrue(newbie.check_password('initpass123'))
        self.assertEqual(newbie.email, 'newbie@example.com')  # email stored for recovery

    def test_password_ignored_without_users_permit(self):
        User = get_user_model()
        plain = User.objects.create_user(username='plain_u', password='pass12345', users=0)
        self.client.force_login(plain)
        self.client.post(reverse('frm_users'), {
            'login': 'newbie2', 'nombre': 'New Bie2', 'is_new': '1', 'password': 'initpass123',
        })
        newbie2 = User.objects.get(username='newbie2')
        self.assertFalse(newbie2.has_usable_password())  # no Users permit -> password not set

    def test_initial_password_not_overwritten_for_existing_user(self):
        # An existing user that already has a password keeps it (the field is
        # only for setting an INITIAL password).
        User = get_user_model()
        admin = User.objects.create_user(username='admin_u2', password='pass12345', users=1)
        target = User.objects.create_user(username='haspw', password='original123', staff=1)
        self.client.force_login(admin)
        self.client.post(reverse('frm_users'), {
            'login': 'haspw', 'nombre': 'Has Pw', 'is_new': '0', 'password': 'hacked999',
        })
        target.refresh_from_db()
        self.assertTrue(target.check_password('original123'))   # unchanged
        self.assertFalse(target.check_password('hacked999'))


class PasswordResetEmailTests(TestCase):
    """The password-reset email is sent through the Resend backend."""

    def test_reset_email_routes_through_resend_backend(self):
        from django.test import override_settings
        from unittest.mock import patch
        User = get_user_model()
        User.objects.create_user(username='resetme', password='pass12345', email='reset@x.com')
        with override_settings(
            EMAIL_BACKEND='oe_inventory_py_web.email_backend.ResendEmailBackend',
            RESEND_API_KEY='test-key', DEFAULT_FROM_EMAIL='noreply@x.com',
        ):
            with patch('oe_inventory_py_web.email_backend.ResendEmailBackend._send') as send:
                resp = self.client.post(reverse('password_reset'), {'email': 'reset@x.com'})
        self.assertEqual(resp.status_code, 302)          # -> password_reset_done
        self.assertEqual(send.call_count, 1)             # the email went via Resend
        msg = send.call_args[0][0]
        self.assertIn('reset@x.com', list(msg.to))
        self.assertEqual(msg.from_email, 'noreply@x.com')


class FooterCountersTests(TestCase):
    """The base_mdi footer counters (pending access cards / pending orders)."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='footer_u', password='pass12345')

    def test_pending_counters(self):
        from datetime import date
        from oe_inventory_py_web.models import (
            OeesAccessCardsStates, OeesAccessCards, OeesOrders,
        )
        pending = OeesAccessCardsStates.objects.create(description='PENDING')
        delivered = OeesAccessCardsStates.objects.create(description='DELIVERED')

        # Access cards: only the one in PENDING should count.
        OeesAccessCards.objects.create(ac_max='C1', fermax_mif='F1', pin_card='1234',
                                       state_card=pending.id_state, obs='', notes='')
        OeesAccessCards.objects.create(ac_max='C2', fermax_mif='F2', pin_card='1234',
                                       state_card=delivered.id_state, obs='', notes='')

        # Orders: only tramitado=0 AND cancelado=0 should count as pending.
        OeesOrders.objects.create(article='A', uds=1, insert_date=date(2026, 1, 1), notes='',
                                  tramitado=0, cancelado=0, recibido=0)   # pending
        OeesOrders.objects.create(article='B', uds=1, insert_date=date(2026, 1, 1), notes='',
                                  tramitado=1, cancelado=0, recibido=0)   # processed -> excluded
        OeesOrders.objects.create(article='C', uds=1, insert_date=date(2026, 1, 1), notes='',
                                  tramitado=0, cancelado=1, recibido=0)   # cancelled -> excluded

        self.client.force_login(self.user)
        response = self.client.get(reverse('mdi_home'))
        self.assertEqual(response.context['total_cards'], 1)
        self.assertEqual(response.context['total_orders'], 1)


class StaffDocStorageTests(TestCase):
    """Staff documents are saved/served via Django's storage API (S3-ready)."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='doc_u', password='pass12345', staff=1, reader=0)
        self.staff = OeesStaff.objects.create(name='Doc Person', persona_fisica=1, notes='',
                                              state=1, department='IT')

    def test_upload_and_serve_via_storage(self):
        import tempfile
        from django.test import override_settings
        from django.core.files.uploadedfile import SimpleUploadedFile
        from oe_inventory_py_web.models import OeesDocs

        with tempfile.TemporaryDirectory() as tmp, override_settings(MEDIA_ROOT=tmp):
            self.client.force_login(self.user)
            pdf = SimpleUploadedFile('contract.pdf', b'%PDF-1.4 test', content_type='application/pdf')
            up = self.client.post(reverse('frm_staff'), {
                'action': 'upload_doc', 'code': str(self.staff.id_staff),
                'doc_name': 'Contract', 'doc_file': pdf,
            })
            self.assertEqual(up.status_code, 302)
            self.assertTrue(OeesDocs.objects.filter(
                id_staff_id=self.staff.id_staff, doc_name='Contract').exists())

            served = self.client.get(reverse('staff_doc', kwargs={
                'staff_id': self.staff.id_staff, 'doc_name': 'Contract'}))
            self.assertEqual(served.status_code, 200)
            self.assertEqual(served['Content-Type'], 'application/pdf')


class OnlineUsersTests(TestCase):
    """Counting users currently connected to the app (option 2: last activity)."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(
            username='online_a', password='pass12345', devices=1
        )
        self.other = User.objects.create_user(
            username='online_b', password='pass12345', devices=1
        )

    def _make_session(self, user, last_activity):
        """Create a DB session for `user` stamped with `last_activity` (epoch)."""
        from django.contrib.sessions.backends.db import SessionStore
        store = SessionStore()
        store['_auth_user_id'] = str(user.pk)
        store['last_activity'] = last_activity
        store.create()
        return store

    def test_counts_recently_active_user(self):
        import time
        from .context_processors import count_online_users
        self._make_session(self.user, time.time())
        self.assertEqual(count_online_users(), 1)

    def test_ignores_stale_session(self):
        import time
        from .context_processors import count_online_users
        # Active 10 minutes ago: outside the 5-minute window.
        self._make_session(self.user, time.time() - 600)
        self.assertEqual(count_online_users(), 0)

    def test_counts_distinct_users_not_sessions(self):
        import time
        from .context_processors import count_online_users
        now = time.time()
        # Same user on two devices must count once; a second user adds one.
        self._make_session(self.user, now)
        self._make_session(self.user, now)
        self._make_session(self.other, now)
        self.assertEqual(count_online_users(), 2)

    def test_middleware_stamps_last_activity(self):
        self.client.force_login(self.user)
        self.client.get(reverse('mdi_home'))
        session = self.client.session
        self.assertIn('last_activity', session)

    def test_online_count_in_footer_context(self):
        import time
        # Another user already has a persisted active session.
        self._make_session(self.other, time.time())
        self.client.force_login(self.user)
        response = self.client.get(reverse('mdi_home'))
        self.assertIn('online_users', response.context)
        self.assertGreaterEqual(response.context['online_users'], 1)

    def test_current_user_always_counted_on_own_footer(self):
        # The requester is online by definition even before their session
        # stamp is flushed to the DB, so the footer must never show 0 for them.
        self.client.force_login(self.user)
        response = self.client.get(reverse('mdi_home'))
        self.assertGreaterEqual(response.context['online_users'], 1)

    def test_current_user_always_listed_in_popup(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_online_users'))
        self.assertIn('online_a', response.json()['users'])

    def test_api_online_users_returns_names(self):
        import time
        self._make_session(self.other, time.time())
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_online_users'))
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertIn('online_b', data['users'])
        self.assertEqual(data['count'], len(data['users']))

    def test_api_online_users_uses_full_name_when_available(self):
        import time
        self.other.first_name = 'Ada'
        self.other.last_name = 'Lovelace'
        self.other.save()
        self._make_session(self.other, time.time())
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_online_users'))
        self.assertIn('Ada Lovelace', response.json()['users'])

    def test_api_online_users_requires_login(self):
        response = self.client.get(reverse('api_online_users'))
        self.assertEqual(response.status_code, 302)


class DevicesGridServerSideTests(TestCase):
    """The Devices grid loads server-side (DataTables) instead of all at once."""

    def setUp(self):
        from oe_inventory_py_web.models import OeesDevices
        User = get_user_model()
        self.user = User.objects.create_user(
            username='dev_admin', password='pass12345', devices=1, reader=0)
        self.company = OeesCompanies.objects.create(name='Acme')
        for i in range(3):
            OeesDevices.objects.create(
                serial_number=f'SN-{i}', type='LAPTOP', brand='Dell',
                model=f'M{i}', origin='New', insert_date=date(2026, 1, 1), value=100.0)
        OeesDevices.objects.create(
            serial_number='PHONE-9', type='PHONE', brand='Apple',
            model='15', origin='New', insert_date=date(2026, 1, 1), value=900.0,
            bill_number='FAC-2026-001')

    def test_frm_devices_does_not_render_rows_inline(self):
        # The page ships an empty tbody; rows arrive via AJAX afterwards.
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_devices'))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'class="device-row"')
        # Totals are computed by aggregation, not by looping a grid in Python.
        self.assertEqual(response.context['total_devices'], 4)
        self.assertEqual(response.context['total_value'], 1200.0)

    def test_finder_searches_devices_by_partial_serial(self):
        # The lupa finder must search by serial number (partial), not by model.
        self.client.force_login(self.user)
        resp = self.client.get(reverse('api_finder'),
                               {'term': 'SN-', 'field': 'serial_number', 'option': 'devices'})
        self.assertEqual(resp.status_code, 200)
        data = resp.json()['data']
        self.assertEqual({d['code'] for d in data}, {'SN-0', 'SN-1', 'SN-2'})
        # The description column shows the model.
        self.assertEqual(next(d for d in data if d['code'] == 'SN-0')['description'], 'M0')

    def test_datatable_returns_paginated_json(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_devices_datatable'), {
            'draw': '2', 'start': '0', 'length': '2',
            'order[0][column]': '0', 'order[0][dir]': 'asc',
        })
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['draw'], 2)
        self.assertEqual(data['recordsTotal'], 4)
        self.assertEqual(data['recordsFiltered'], 4)
        self.assertEqual(len(data['data']), 2)  # only the requested page

    def test_datatable_global_search_filters(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('api_devices_datatable'), {
            'draw': '1', 'start': '0', 'length': '50',
            'search[value]': 'PHONE',
        })
        data = response.json()
        self.assertEqual(data['recordsFiltered'], 1)
        self.assertEqual(data['data'][0]['serial'], 'PHONE-9')
        # The Bill Number column must carry its value through to the grid.
        self.assertEqual(data['data'][0]['bill'], 'FAC-2026-001')

    def test_datatable_requires_login(self):
        response = self.client.get(reverse('api_devices_datatable'))
        self.assertEqual(response.status_code, 302)

    def test_save_creates_a_brand_new_device(self):
        # Typing a new serial and pressing Save must create the device, like
        # frmPhones/frmLicenses do.
        from oe_inventory_py_web.models import OeesDevices
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_devices'), {
            'action': 'save', 'serial_number': 'BRAND-NEW-1', 'id_company': str(self.company.id_company),
            'type': 'LAPTOP', 'brand': 'HP', 'model': 'X1', 'value': '0',
        })
        self.assertEqual(response.status_code, 200)
        self.assertTrue(OeesDevices.objects.filter(serial_number='BRAND-NEW-1').exists())

    def test_save_requires_a_company(self):
        from oe_inventory_py_web.models import OeesDevices
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_devices'), {
            'action': 'save', 'serial_number': 'NO-COMPANY-1',
            'type': 'LAPTOP', 'brand': 'HP', 'model': 'X1', 'value': '0',
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Please select a company')
        self.assertFalse(OeesDevices.objects.filter(serial_number='NO-COMPANY-1').exists())

    def test_api_get_device_returns_obs(self):
        # The lookup must return obs so the form can populate the Obs field;
        # otherwise saving would post an empty obs and wipe it.
        from oe_inventory_py_web.models import OeesDevices
        OeesDevices.objects.filter(serial_number='SN-1').update(obs='check battery')
        self.client.force_login(self.user)
        data = self.client.get(reverse('api_get_device'), {'serial_number': 'SN-1'}).json()
        self.assertEqual(data['obs'], 'check battery')

    def test_save_persists_obs_and_logs_history(self):
        from oe_inventory_py_web.models import OeesDevices
        self.client.force_login(self.user)
        self.client.post(reverse('frm_devices'), {
            'action': 'save', 'serial_number': 'SN-1', 'id_company': str(self.company.id_company),
            'type': 'LAPTOP', 'brand': 'Dell', 'model': 'M1', 'value': '100',
            'obs': 'needs cleaning', 'notes': '',
        })
        d = OeesDevices.objects.get(serial_number='SN-1')
        self.assertEqual(d.obs, 'needs cleaning')
        self.assertIn('Updated by', d.notes)  # a history entry was logged

    def test_save_keeps_existing_history_and_prepends(self):
        from oe_inventory_py_web.models import OeesDevices
        OeesDevices.objects.filter(serial_number='SN-2').update(notes='OLD ENTRY')
        self.client.force_login(self.user)
        self.client.post(reverse('frm_devices'), {
            'action': 'save', 'serial_number': 'SN-2', 'id_company': str(self.company.id_company),
            'type': 'LAPTOP', 'brand': 'Dell', 'model': 'M2', 'value': '100',
            'obs': '', 'notes': 'OLD ENTRY',
        })
        d = OeesDevices.objects.get(serial_number='SN-2')
        self.assertIn('Updated by', d.notes)
        self.assertIn('OLD ENTRY', d.notes)  # previous history preserved

    def test_api_get_device_reports_under_repair_flag(self):
        # The AJAX lookup must say whether the device is in maintenance, so the
        # banner can be toggled when switching devices without a page reload.
        from datetime import date as _date
        OeesUnderRepair.objects.create(
            serial_number='SN-0', type='1', date_out=_date(2026, 1, 2),
            destiny='Technical Service', notes='', value=0.0)  # date_in NULL = in repair
        self.client.force_login(self.user)
        in_repair = self.client.get(reverse('api_get_device'), {'serial_number': 'SN-0'}).json()
        self.assertTrue(in_repair['under_repair'])
        not_repair = self.client.get(reverse('api_get_device'), {'serial_number': 'SN-1'}).json()
        self.assertFalse(not_repair['under_repair'])

    def test_unassigned_device_renders_without_fk_error(self):
        # SN-0 has no assigned person; loading it must not raise on the
        # persone FK and must show "Unassigned".
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_devices'), {
            'action': 'find', 'serial_number': 'SN-0',
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['device_staff'], 'Unassigned')

    def test_support_send_stores_destiny_type_and_note(self):
        self.client.force_login(self.user)
        self.client.post(reverse('frm_devices'), {
            'action': 'support', 'serial_number': 'SN-0', 'repair_destiny': 'Madrid SAT',
        })
        rep = OeesUnderRepair.objects.filter(serial_number='SN-0', date_in__isnull=True).first()
        self.assertIsNotNone(rep)
        self.assertEqual(rep.destiny, 'Madrid SAT')
        self.assertEqual(rep.type, 'D')          # marked as a device
        dev = OeesDevices.objects.get(serial_number='SN-0')
        self.assertIn('Sent to maintenance (Madrid SAT)', dev.notes)

    def test_support_message_rendered_on_same_page(self):
        # The success message must appear on this render (consumed here), not
        # leak to the next screen the user visits.
        self.client.force_login(self.user)
        response = self.client.post(reverse('frm_devices'), {
            'action': 'support', 'serial_number': 'SN-0', 'repair_destiny': 'SAT',
        })
        self.assertContains(response, 'sent to technical support')

    def test_sent_device_appears_in_under_repair_pending(self):
        from oe_inventory_py_web.views import _under_repair_rows
        self.client.force_login(self.user)
        self.client.post(reverse('frm_devices'), {
            'action': 'support', 'serial_number': 'SN-0', 'repair_destiny': 'SAT',
        })
        serials = [r['serial'] for r in _under_repair_rows(repaired=False)]
        self.assertIn('SN-0', serials)

    def test_support_receive_records_cost_and_note(self):
        from datetime import date as _date
        OeesUnderRepair.objects.create(
            serial_number='SN-1', type='D', date_out=_date(2026, 1, 2),
            destiny='SAT', notes='', value=0.0)
        self.client.force_login(self.user)
        self.client.post(reverse('frm_devices'), {
            'action': 'support', 'serial_number': 'SN-1', 'repair_value': '45.50',
        })
        rep = OeesUnderRepair.objects.get(serial_number='SN-1')
        self.assertIsNotNone(rep.date_in)       # received
        self.assertEqual(rep.value, 45.50)       # cost recorded
        dev = OeesDevices.objects.get(serial_number='SN-1')
        self.assertIn('Received from maintenance', dev.notes)


class UserManualTests(TestCase):
    """In-app user manual: rendering, language switch and contextual help."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(
            username='manual_u', password='pass12345', devices=1)

    def test_manual_requires_login(self):
        response = self.client.get(reverse('manual'))
        self.assertEqual(response.status_code, 302)

    def test_manual_renders_spanish_by_default(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('manual'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Manual de Usuario')
        # A stable per-screen anchor must be present for deep links.
        self.assertContains(response, 'id="devices"')

    def test_manual_renders_english(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('manual'), {'lang': 'en'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'User Manual')

    def test_manual_shows_both_language_buttons(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('manual'))
        self.assertContains(response, '?lang=es')
        self.assertContains(response, '?lang=en')

    def test_manual_rewrites_image_paths_to_static(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('manual'))
        body = response.content.decode()
        self.assertIn('/static/manual_images/', body)
        self.assertNotIn('src="images/', body)

    def test_help_anchor_in_context_for_a_form(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_devices'))
        self.assertEqual(response.context['manual_anchor'], 'devices')

    def test_navbar_has_manual_link(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('mdi_home'))
        self.assertContains(response, reverse('manual'))


class ResendEmailBackendTests(TestCase):
    """Email goes through the Resend API backend (replaces the SMTP backend)."""

    def test_payload_plain_text_message(self):
        from django.core.mail import EmailMessage
        from oe_inventory_py_web.email_backend import build_resend_payload
        msg = EmailMessage(subject='Reset', body='Click here', to=['u@example.com'])
        payload = build_resend_payload(msg)
        self.assertEqual(payload['to'], ['u@example.com'])
        self.assertEqual(payload['subject'], 'Reset')
        self.assertEqual(payload['text'], 'Click here')
        self.assertNotIn('html', payload)

    def test_payload_html_alternative_becomes_html(self):
        from django.core.mail import EmailMultiAlternatives
        from oe_inventory_py_web.email_backend import build_resend_payload
        msg = EmailMultiAlternatives(subject='Hi', body='plain', to=['u@example.com'])
        msg.attach_alternative('<b>rich</b>', 'text/html')
        payload = build_resend_payload(msg)
        self.assertEqual(payload['html'], '<b>rich</b>')

    def test_payload_includes_pdf_attachment(self):
        from django.core.mail import EmailMessage
        from oe_inventory_py_web.email_backend import build_resend_payload
        msg = EmailMessage(subject='Inv', body='see attached', to=['u@example.com'])
        msg.attach('Inventory.pdf', b'%PDF-1.4 xyz', 'application/pdf')
        payload = build_resend_payload(msg)
        self.assertEqual(len(payload['attachments']), 1)
        att = payload['attachments'][0]
        self.assertEqual(att['filename'], 'Inventory.pdf')
        # Bytes are serialised as a list of integers (Resend SDK format).
        self.assertEqual(bytes(att['content']), b'%PDF-1.4 xyz')

    def test_payload_uses_default_from_when_unset(self):
        from django.core.mail import EmailMessage
        from django.test import override_settings
        from oe_inventory_py_web.email_backend import build_resend_payload
        with override_settings(DEFAULT_FROM_EMAIL='noreply@octoenergy.com'):
            msg = EmailMessage(subject='x', body='y', to=['u@example.com'])
            payload = build_resend_payload(msg)
            self.assertEqual(payload['from'], 'noreply@octoenergy.com')

    def test_backend_errors_without_api_key(self):
        from django.core.mail import EmailMessage
        from django.test import override_settings
        from oe_inventory_py_web.email_backend import ResendEmailBackend
        msg = EmailMessage(subject='x', body='y', to=['u@example.com'])
        with override_settings(RESEND_API_KEY=''):
            backend = ResendEmailBackend(fail_silently=False)
            with self.assertRaises(ValueError):
                backend.send_messages([msg])
            # With fail_silently the call is a no-op (0 sent), never the network.
            backend_silent = ResendEmailBackend(fail_silently=True)
            self.assertEqual(backend_silent.send_messages([msg]), 0)


class SessionSecurityTests(TestCase):
    """Sessions expire after inactivity and when the browser closes."""

    def test_session_settings_are_configured(self):
        from django.conf import settings
        self.assertEqual(settings.SESSION_COOKIE_AGE, 30 * 60)
        self.assertTrue(settings.SESSION_SAVE_EVERY_REQUEST)
        self.assertTrue(settings.SESSION_EXPIRE_AT_BROWSER_CLOSE)

    def test_session_is_browser_close_with_30min_idle(self):
        User = get_user_model()
        user = User.objects.create_user(username='sess_u', password='pass12345')
        self.client.force_login(user)
        self.client.get(reverse('mdi_home'))
        session = self.client.session
        # Cookie dies on browser close, and the idle window is 30 minutes.
        self.assertTrue(session.get_expire_at_browser_close())
        self.assertEqual(session.get_expiry_age(), 30 * 60)


class NotReturnedScreenTests(TestCase):
    """Material still assigned to staff who already left (fecha_baja set)."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(
            username='nr_user', password='pass12345', not_returned=1, reader=0)
        self.gone = OeesStaff.objects.create(
            name='Zoe Gone', notes='', persona_fisica=1, state=0, fecha_baja='2026-01-01')
        self.active = OeesStaff.objects.create(
            name='Andy Active', notes='', persona_fisica=1, state=1)
        OeesDevices.objects.create(
            serial_number='DEV-GONE', type='LAPTOP', brand='Dell', model='X',
            origin='New', insert_date=date(2026, 1, 1), value=500.0,
            persone=self.gone, mobile_line=0)
        OeesDevices.objects.create(
            serial_number='DEV-ACTIVE', type='LAPTOP', brand='HP', model='Y',
            origin='New', insert_date=date(2026, 1, 1), value=300.0,
            persone=self.active, mobile_line=0)

    def test_requires_login(self):
        self.assertEqual(self.client.get(reverse('frm_not_returned')).status_code, 302)

    def test_lists_only_terminated_staff_items(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_not_returned'))
        self.assertEqual(response.status_code, 200)
        serials = [r['serial'] for r in response.context['rows']]
        self.assertIn('DEV-GONE', serials)         # left -> shown
        self.assertNotIn('DEV-ACTIVE', serials)    # active -> not shown

    def test_row_has_person_date_aging_and_value(self):
        self.client.force_login(self.user)
        rows = self.client.get(reverse('frm_not_returned')).context['rows']
        row = next(r for r in rows if r['serial'] == 'DEV-GONE')
        self.assertEqual(row['person'], 'Zoe Gone')
        self.assertEqual(row['termination_date'], '2026-01-01')
        self.assertEqual(row['category'], 'Device')
        self.assertEqual(row['value'], 500.0)
        self.assertIsNotNone(row['aging_days'])

    def test_total_value_only_counts_terminated(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('frm_not_returned'))
        self.assertEqual(response.context['total_value'], 500.0)

    def test_includes_access_keys(self):
        from oe_inventory_py_web.models import OeesAccessKeys
        OeesAccessKeys.objects.create(type='Main door', id_staff=self.gone, insert_date=date(2026, 1, 1))
        self.client.force_login(self.user)
        rows = self.client.get(reverse('frm_not_returned')).context['rows']
        self.assertTrue(any(r['category'] == 'Access Key' and r['person'] == 'Zoe Gone' for r in rows))

    def test_spanish_date_normalized_to_english_and_aged(self):
        spanish = OeesStaff.objects.create(
            name='Bob Spanish', notes='', persona_fisica=1, state=0, fecha_baja='15-03-2026')
        OeesDevices.objects.create(
            serial_number='DEV-ES', type='LAPTOP', brand='Dell', model='Z',
            origin='New', insert_date=date(2026, 1, 1), value=100.0,
            persone=spanish, mobile_line=0)
        self.client.force_login(self.user)
        rows = self.client.get(reverse('frm_not_returned')).context['rows']
        row = next(r for r in rows if r['serial'] == 'DEV-ES')
        self.assertEqual(row['termination_date'], '2026-03-15')   # dd-mm-yyyy -> yyyy-mm-dd
        self.assertIsNotNone(row['aging_days'])                   # parsed -> aging computed

    def test_subtotals_row_inserted_per_person(self):
        self.client.force_login(self.user)
        rows = self.client.get(reverse('frm_not_returned'), {'subtotals': 'on'}).context['rows']
        subs = [r for r in rows if r.get('sub')]
        self.assertTrue(any(r['person'] == 'Total Zoe Gone' and r['value'] == 500.0 for r in subs))


class NebulaGatewayWanTests(TestCase):
    """nebula._gateway_wan: enabled WAN count + operational link count."""

    def _settings(self):
        from django.test import override_settings
        return override_settings(NEBULA_BASE_URL='https://x', NEBULA_API_KEY='t')

    def test_enabled_count_and_operational_unknown_when_ports_status_fails(self):
        from unittest.mock import patch
        from oe_inventory_py_web import nebula

        def fake_request(path, method='GET', body=None):
            if path.endswith('/interface-settings'):
                return {'wan': [{'interface': 'ge1', 'enabled': True},
                                {'interface': 'ge2', 'enabled': True},
                                {'interface': 'ge3', 'enabled': False}]}
            if path.endswith('/ports-status'):
                raise nebula.NebulaError('500')
            return None

        with self._settings(), patch('oe_inventory_py_web.nebula._request', side_effect=fake_request):
            w = nebula._gateway_wan('site1', 'dev1')
        self.assertEqual(w['enabled'], 2)        # ge1, ge2 (ge3 disabled)
        self.assertIsNone(w['operational'])      # ports-status failed -> unknown

    def test_operational_counts_wan_ports_with_link(self):
        from unittest.mock import patch
        from oe_inventory_py_web import nebula

        def fake_request(path, method='GET', body=None):
            if path.endswith('/interface-settings'):
                return {'wan': [{'interface': 'ge1', 'enabled': True},
                                {'interface': 'ge2', 'enabled': True}]}
            if path.endswith('/ports-status'):
                return [{'portNumber': '1', 'linkSpeed': '1000M'},   # WAN, up
                        {'portNumber': '2', 'linkSpeed': None},       # WAN, down
                        {'portNumber': '13', 'linkSpeed': '1000M'}]   # LAN, ignored
            return None

        with self._settings(), patch('oe_inventory_py_web.nebula._request', side_effect=fake_request):
            w = nebula._gateway_wan('site1', 'dev1')
        self.assertEqual(w['enabled'], 2)
        self.assertEqual(w['operational'], 1)    # only ge1 has a live link

    def test_gateway_system_returns_cpu_and_mem(self):
        from unittest.mock import patch
        from oe_inventory_py_web import nebula
        resp = {'cpuUsage': 22, 'memUsage': 69, 'sessions': 260}
        with patch('oe_inventory_py_web.nebula._request', return_value=resp):
            m = nebula._gateway_system('site1', 'dev1')
        self.assertEqual(m, {'cpu': 22, 'mem': 69})

    def test_gateway_system_empty_when_unavailable(self):
        from unittest.mock import patch
        from oe_inventory_py_web import nebula
        with patch('oe_inventory_py_web.nebula._request', side_effect=nebula.NebulaError('500')):
            self.assertEqual(nebula._gateway_system('site1', 'dev1'), {})


class NebulaFirmwareTests(TestCase):
    """Outdated-firmware detection (per-type counters + alerts)."""

    def test_firmware_status_maps_devid_to_info(self):
        from unittest.mock import patch
        from oe_inventory_py_web import nebula
        data = [
            {'devId': 'a', 'currentVersion': '1.0', 'latestVersion': '1.0', 'status': 'UP_TO_DATE'},
            {'devId': 'b', 'currentVersion': '1.0', 'latestVersion': '2.0', 'status': 'NOT_UP_TO_DATE'},
        ]
        with patch('oe_inventory_py_web.nebula._request', return_value=data):
            m = nebula._firmware_status('site1')
        self.assertEqual(m['b']['status'], 'NOT_UP_TO_DATE')
        self.assertEqual(m['b']['latest'], '2.0')

    def test_device_stats_counts_outdated_per_family(self):
        from oe_inventory_py_web import nebula
        devices = [
            {'devId': 's1', 'type': 'SW'}, {'devId': 's2', 'type': 'SW'},
            {'devId': 'a1', 'type': 'AP'},
            {'devId': 'g1', 'type': 'GWH'},
        ]
        online = {'s1': True, 's2': True, 'a1': True, 'g1': True}
        outdated = {'s2', 'a1'}
        sw, ap, fw, offline = nebula._device_stats(devices, online, outdated)
        self.assertEqual(sw['outdated'], 1)   # s2
        self.assertEqual(ap['outdated'], 1)   # a1
        self.assertEqual(fw['outdated'], 0)
        self.assertEqual(offline, [])


class NebulaClientCountsTests(TestCase):
    """nebula._online_clients / _client_counts: only ONLINE clients, split wifi/wired."""

    BY_ID = {'ap1': {'devId': 'ap1', 'type': 'AP'},
             'sw1': {'devId': 'sw1', 'type': 'SW'},
             'gw1': {'devId': 'gw1', 'type': 'GWH'}}

    def test_online_clients_filters_offline(self):
        from unittest.mock import patch
        from oe_inventory_py_web import nebula
        resp = {'data': [
            {'status': 'ONLINE', 'connectedTo': 'ap1'},
            {'status': 'OFFLINE', 'connectedTo': 'ap1'},
            {'status': 'ONLINE', 'connectedTo': 'sw1'},
        ]}
        with patch('oe_inventory_py_web.nebula._request', return_value=resp):
            clients = nebula._online_clients('site1')
        self.assertEqual(len(clients), 2)
        self.assertTrue(all(c['status'] == 'ONLINE' for c in clients))

    def test_online_clients_none_when_endpoint_fails(self):
        from unittest.mock import patch
        from oe_inventory_py_web import nebula
        with patch('oe_inventory_py_web.nebula._request', side_effect=nebula.NebulaError('boom')):
            self.assertIsNone(nebula._online_clients('site1'))

    def test_counts_split_by_connected_device(self):
        from oe_inventory_py_web import nebula
        clients = [
            {'macAddress': 'AA', 'connectedTo': 'ap1'},
            {'macAddress': 'BB', 'connectedTo': 'ap1'},   # wifi
            {'macAddress': 'CC', 'connectedTo': 'sw1'},   # wired
            {'macAddress': 'DD', 'connectedTo': 'gw1'},   # wired (gateway)
            {'macAddress': 'EE', 'connectedTo': 'unknown'},  # wired (fallback)
        ]
        self.assertEqual(nebula._client_counts(clients, self.BY_ID),
                         {'wifi': 2, 'wired': 3, 'total': 5})

    def test_dedupes_same_device_across_connections(self):
        from oe_inventory_py_web import nebula
        # Same MAC seen on two switches + once on an AP -> 1 wired, 1 wifi, 1 device.
        clients = [
            {'macAddress': 'AA', 'connectedTo': 'sw1'},
            {'macAddress': 'AA', 'connectedTo': 'sw1'},
            {'macAddress': 'AA', 'connectedTo': 'ap1'},
            {'macAddress': 'BB', 'connectedTo': 'sw1'},
        ]
        self.assertEqual(nebula._client_counts(clients, self.BY_ID),
                         {'wifi': 1, 'wired': 2, 'total': 2})

    def test_counts_none_when_clients_unavailable(self):
        from oe_inventory_py_web import nebula
        self.assertEqual(nebula._client_counts(None, self.BY_ID),
                         {'wifi': None, 'wired': None, 'total': None})

    def test_build_topology_groups_and_counts(self):
        from oe_inventory_py_web import nebula
        devices = [
            {'devId': 'sw1', 'type': 'SW', 'name': 'SW-A', 'model': 'X1'},
            {'devId': 'ap1', 'type': 'AP', 'name': 'AP-A', 'model': 'W1'},
            {'devId': 'gw1', 'type': 'GWH', 'name': 'FW-A', 'model': 'USG'},
        ]
        online = {'sw1': True, 'ap1': False, 'gw1': True}
        clients = [{'connectedTo': 'ap1'}, {'connectedTo': 'sw1'}, {'connectedTo': 'sw1'}]
        topo = nebula._build_topology('Site X', devices, online, clients)
        self.assertEqual(topo['site'], 'Site X')
        self.assertEqual(len(topo['gateways']), 1)
        self.assertEqual(topo['switches'][0]['clients'], 2)
        self.assertFalse(topo['aps'][0]['online'])
        # No metrics for switches/APs (API doesn't expose them).
        self.assertEqual(topo['switches'][0]['metrics'], [])

    def test_build_topology_attaches_gateway_metrics(self):
        from oe_inventory_py_web import nebula
        devices = [{'devId': 'gw1', 'type': 'GWH', 'name': 'FW', 'model': 'USG'}]
        online = {'gw1': True}
        gw_metrics = {'gw1': {'cpu': 22.0, 'mem': 69.4}}
        topo = nebula._build_topology('S', devices, online, [], gw_metrics)
        self.assertEqual(topo['gateways'][0]['metrics'],
                         [{'label': 'CPU', 'value': 22}, {'label': 'Memory', 'value': 69}])


class AnyDeskScreenTests(TestCase):
    """AnyDesk screen: cards with description + online/offline dot (gated by
    net_overview). The check itself runs in the background (status_cache)."""

    def setUp(self):
        from django.db import connection
        from django.core.cache import cache
        cache.clear()
        with connection.cursor() as c:
            c.execute("CREATE TABLE IF NOT EXISTS oees_anydesk "
                      "(id_anydesk INTEGER PRIMARY KEY, code VARCHAR(50), "
                      "description VARCHAR(100), last_connection DATETIME NULL)")
            c.execute("DELETE FROM oees_anydesk")
            c.execute("INSERT INTO oees_anydesk VALUES (1, 'AD-111', 'Reception PC', '2026-06-22 10:00:00')")
            c.execute("INSERT INTO oees_anydesk VALUES (2, 'AD-222', 'Warehouse PC', NULL)")
        User = get_user_model()
        self.user = User.objects.create_user(
            username='ad_user', password='pass12345', net_overview=1, reader=0)

    def test_requires_login(self):
        self.assertEqual(self.client.get(reverse('frm_remote_machines')).status_code, 302)

    def test_forbidden_without_permission(self):
        User = get_user_model()
        other = User.objects.create_user(
            username='ad_noperm', password='pass12345', net_overview=0, reader=1)
        self.client.force_login(other)
        self.assertRedirects(self.client.get(reverse('frm_remote_machines')), reverse('mdi_home'))

    def test_lists_machines_as_cards(self):
        self.client.force_login(self.user)
        resp = self.client.get(reverse('frm_remote_machines'))
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.context['available'])
        self.assertEqual(len(resp.context['machines']), 2)
        self.assertContains(resp, 'Reception PC')
        self.assertContains(resp, 'Warehouse PC')
        self.assertContains(resp, 'AD-111')

    def test_status_falls_back_to_last_connection_without_check(self):
        # No background status cached -> dot from last_connection presence.
        self.client.force_login(self.user)
        resp = self.client.get(reverse('frm_remote_machines'))
        by_code = {m['code']: m['online'] for m in resp.context['machines']}
        self.assertTrue(by_code['AD-111'])    # has last_connection -> green
        self.assertFalse(by_code['AD-222'])   # never connected -> red

    def test_uses_background_status_when_available(self):
        from django.core.cache import cache
        from oe_inventory_py_web import status_cache
        # A real check result overrides the last_connection fallback.
        cache.set(status_cache.ANYDESK_STATUS_KEY, {'AD-111': False, 'AD-222': True}, None)
        self.client.force_login(self.user)
        resp = self.client.get(reverse('frm_remote_machines'))
        by_code = {m['code']: m['online'] for m in resp.context['machines']}
        self.assertFalse(by_code['AD-111'])
        self.assertTrue(by_code['AD-222'])


class StatusCacheAnyDeskTests(TestCase):
    """Background AnyDesk check (status_cache._anydesk_check)."""

    def setUp(self):
        from django.db import connection
        from django.core.cache import cache
        cache.clear()
        with connection.cursor() as c:
            c.execute("CREATE TABLE IF NOT EXISTS oees_anydesk "
                      "(id_anydesk INTEGER PRIMARY KEY, code VARCHAR(50), "
                      "description VARCHAR(100), last_connection DATETIME NULL)")
            c.execute("DELETE FROM oees_anydesk")
            c.execute("INSERT INTO oees_anydesk VALUES (1, 'AD-111', 'PC1', NULL)")
            c.execute("INSERT INTO oees_anydesk VALUES (2, 'AD-222', 'PC2', NULL)")

    def test_not_configured_falls_back_to_last_connection(self):
        # Without the API key, machines without a last_connection count as
        # offline (matching the screen's dots), so the footer still shows alerts.
        from oe_inventory_py_web import status_cache
        alerts, status = status_cache._anydesk_check()
        self.assertEqual(alerts, 2)   # both AD-111 and AD-222 have NULL last_connection
        self.assertEqual(status, {'AD-111': False, 'AD-222': False})

    def test_table_missing_returns_none(self):
        from django.db import connection
        from oe_inventory_py_web import status_cache
        with connection.cursor() as c:
            c.execute("DROP TABLE oees_anydesk")
        self.assertEqual(status_cache._anydesk_check(), (None, {}))

    def test_counts_offline_and_stamps_online(self):
        from django.test import override_settings
        from django.db import connection
        from unittest.mock import patch
        from oe_inventory_py_web import status_cache
        with override_settings(ANYDESK_API_TOKEN='token'):
            with patch('oe_inventory_py_web.anydesk.online_map',
                       return_value={'AD-111': True, 'AD-222': False}):
                alerts, status = status_cache._anydesk_check()
        self.assertEqual(alerts, 1)                       # AD-222 offline
        self.assertEqual(status, {'AD-111': True, 'AD-222': False})
        with connection.cursor() as c:
            c.execute("SELECT last_connection FROM oees_anydesk WHERE code='AD-111'")
            self.assertIsNotNone(c.fetchone()[0])         # online -> stamped
            c.execute("SELECT last_connection FROM oees_anydesk WHERE code='AD-222'")
            self.assertIsNone(c.fetchone()[0])            # offline -> untouched

    def test_api_error_falls_back_to_last_connection(self):
        # A configured-but-failing API (e.g. bad/invalid key) must NOT hide the
        # badge: it falls back to the last_connection logic so the footer still
        # reflects the unreachable machines.
        from django.test import override_settings
        from unittest.mock import patch
        from oe_inventory_py_web import status_cache, anydesk
        with override_settings(ANYDESK_API_TOKEN='token'):
            with patch('oe_inventory_py_web.anydesk.online_map',
                       side_effect=anydesk.AnydeskError('boom')):
                alerts, status = status_cache._anydesk_check({'anydesk_alerts': 5})
        self.assertEqual(alerts, 2)   # both machines have NULL last_connection -> offline
        self.assertEqual(status, {'AD-111': False, 'AD-222': False})


class StatusCacheOmadaMergeTests(TestCase):
    """status_cache._apply_omada: fold Omada per-site data into Nebula rows."""

    def _rows(self):
        return [
            {'site': 'DOCTOR ROMAGOSA',
             'switches': {'total': 0, 'online': 0, 'offline': 0, 'outdated': 0},
             'aps': {'total': 0, 'online': 0, 'offline': 0, 'outdated': 0},
             'clients': {'total': 5, 'wifi': 2, 'wired': 3},
             'alerts': 0, 'alert_list': [],
             'topology': {'gateways': [{'name': 'FW'}], 'switches': [], 'aps': []}},
            {'site': 'OTHER SITE',
             'switches': {'total': 1, 'online': 1, 'offline': 0, 'outdated': 0},
             'aps': {'total': 0, 'online': 0, 'offline': 0, 'outdated': 0},
             'clients': {'total': 9, 'wifi': 9, 'wired': 0},
             'alerts': 0, 'alert_list': [], 'topology': {}},
        ]

    def _details(self):
        return [
            {'name': 'Octopus Valencia',
             'switches': {'total': 11, 'online': 10, 'offline': 1, 'outdated': 2},
             'aps': {'total': 7, 'online': 7, 'offline': 0, 'outdated': 1},
             'clients': {'total': 240, 'wifi': 130, 'wired': 110},
             'offline_devices': [{'name': 'CORE 10G', 'issue': 'Offline'}],
             'outdated_devices': [{'name': 'SW01', 'issue': 'Outdated firmware'},
                                  {'name': 'SW02', 'issue': 'Outdated firmware'},
                                  {'name': 'AP01', 'issue': 'Outdated firmware'}],
             'topology': {'switches': [{'name': 'CORE 10G', 'online': False}],
                          'aps': [{'name': 'AP01', 'online': True}]}},
            {'name': 'Unmapped Site', 'switches': {'total': 3, 'online': 3, 'offline': 0, 'outdated': 0},
             'aps': {'total': 0, 'online': 0, 'offline': 0, 'outdated': 0},
             'clients': {'total': 1, 'wifi': 1, 'wired': 0}, 'offline_devices': [],
             'outdated_devices': [], 'topology': {'switches': [], 'aps': []}},
        ]

    def test_merge_folds_into_mapped_site(self):
        from oe_inventory_py_web import status_cache
        rows = self._rows()
        status_cache._apply_omada(rows, self._details(),
                                  {'Octopus Valencia': 'DOCTOR ROMAGOSA'})
        dr = next(r for r in rows if r['site'] == 'DOCTOR ROMAGOSA')
        # switches/APs (incl. outdated) added on top of the (zero) Nebula counts
        self.assertEqual(dr['switches'], {'total': 11, 'online': 10, 'offline': 1, 'outdated': 2})
        self.assertEqual(dr['aps']['total'], 7)
        self.assertEqual(dr['aps']['outdated'], 1)
        # clients replaced with Omada's (no double counting)
        self.assertEqual(dr['clients'], {'total': 240, 'wifi': 130, 'wired': 110})
        # offline (1) + outdated-firmware (3) devices all became alerts
        self.assertEqual(dr['alerts'], 4)
        self.assertEqual(len(dr['alert_list']), 4)
        # topology tiers extended (gateways kept from Nebula)
        self.assertEqual([d['name'] for d in dr['topology']['switches']], ['CORE 10G'])
        self.assertEqual([d['name'] for d in dr['topology']['aps']], ['AP01'])
        self.assertEqual(dr['topology']['gateways'], [{'name': 'FW'}])
        self.assertTrue(dr['omada'])

    def test_merge_iterates_controllers_with_dedupe(self):
        from django.test import override_settings
        from unittest.mock import patch
        from oe_inventory_py_web import status_cache

        def blank(total, sw, ap, clients):
            return {'switches': {'total': sw, 'online': sw, 'offline': 0, 'outdated': 0},
                    'aps': {'total': ap, 'online': ap, 'offline': 0, 'outdated': 0},
                    'clients': clients}
        rows = [
            {'site': 'DOCTOR ROMAGOSA', **blank(0, 0, 0, {'total': 0, 'wifi': 0, 'wired': 0}),
             'alerts': 0, 'alert_list': [], 'topology': {}},
            {'site': 'MADRID', **blank(0, 0, 0, {'total': 0, 'wifi': 0, 'wired': 0}),
             'alerts': 0, 'alert_list': [], 'topology': {}},
        ]
        val = {'name': 'Octopus Valencia', 'site_id': 'v1', 'offline_devices': [],
               'outdated_devices': [], 'topology': {'switches': [], 'aps': []},
               **blank(0, 5, 2, {'total': 50, 'wifi': 30, 'wired': 20})}
        mad = {'name': 'Octopus Madrid', 'site_id': 'm1', 'offline_devices': [],
               'outdated_devices': [], 'topology': {'switches': [], 'aps': []},
               **blank(0, 3, 1, {'total': 10, 'wifi': 10, 'wired': 0})}
        # Controller A listed twice (same site) must merge once; B is a 2nd site.
        controllers = [{'omadac_id': 'A', 'base_url': 'a'},
                       {'omadac_id': 'A', 'base_url': 'a'},
                       {'omadac_id': 'B', 'base_url': 'b'}]
        site_map = {'Octopus Valencia': 'DOCTOR ROMAGOSA', 'Octopus Madrid': 'MADRID'}
        with override_settings(OMADA_NEBULA_SITE_MAP=site_map):
            with patch('oe_inventory_py_web.omada.controllers', return_value=controllers), \
                 patch('oe_inventory_py_web.omada.site_details',
                       side_effect=lambda creds: [val] if creds['omadac_id'] == 'A' else [mad]):
                status_cache._merge_omada(rows)
        dr = next(r for r in rows if r['site'] == 'DOCTOR ROMAGOSA')
        mr = next(r for r in rows if r['site'] == 'MADRID')
        # Valencia merged once despite controller A appearing twice (dedupe).
        self.assertEqual(dr['switches']['total'], 5)
        self.assertEqual(dr['clients']['total'], 50)
        # The second controller's site folded into the Madrid card.
        self.assertEqual(mr['switches']['total'], 3)
        self.assertEqual(mr['clients']['total'], 10)

    def test_unmapped_and_other_sites_untouched(self):
        from oe_inventory_py_web import status_cache
        rows = self._rows()
        status_cache._apply_omada(rows, self._details(),
                                  {'Octopus Valencia': 'DOCTOR ROMAGOSA'})
        other = next(r for r in rows if r['site'] == 'OTHER SITE')
        self.assertEqual(other['switches'], {'total': 1, 'online': 1, 'offline': 0, 'outdated': 0})
        self.assertEqual(other['clients'], {'total': 9, 'wifi': 9, 'wired': 0})
        self.assertNotIn('omada', other)
        # The unmapped Omada site created no new row.
        self.assertEqual({r['site'] for r in rows}, {'DOCTOR ROMAGOSA', 'OTHER SITE'})


class LogitechClientTests(TestCase):
    """logitech.rooms_overview: parse the Sync Cloud /places payload."""

    def test_rooms_overview_parses_places(self):
        from unittest.mock import patch
        from oe_inventory_py_web import logitech
        payload = {'places': [{
            'id': 'r1', 'name': 'Sala de Juntas',
            'insights': {'isOccupied': True, 'occupancyCount': 4, 'inMeeting': True},
            'devices': [{'model': 'Rally Bar Mini', 'firmwareVersion': 'CollabOS 1.12.x',
                         'status': 'connected'}],
        }]}
        with patch('oe_inventory_py_web.logitech._request', return_value=payload):
            rooms = logitech.rooms_overview()
        self.assertEqual(len(rooms), 1)
        r = rooms[0]
        self.assertEqual(r['name'], 'Sala de Juntas')
        self.assertTrue(r['in_meeting'] and r['occupied'] and r['connected'])
        self.assertEqual(r['occupancy'], 4)
        self.assertEqual(r['devices'][0]['model'], 'Rally Bar Mini')

    def test_rooms_overview_raises_on_error(self):
        from unittest.mock import patch
        from oe_inventory_py_web import logitech
        with patch('oe_inventory_py_web.logitech._request', side_effect=ValueError('boom')):
            with self.assertRaises(logitech.LogitechError):
                logitech.rooms_overview()

    def test_organizer_title_and_occupied_empty_alert(self):
        from unittest.mock import patch
        from oe_inventory_py_web import logitech
        payload = {'places': [{
            'id': 'r1', 'name': 'Sala', 'devices': [],
            'insights': {'isOccupied': True, 'inMeeting': True, 'occupancyCount': 0,
                         'meetingTitle': 'Daily', 'organizer': {'name': 'Ana García'}},
        }]}
        with patch('oe_inventory_py_web.logitech._request', return_value=payload):
            r = logitech.rooms_overview()[0]
        self.assertEqual(r['title'], 'Daily')
        self.assertEqual(r['organizer'], 'Ana García')   # extracted from {name: ...}
        self.assertTrue(r['alert'])                       # occupied/in-meeting but 0 people

    def test_disconnected_room_is_alert(self):
        from unittest.mock import patch
        from oe_inventory_py_web import logitech
        payload = {'places': [{
            'id': 'r1', 'name': 'Sala', 'insights': {'isOccupied': False, 'occupancyCount': 0},
            'devices': [{'model': 'Rally Bar', 'status': 'disconnected'}],
        }]}
        with patch('oe_inventory_py_web.logitech._request', return_value=payload):
            r = logitech.rooms_overview()[0]
        self.assertFalse(r['connected'])
        self.assertTrue(r['alert'])     # disconnected -> alert even if free

    def test_parse_dt_iso_and_epoch(self):
        import datetime
        from oe_inventory_py_web import logitech
        self.assertIsNone(logitech._parse_dt(''))
        self.assertEqual(logitech._parse_dt('2026-06-25T09:00:00+00:00'),
                         datetime.datetime(2026, 6, 25, 9, 0, tzinfo=datetime.timezone.utc))
        # epoch milliseconds
        self.assertEqual(logitech._parse_dt(1782896400000),
                         datetime.datetime.fromtimestamp(1782896400, tz=datetime.timezone.utc))

    def test_no_alert_when_people_present_or_unknown(self):
        from oe_inventory_py_web import logitech
        # occupancy unknown (None) -> not an alert even if occupied
        self.assertFalse(logitech._is_occupied_but_empty(True, False, None))
        # people present -> not an alert
        self.assertFalse(logitech._is_occupied_but_empty(True, True, 3))
        # free + 0 -> not an alert
        self.assertFalse(logitech._is_occupied_but_empty(False, False, 0))
        # occupied + 0 -> alert
        self.assertTrue(logitech._is_occupied_but_empty(True, False, 0))


class VideoRoomsScreenTests(TestCase):
    """Video Rooms screen (Logitech Sync; gated by net_overview, client mocked)."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(
            username='vr_user', password='pass12345', net_overview=1, reader=0)

    def test_requires_login(self):
        self.assertEqual(self.client.get(reverse('frm_video_rooms')).status_code, 302)

    def test_forbidden_without_permission(self):
        User = get_user_model()
        other = User.objects.create_user(
            username='vr_noperm', password='pass12345', net_overview=0, reader=1)
        self.client.force_login(other)
        self.assertRedirects(self.client.get(reverse('frm_video_rooms')), reverse('mdi_home'))

    def test_not_configured_shows_demo_rooms(self):
        # In test mode the LOGITECH cert/key paths are blanked -> not configured,
        # so the screen shows sample rooms (with a "Sample data" notice) so the
        # design can be worked on before the certificate is available.
        self.client.force_login(self.user)
        resp = self.client.get(reverse('frm_video_rooms'))
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(resp.context['configured'])
        self.assertTrue(resp.context['demo'])
        self.assertContains(resp, 'Sample data')
        self.assertContains(resp, 'Sala de Juntas')   # a demo room
        # Right panel: future-booking incidences (the demo includes an unknown email).
        self.assertContains(resp, 'Future-booking incidences')
        self.assertContains(resp, 'is not found in the users table')
        # Right panel: under-used meetings table (demo rows).
        self.assertContains(resp, 'Under-used meetings')
        self.assertContains(resp, 'Comité de Dirección')
        # Right panel: organizers no-show ranking (demo rows).
        self.assertContains(resp, 'Organizers ranking')
        self.assertContains(resp, 'Ana García')
        # Room cards show the scheduled meeting's start–end time.
        self.assertContains(resp, '09:00')
        self.assertContains(resp, '10:00')

    def test_renders_rooms_when_configured(self):
        from unittest.mock import patch
        rows = [{
            'id': 'r1', 'name': 'Sala de Juntas', 'occupied': True, 'in_meeting': True,
            'occupancy': 0, 'connected': True, 'alert': True,
            'organizer': 'Ana García', 'title': 'Comité',
            'devices': [{'model': 'Rally Bar Mini', 'firmware': 'CollabOS 1.12.x', 'status': 'connected'}],
        }]
        with patch('oe_inventory_py_web.logitech.logitech_configured', return_value=True), \
             patch('oe_inventory_py_web.logitech.rooms_overview', return_value=rows):
            self.client.force_login(self.user)
            resp = self.client.get(reverse('frm_video_rooms'))
            self.assertTrue(resp.context['configured'])
            self.assertContains(resp, 'Sala de Juntas')
            self.assertContains(resp, 'Rally Bar Mini')
            self.assertContains(resp, 'In meeting')
            self.assertContains(resp, 'Ana García')          # organizer shown
            self.assertContains(resp, 'Comité')              # meeting title shown
            self.assertContains(resp, 'Occupied with no people')  # alert highlighted

    def test_video_rooms_alert_count_from_demo(self):
        # Not configured -> demo rooms; alerts = occupied-but-empty + disconnected
        # (Sala Formación occupied with 0 people, Sala Goya disconnected) -> 2.
        from oe_inventory_py_web import status_cache
        alerts, rooms, is_real = status_cache._video_rooms_check()
        self.assertEqual(alerts, 2)
        self.assertFalse(is_real)        # demo data -> not persisted
        self.assertTrue(rooms)

    def test_track_meetings_inserts_then_increments(self):
        from oe_inventory_py_web import status_cache
        from oe_inventory_py_web.models import OeesMeetingRoom
        import datetime
        start = datetime.datetime(2026, 6, 25, 9, 0, tzinfo=datetime.timezone.utc)
        end = datetime.datetime(2026, 6, 25, 10, 0, tzinfo=datetime.timezone.utc)  # 60 min reserved
        occupied = [{'meet_id': 'M1', 'title': 'Daily', 'organizer_email': 'a@x.com',
                     'occupied': True, 'start_time': start, 'end_time': end}]
        # 1st cycle: insert. duration = initial reserved length (60); occupied 0.
        status_cache._track_meetings(occupied)
        m = OeesMeetingRoom.objects.get(meet_id='M1')
        self.assertEqual(m.duration, 60)
        self.assertEqual(m.occupied, 0)
        self.assertEqual(m.description, 'Daily')
        self.assertEqual(m.org_email, 'a@x.com')
        self.assertEqual(m.start_time, start)
        self.assertEqual(m.end_time, end)
        # 2nd cycle, occupied: occupied +5, duration unchanged.
        status_cache._track_meetings(occupied)
        m = OeesMeetingRoom.objects.get(meet_id='M1')
        self.assertEqual(m.duration, 60)
        self.assertEqual(m.occupied, 5)
        # 3rd cycle, NOT occupied: no change to occupied (or duration).
        status_cache._track_meetings([{'meet_id': 'M1', 'occupied': False}])
        m = OeesMeetingRoom.objects.get(meet_id='M1')
        self.assertEqual(m.duration, 60)
        self.assertEqual(m.occupied, 5)

    def test_track_meetings_skips_rooms_without_meeting_id(self):
        from oe_inventory_py_web import status_cache
        from oe_inventory_py_web.models import OeesMeetingRoom
        status_cache._track_meetings([{'meet_id': '', 'occupied': True}])
        self.assertEqual(OeesMeetingRoom.objects.count(), 0)

    def test_organizer_no_show_ranking(self):
        from oe_inventory_py_web.models import OeesMeetingRoom, OeesStaff
        from oe_inventory_py_web.views import _organizer_no_show_ranking
        OeesStaff.objects.create(name='Ana García', notes='', persona_fisica=1, email='ana@x.com')
        # ana: 2 meetings with occupied <= 10; bob: 1; ana also has 1 that doesn't count.
        OeesMeetingRoom.objects.create(meet_id='1', org_email='ana@x.com', duration=60, occupied=0)
        OeesMeetingRoom.objects.create(meet_id='2', org_email='ana@x.com', duration=60, occupied=10)
        OeesMeetingRoom.objects.create(meet_id='3', org_email='ana@x.com', duration=60, occupied=40)  # excluded
        OeesMeetingRoom.objects.create(meet_id='4', org_email='bob@x.com', duration=60, occupied=5)
        ranking = _organizer_no_show_ranking()
        # ana first (2), resolved to her staff name; bob second (1), shown as email.
        # Totals sum duration/occupied over the counted (occupied <= 10) meetings.
        self.assertEqual(ranking[0], {'organizer': 'Ana García', 'count': 2,
                                      'total_duration': 120, 'total_occupied': 10})
        self.assertEqual(ranking[1], {'organizer': 'bob@x.com', 'count': 1,
                                      'total_duration': 60, 'total_occupied': 5})

    def test_low_occupancy_meetings_filter_and_pct(self):
        import datetime
        from oe_inventory_py_web.models import OeesMeetingRoom
        from oe_inventory_py_web.views import _low_occupancy_meetings
        start = datetime.datetime(2026, 6, 24, 9, 0, tzinfo=datetime.timezone.utc)
        end = datetime.datetime(2026, 6, 24, 10, 0, tzinfo=datetime.timezone.utc)
        OeesMeetingRoom.objects.create(meet_id='A', description='Low', org_email='a@x.com',
                                       duration=60, occupied=12, start_time=start, end_time=end)  # 20%
        OeesMeetingRoom.objects.create(meet_id='B', description='Half', org_email='b@x.com',
                                       duration=60, occupied=30, start_time=start)   # 50%, no end_time
        OeesMeetingRoom.objects.create(meet_id='C', description='High', org_email='c@x.com',
                                       duration=60, occupied=45, start_time=start)   # 75% -> excluded
        OeesMeetingRoom.objects.create(meet_id='D', description='NoDur', org_email='d@x.com',
                                       duration=0, occupied=0, start_time=start)     # duration 0 -> excluded
        rows = _low_occupancy_meetings()
        by_title = {r['title']: r for r in rows}
        self.assertEqual(set(by_title), {'Low', 'Half'})
        self.assertEqual(by_title['Low']['pct'], 20)
        self.assertEqual(by_title['Half']['pct'], 50)
        self.assertEqual(rows[0]['date'], '24-06-2026')
        # 10:00 UTC shown in Madrid time (CEST, +2 in June) -> 12:00.
        self.assertEqual(by_title['Low']['end'], '12:00')   # end time when known
        self.assertIsNone(by_title['Half']['end'])          # falls back to duration
        self.assertEqual(by_title['Half']['duration'], 60)

    def test_booking_incidences_deactivated_and_unknown(self):
        from oe_inventory_py_web.models import OeesStaff
        from oe_inventory_py_web.views import _booking_incidences
        OeesStaff.objects.create(name='Jane Doe', notes='', persona_fisica=1,
                                 email='jane@x.com', fecha_baja='06-10-2025', state=0)
        OeesStaff.objects.create(name='Active Al', notes='', persona_fisica=1,
                                 email='al@x.com', fecha_baja='', state=1)
        bookings = [
            {'organizer_email': 'Jane@x.com'}, {'organizer_email': 'jane@x.com'},  # 2, case-insensitive
            {'organizer_email': 'al@x.com'},        # active -> no incidence
            {'organizer_email': 'ghost@x.com'},     # not in staff
        ]
        joined = ' | '.join(_booking_incidences(bookings))
        self.assertIn('User Jane Doe, deactivated since 06-10-2025, has 2 future bookings', joined)
        self.assertIn('The email ghost@x.com is not found in the users table', joined)
        self.assertNotIn('al@x.com', joined)    # active organizer produces no line

    def test_shows_error_when_api_fails(self):
        from unittest.mock import patch
        with patch('oe_inventory_py_web.logitech.logitech_configured', return_value=True), \
             patch('oe_inventory_py_web.logitech.rooms_overview', side_effect=Exception('boom')):
            self.client.force_login(self.user)
            resp = self.client.get(reverse('frm_video_rooms'))
            self.assertIsNotNone(resp.context['error'])
            self.assertContains(resp, 'Could not reach the Logitech Sync Cloud API')


class OmadaScreenTests(TestCase):
    """Omada sites-overview screen (server-rendered; API client mocked)."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(
            username='om_user', password='pass12345', omada=1, reader=0)

    def test_requires_login(self):
        self.assertEqual(self.client.get(reverse('frm_omada')).status_code, 302)

    def test_not_configured_shows_notice(self):
        from django.test import override_settings
        with override_settings(OMADA_BASE_URL='', OMADA_OMADAC_ID='',
                               OMADA_CLIENT_ID='', OMADA_CLIENT_SECRET=''):
            self.client.force_login(self.user)
            resp = self.client.get(reverse('frm_omada'))
            self.assertEqual(resp.status_code, 200)
            self.assertFalse(resp.context['configured'])
            self.assertContains(resp, 'not configured')

    def test_renders_site_overview_when_configured(self):
        from django.test import override_settings
        from unittest.mock import patch
        rows = [{'site': 'HQ', 'devices': 5, 'clients': 12}]
        with override_settings(OMADA_BASE_URL='https://x', OMADA_OMADAC_ID='id',
                               OMADA_CLIENT_ID='c', OMADA_CLIENT_SECRET='s'):
            with patch('oe_inventory_py_web.omada.site_overview', return_value=rows):
                self.client.force_login(self.user)
                resp = self.client.get(reverse('frm_omada'))
                self.assertTrue(resp.context['configured'])
                self.assertContains(resp, 'HQ')

    def test_shows_error_when_api_fails(self):
        from django.test import override_settings
        from unittest.mock import patch
        with override_settings(OMADA_BASE_URL='https://x', OMADA_OMADAC_ID='id',
                               OMADA_CLIENT_ID='c', OMADA_CLIENT_SECRET='s'):
            with patch('oe_inventory_py_web.omada.site_overview', side_effect=Exception('boom')):
                self.client.force_login(self.user)
                resp = self.client.get(reverse('frm_omada'))
                self.assertIsNotNone(resp.context['error'])
                self.assertContains(resp, 'Could not reach the Omada controller')


class NetOverviewScreenTests(TestCase):
    """Net Overview dashboard (Nebula-backed, AJAX + spinner; client mocked)."""

    def setUp(self):
        from django.core.cache import cache
        cache.clear()  # the screen now reads cached site_overview rows
        User = get_user_model()
        self.user = User.objects.create_user(
            username='no_user', password='pass12345', net_overview=1, reader=0)

    def test_requires_login(self):
        self.assertEqual(self.client.get(reverse('frm_net_overview')).status_code, 302)

    def test_shell_loads(self):
        # The initial page is a lightweight shell; data arrives via AJAX.
        self.client.force_login(self.user)
        resp = self.client.get(reverse('frm_net_overview'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'net-overview-content')

    def test_partial_not_configured_shows_notice(self):
        from django.test import override_settings
        with override_settings(NEBULA_BASE_URL='', NEBULA_API_KEY='', NEBULA_ORG_ID=''):
            self.client.force_login(self.user)
            resp = self.client.get(reverse('frm_net_overview'), {'partial': '1'})
            self.assertEqual(resp.status_code, 200)
            self.assertFalse(resp.context['configured'])
            self.assertContains(resp, 'not configured')

    def test_partial_renders_site_overview_when_configured(self):
        from django.test import override_settings
        from unittest.mock import patch
        rows = [{
            'site': 'Madrid',
            'wan': {'enabled': 2, 'operational': None},
            'switches': {'total': 2, 'online': 2, 'offline': 0, 'outdated': 0},
            'aps': {'total': 3, 'online': 3, 'offline': 0, 'outdated': 1},
            'firewalls': {'total': 1, 'online': 1, 'offline': 0, 'outdated': 0},
            'clients': {'wifi': 10, 'wired': 5, 'total': 15},
            'outdated': 1,
            'alerts': 1,
        }]
        with override_settings(NEBULA_BASE_URL='https://x', NEBULA_API_KEY='t', NEBULA_ORG_ID='o'):
            with patch('oe_inventory_py_web.nebula.site_overview', return_value=rows):
                self.client.force_login(self.user)
                resp = self.client.get(reverse('frm_net_overview'), {'partial': '1'})
                self.assertTrue(resp.context['configured'])
                self.assertContains(resp, 'Madrid')
                self.assertContains(resp, 'Access Points')
                self.assertContains(resp, 'Firewalls')
                self.assertContains(resp, 'WAN')
                self.assertContains(resp, 'enabled')
                # operational unknown (ports-status 500) -> shows a dash, not 0
                self.assertContains(resp, 'operational')
                # orange outdated-firmware counter on the AP panel
                self.assertContains(resp, '1 outdated')

    def test_partial_shows_error_when_api_fails(self):
        from django.test import override_settings
        from unittest.mock import patch
        with override_settings(NEBULA_BASE_URL='https://x', NEBULA_API_KEY='t', NEBULA_ORG_ID='o'):
            with patch('oe_inventory_py_web.nebula.site_overview', side_effect=Exception('boom')):
                self.client.force_login(self.user)
                resp = self.client.get(reverse('frm_net_overview'), {'partial': '1'})
                self.assertIsNotNone(resp.context['error'])
                self.assertContains(resp, 'Could not reach the Nebula API')

    def test_partial_returns_202_while_preparing(self):
        # Cold start: cache not ready yet -> 202 so the shell keeps the spinner
        # and retries, instead of computing synchronously (which 502'd on EB).
        from django.test import override_settings
        from unittest.mock import patch
        with override_settings(NEBULA_BASE_URL='https://x', NEBULA_API_KEY='t'):
            with patch('oe_inventory_py_web.status_cache.get_net_overview', return_value=(None, None)):
                self.client.force_login(self.user)
                resp = self.client.get(reverse('frm_net_overview'), {'partial': '1'})
                self.assertEqual(resp.status_code, 202)

    def test_alerts_excel_export(self):
        from unittest.mock import patch
        import io
        import openpyxl
        rows = [{'site': 'Madrid', 'alert_list': [
            {'name': 'SW1', 'type': 'SWITCH', 'model': 'X1', 'mac': 'AA-BB',
             'issue': 'Outdated firmware', 'detail': '1.0 -> 2.0'}]}]
        with patch('oe_inventory_py_web.status_cache.get_net_overview', return_value=(rows, None)):
            self.client.force_login(self.user)
            resp = self.client.get(reverse('frm_net_overview'),
                                   {'export': 'excel', 'site': 'Madrid'})
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp['Content-Type'])
        self.assertIn('attachment', resp['Content-Disposition'])
        ws = openpyxl.load_workbook(io.BytesIO(resp.content)).active
        self.assertEqual(ws.max_row, 2)  # header + one alert row
        self.assertEqual(ws.cell(2, 6).value, 'Outdated firmware')
        self.assertEqual(ws.cell(2, 7).value, '1.0 -> 2.0')

    def test_topology_pdf_export(self):
        from unittest.mock import patch
        rows = [{'site': 'Madrid', 'topology': {
            'gateways': [{'name': 'FW', 'model': 'Z', 'online': True, 'clients': 0,
                          'metrics': [{'label': 'CPU', 'value': 10}, {'label': 'Memory', 'value': 40}],
                          'outdated': False}],
            'switches': [{'name': 'SW1', 'model': 'X', 'online': True, 'clients': 5,
                          'metrics': [], 'outdated': True}],
            'aps': []}}]
        with patch('oe_inventory_py_web.status_cache.get_net_overview', return_value=(rows, None)):
            self.client.force_login(self.user)
            resp = self.client.get(reverse('frm_net_overview'),
                                   {'export': 'pdf', 'site': 'Madrid'})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')
        self.assertTrue(resp.content[:5] == b'%PDF-')

    def test_alerts_excel_export_requires_permission(self):
        User = get_user_model()
        other = User.objects.create_user(
            username='no_perm_exp', password='pass12345', net_overview=0, reader=1)
        self.client.force_login(other)
        resp = self.client.get(reverse('frm_net_overview'),
                               {'export': 'excel', 'site': 'Madrid'})
        self.assertEqual(resp.status_code, 302)


class NetAlertsBadgeTests(TestCase):
    """Footer 'Net Alerts' badge endpoint (api_net_alerts)."""

    def setUp(self):
        from django.core.cache import cache
        cache.clear()  # avoid the last-known net_alerts carrying across tests
        User = get_user_model()
        self.user = User.objects.create_user(
            username='na_user', password='pass12345', net_overview=1, reader=0)

    def test_requires_login(self):
        self.assertEqual(self.client.get(reverse('api_net_alerts')).status_code, 302)

    def test_forbidden_without_permission(self):
        User = get_user_model()
        other = User.objects.create_user(
            username='na_noperm', password='pass12345', net_overview=0, reader=0)
        self.client.force_login(other)
        resp = self.client.get(reverse('api_net_alerts'))
        self.assertEqual(resp.status_code, 403)
        self.assertFalse(resp.json()['ok'])

    def test_sums_alerts_across_sites(self):
        from django.test import override_settings
        from unittest.mock import patch
        rows = [{'alerts': 1}, {'alerts': 2}, {'alerts': 0}]
        with override_settings(NEBULA_BASE_URL='https://x', NEBULA_API_KEY='t', NEBULA_ORG_ID='o'):
            with patch('oe_inventory_py_web.nebula.site_overview', return_value=rows):
                self.client.force_login(self.user)
                resp = self.client.get(reverse('api_net_alerts'))
                self.assertEqual(resp.status_code, 200)
                self.assertEqual(resp.json()['alerts'], 3)
                self.assertTrue(resp.json()['ok'])

    def test_not_configured_returns_zero_not_ok(self):
        from django.test import override_settings
        with override_settings(NEBULA_BASE_URL='', NEBULA_API_KEY='', NEBULA_ORG_ID=''):
            self.client.force_login(self.user)
            resp = self.client.get(reverse('api_net_alerts'))
            self.assertEqual(resp.status_code, 200)
            self.assertEqual(resp.json()['alerts'], 0)
            self.assertFalse(resp.json()['ok'])

    def test_api_failure_returns_zero_not_ok(self):
        from django.test import override_settings
        from unittest.mock import patch
        with override_settings(NEBULA_BASE_URL='https://x', NEBULA_API_KEY='t', NEBULA_ORG_ID='o'):
            with patch('oe_inventory_py_web.nebula.site_overview', side_effect=Exception('boom')):
                self.client.force_login(self.user)
                resp = self.client.get(reverse('api_net_alerts'))
                self.assertEqual(resp.status_code, 200)
                self.assertFalse(resp.json()['ok'])

    def test_badge_rendered_for_net_overview_user(self):
        self.client.force_login(self.user)
        resp = self.client.get(reverse('mdi_home'))
        # The div markup is only emitted for users with the permission (the JS
        # that references the id by string is always present, hence id="..").
        self.assertContains(resp, 'id="net-alerts-panel"')
        self.assertContains(resp, 'Net Alerts:')

    def test_badge_hidden_for_user_without_permission(self):
        User = get_user_model()
        other = User.objects.create_user(
            username='na_noperm2', password='pass12345', net_overview=0, reader=1)
        self.client.force_login(other)
        resp = self.client.get(reverse('mdi_home'))
        self.assertNotContains(resp, 'id="net-alerts-panel"')

    def test_api_includes_anydesk_count(self):
        from unittest.mock import patch
        with patch('oe_inventory_py_web.status_cache.get_status',
                   return_value={'net_alerts': 2, 'anydesk_alerts': 3}):
            self.client.force_login(self.user)
            data = self.client.get(reverse('api_net_alerts')).json()
        self.assertEqual(data['alerts'], 2)
        self.assertEqual(data['anydesk_alerts'], 3)
        self.assertTrue(data['anydesk_ok'])

    def test_anydesk_badge_rendered_for_net_overview_user(self):
        self.client.force_login(self.user)
        resp = self.client.get(reverse('mdi_home'))
        self.assertContains(resp, 'id="anydesk-alerts-panel"')
        self.assertContains(resp, 'Remote Machines Alerts:')

    def test_anydesk_badge_hidden_for_user_without_permission(self):
        User = get_user_model()
        other = User.objects.create_user(
            username='na_noperm3', password='pass12345', net_overview=0, reader=1)
        self.client.force_login(other)
        resp = self.client.get(reverse('mdi_home'))
        self.assertNotContains(resp, 'id="anydesk-alerts-panel"')


class StatusCacheTests(TestCase):
    """Background-refreshed footer status cache (status_cache)."""

    def setUp(self):
        from django.core.cache import cache
        cache.clear()

    def test_compute_and_store_caches_values(self):
        from django.test import override_settings
        from unittest.mock import patch
        from oe_inventory_py_web import status_cache
        with override_settings(NEBULA_BASE_URL='x', NEBULA_API_KEY='t'):
            with patch('oe_inventory_py_web.context_processors.pending_counts', return_value=(3, 5)), \
                 patch('oe_inventory_py_web.nebula.site_overview', return_value=[{'alerts': 2}, {'alerts': 4}]):
                data = status_cache.compute_and_store()
        self.assertEqual(data['total_orders'], 3)
        self.assertEqual(data['total_cards'], 5)
        self.assertEqual(data['net_alerts'], 6)

    def test_net_alerts_none_when_not_configured(self):
        from django.test import override_settings
        from unittest.mock import patch
        from oe_inventory_py_web import status_cache
        with override_settings(NEBULA_BASE_URL='', NEBULA_API_KEY=''):
            with patch('oe_inventory_py_web.context_processors.pending_counts', return_value=(0, 0)):
                data = status_cache.compute_and_store()
        self.assertIsNone(data['net_alerts'])

    def test_site_overview_rows_are_cached(self):
        from django.test import override_settings
        from unittest.mock import patch
        from oe_inventory_py_web import status_cache
        rows = [{'site': 'X', 'alerts': 2}]
        with override_settings(NEBULA_BASE_URL='x', NEBULA_API_KEY='t'):
            with patch('oe_inventory_py_web.context_processors.pending_counts', return_value=(0, 0)), \
                 patch('oe_inventory_py_web.nebula.site_overview', return_value=rows):
                status_cache.compute_and_store()
                with override_settings(MDI_STATUS_REFRESH_IN_BACKGROUND=True):
                    got, err = status_cache.get_net_overview(trigger=False)
        self.assertEqual(got, rows)
        self.assertIsNone(err)

    def test_get_net_overview_keeps_rows_and_reports_error_on_failure(self):
        from django.test import override_settings
        from unittest.mock import patch
        from oe_inventory_py_web import status_cache
        with override_settings(NEBULA_BASE_URL='x', NEBULA_API_KEY='t'):
            with patch('oe_inventory_py_web.context_processors.pending_counts', return_value=(0, 0)), \
                 patch('oe_inventory_py_web.nebula.site_overview', side_effect=Exception('boom')):
                rows, err = status_cache.get_net_overview()
        self.assertIsNone(rows)
        self.assertIn('Nebula', err)

    def test_get_status_serves_cache_without_recompute_when_fresh(self):
        import time
        from django.test import override_settings
        from django.core.cache import cache
        from unittest.mock import patch
        from oe_inventory_py_web import status_cache
        cache.set(status_cache.DATA_KEY, {'total_orders': 1, 'total_cards': 2, 'net_alerts': 9}, None)
        cache.set(status_cache.TS_KEY, time.time(), None)
        with override_settings(MDI_STATUS_REFRESH_IN_BACKGROUND=True):
            with patch('oe_inventory_py_web.status_cache._trigger_refresh') as trig:
                data = status_cache.get_status()
        self.assertEqual(data['net_alerts'], 9)
        trig.assert_not_called()

    def test_get_status_triggers_refresh_when_stale(self):
        import time
        from django.test import override_settings
        from django.core.cache import cache
        from unittest.mock import patch
        from oe_inventory_py_web import status_cache
        cache.set(status_cache.DATA_KEY, {'total_orders': 0, 'total_cards': 0, 'net_alerts': 0}, None)
        cache.set(status_cache.TS_KEY, time.time() - 9999, None)  # stale
        with override_settings(MDI_STATUS_REFRESH_IN_BACKGROUND=True, MDI_STATUS_REFRESH_SECONDS=300):
            with patch('oe_inventory_py_web.status_cache._trigger_refresh') as trig:
                status_cache.get_status()
        trig.assert_called_once()


class FooterCountsApiTests(TestCase):
    """Footer live counters endpoint (api_footer_counts)."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(
            username='fc_user', password='pass12345', reader=1)

    def test_requires_login(self):
        self.assertEqual(self.client.get(reverse('api_footer_counts')).status_code, 302)

    def test_returns_pending_counts(self):
        from unittest.mock import patch
        with patch('oe_inventory_py_web.context_processors.pending_counts', return_value=(7, 4)):
            self.client.force_login(self.user)
            resp = self.client.get(reverse('api_footer_counts'))
            self.assertEqual(resp.status_code, 200)
            data = resp.json()
            self.assertEqual(data['total_orders'], 7)
            self.assertEqual(data['total_cards'], 4)
            # The requester is always counted as online.
            self.assertGreaterEqual(data['online_users'], 1)
